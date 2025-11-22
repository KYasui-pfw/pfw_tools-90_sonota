# -*- coding: utf-8 -*-
"""
出荷シール作成システム - Streamlit版（完全版）
作成日：2025-01-22
元のfactcnv.pyの全機能を統合
"""

import streamlit as st
import pandas as pd
import openpyxl
import openpyxl.styles
from glob import glob
import datetime
from io import BytesIO
import os

# ===== ヘルパー関数 =====

def print_excel_make1(output_list, excel_name, pg_folder_path):
    """テンプレート1（ドライビングバー）でExcel生成"""
    list_len = len(output_list)

    template_path = os.path.join(pg_folder_path, 'テンプレート１.xlsx')
    with open(template_path, 'rb') as f:
        template_bytes = f.read()

    wb_io = BytesIO(template_bytes)
    out_wb = openpyxl.load_workbook(wb_io)
    out_wb.active.title = excel_name

    if list_len == 0:
        out_wb.active.cell(1, 1).value = "対象データなし"
    else:
        for i in range(list_len):
            x = i % 2 * 5
            y = i / 2 * 9
            if i % 2 == 1:
                y = (i - i % 2) / 2 * 9

            out_wb.active.cell(int(y+1), x+1).value = output_list[i][0]  # 製造番号
            out_wb.active.cell(int(y+1), x+3).value = output_list[i][1]  # 月次
            out_wb.active.cell(int(y+2), x+1).value = output_list[i][2]  # 国名
            out_wb.active.cell(int(y+2), x+2).value = output_list[i][3]  # 製品名称
            out_wb.active.cell(int(y+3), x+1).value = output_list[i][4]  # 部品補足1
            out_wb.active.cell(int(y+4), x+1).value = output_list[i][5]  # 部品補足2
            out_wb.active.cell(int(y+5), x+1).value = output_list[i][6]  # 部品1名称
            out_wb.active.cell(int(y+5), x+3).value = output_list[i][7]  # 部品1数量
            out_wb.active.cell(int(y+6), x+1).value = output_list[i][8]  # 部品2名称
            out_wb.active.cell(int(y+6), x+3).value = output_list[i][9]  # 部品2数量
            out_wb.active.cell(int(y+7), x+1).value = output_list[i][10]  # 部品3名称
            out_wb.active.cell(int(y+7), x+3).value = output_list[i][11]  # 部品3数量
            out_wb.active.cell(int(y+8), x+1).value = output_list[i][12]  # 部品4名称
            out_wb.active.cell(int(y+8), x+3).value = output_list[i][13]  # 部品4数量

        out_wb.active.print_area = 'A1:I' + str(int(y+8))

    output_buffer = BytesIO()
    out_wb.save(output_buffer)
    output_buffer.seek(0)
    out_wb.close()

    return output_buffer


def print_excel_make2(output_list, excel_name, pg_folder_path):
    """テンプレート2（キャスター、KYK）でExcel生成"""
    list_len = len(output_list)

    template_path = os.path.join(pg_folder_path, 'テンプレート２.xlsx')
    with open(template_path, 'rb') as f:
        template_bytes = f.read()

    wb_io = BytesIO(template_bytes)
    out_wb = openpyxl.load_workbook(wb_io)
    out_wb.active.title = excel_name

    if list_len == 0:
        out_wb.active.cell(1, 1).value = "対象データなし"
    else:
        for i in range(list_len):
            x = i % 2 * 5
            y = i / 2 * 9
            if i % 2 == 1:
                y = (i - i % 2) / 2 * 9

            out_wb.active.cell(int(y+1), x+1).value = output_list[i][0]  # 製造番号
            out_wb.active.cell(int(y+1), x+3).value = output_list[i][1]  # 月次
            out_wb.active.cell(int(y+1), x+4).value = output_list[i][2]  # 分
            out_wb.active.cell(int(y+2), x+1).value = output_list[i][3]  # 製品名称1
            out_wb.active.cell(int(y+2), x+4).value = output_list[i][4]  # チェックボックス
            out_wb.active.cell(int(y+3), x+1).value = output_list[i][5]  # 国名
            out_wb.active.cell(int(y+4), x+1).value = output_list[i][6]  # 空欄1
            out_wb.active.cell(int(y+3), x+2).value = output_list[i][7]  # 補足1
            out_wb.active.cell(int(y+5), x+1).value = output_list[i][8]  # 製品名称2
            out_wb.active.cell(int(y+7), x+1).value = output_list[i][9]  # 空欄2

        out_wb.active.print_area = 'A1:I' + str(int(y+8))

    output_buffer = BytesIO()
    out_wb.save(output_buffer)
    output_buffer.seek(0)
    out_wb.close()

    return output_buffer


def print_excel_make3(output_list, excel_name, pg_folder_path):
    """テンプレート3（スクリュー、電柄Y付きチップ箱、ガードネットプレート）でExcel生成"""
    list_len = len(output_list)

    template_path = os.path.join(pg_folder_path, 'テンプレート３.xlsx')
    with open(template_path, 'rb') as f:
        template_bytes = f.read()

    wb_io = BytesIO(template_bytes)
    out_wb = openpyxl.load_workbook(wb_io)
    out_wb.active.title = excel_name

    if list_len == 0:
        out_wb.active.cell(1, 1).value = "対象データなし"
    else:
        for i in range(list_len):
            x = i % 2 * 5
            y = i / 2 * 12
            if i % 2 == 1:
                y = (i - i % 2) / 2 * 12

            # フラグ確認（部品数が3以下でフラグ=1）
            if output_list[i][20] == 0:
                # セル結合
                if i % 2 == 0:
                    out_wb.active.merge_cells('A'+str(int(y+5))+':B'+str(int(y+5)))
                    out_wb.active.merge_cells('C'+str(int(y+5))+':D'+str(int(y+5)))
                    out_wb.active.merge_cells('A'+str(int(y+6))+':B'+str(int(y+6)))
                    out_wb.active.merge_cells('C'+str(int(y+6))+':D'+str(int(y+6)))
                    out_wb.active.merge_cells('A'+str(int(y+7))+':B'+str(int(y+7)))
                    out_wb.active.merge_cells('C'+str(int(y+7))+':D'+str(int(y+7)))
                    out_wb.active.merge_cells('A'+str(int(y+8))+':B'+str(int(y+8)))
                    out_wb.active.merge_cells('C'+str(int(y+8))+':D'+str(int(y+8)))
                    out_wb.active.merge_cells('A'+str(int(y+9))+':B'+str(int(y+9)))
                    out_wb.active.merge_cells('C'+str(int(y+9))+':D'+str(int(y+9)))
                    out_wb.active.merge_cells('A'+str(int(y+10))+':B'+str(int(y+10)))
                    out_wb.active.merge_cells('C'+str(int(y+10))+':D'+str(int(y+10)))
                    out_wb.active.merge_cells('A'+str(int(y+11))+':B'+str(int(y+11)))
                    out_wb.active.merge_cells('C'+str(int(y+11))+':D'+str(int(y+11)))
                else:
                    out_wb.active.merge_cells('F'+str(int(y+5))+':G'+str(int(y+5)))
                    out_wb.active.merge_cells('H'+str(int(y+5))+':I'+str(int(y+5)))
                    out_wb.active.merge_cells('F'+str(int(y+6))+':G'+str(int(y+6)))
                    out_wb.active.merge_cells('H'+str(int(y+6))+':I'+str(int(y+6)))
                    out_wb.active.merge_cells('F'+str(int(y+7))+':G'+str(int(y+7)))
                    out_wb.active.merge_cells('H'+str(int(y+7))+':I'+str(int(y+7)))
                    out_wb.active.merge_cells('F'+str(int(y+8))+':G'+str(int(y+8)))
                    out_wb.active.merge_cells('H'+str(int(y+8))+':I'+str(int(y+8)))
                    out_wb.active.merge_cells('F'+str(int(y+9))+':G'+str(int(y+9)))
                    out_wb.active.merge_cells('H'+str(int(y+9))+':I'+str(int(y+9)))
                    out_wb.active.merge_cells('F'+str(int(y+10))+':G'+str(int(y+10)))
                    out_wb.active.merge_cells('H'+str(int(y+10))+':I'+str(int(y+10)))
                    out_wb.active.merge_cells('F'+str(int(y+11))+':G'+str(int(y+11)))
                    out_wb.active.merge_cells('H'+str(int(y+11))+':I'+str(int(y+11)))

                out_wb.active.cell(int(y+1), x+1).value = output_list[i][0]
                out_wb.active.cell(int(y+1), x+3).value = output_list[i][1]
                out_wb.active.cell(int(y+2), x+1).value = output_list[i][2]
                out_wb.active.cell(int(y+2), x+2).value = output_list[i][3]
                out_wb.active.cell(int(y+3), x+1).value = output_list[i][4]
                out_wb.active.cell(int(y+4), x+1).value = output_list[i][5]
                out_wb.active.cell(int(y+5), x+1).value = output_list[i][6]
                out_wb.active.cell(int(y+5), x+3).value = output_list[i][7]
                out_wb.active.cell(int(y+6), x+1).value = output_list[i][8]
                out_wb.active.cell(int(y+6), x+3).value = output_list[i][9]
                out_wb.active.cell(int(y+7), x+1).value = output_list[i][10]
                out_wb.active.cell(int(y+7), x+3).value = output_list[i][11]
                out_wb.active.cell(int(y+8), x+1).value = output_list[i][12]
                out_wb.active.cell(int(y+8), x+3).value = output_list[i][13]
                out_wb.active.cell(int(y+9), x+1).value = output_list[i][14]
                out_wb.active.cell(int(y+9), x+3).value = output_list[i][15]
                out_wb.active.cell(int(y+10), x+1).value = output_list[i][16]
                out_wb.active.cell(int(y+10), x+3).value = output_list[i][17]
                out_wb.active.cell(int(y+11), x+1).value = output_list[i][18]
                out_wb.active.cell(int(y+11), x+3).value = output_list[i][19]

            elif output_list[i][20] == 1:
                # セル結合（部品3つ以下）
                if i % 2 == 0:
                    out_wb.active.merge_cells('A'+str(int(y+5))+':B'+str(int(y+6)))
                    out_wb.active.merge_cells('C'+str(int(y+5))+':D'+str(int(y+6)))
                    out_wb.active.merge_cells('A'+str(int(y+7))+':B'+str(int(y+8)))
                    out_wb.active.merge_cells('C'+str(int(y+7))+':D'+str(int(y+8)))
                    out_wb.active.merge_cells('A'+str(int(y+9))+':B'+str(int(y+10)))
                    out_wb.active.merge_cells('C'+str(int(y+9))+':D'+str(int(y+10)))
                else:
                    out_wb.active.merge_cells('F'+str(int(y+5))+':G'+str(int(y+6)))
                    out_wb.active.merge_cells('H'+str(int(y+5))+':I'+str(int(y+6)))
                    out_wb.active.merge_cells('F'+str(int(y+7))+':G'+str(int(y+8)))
                    out_wb.active.merge_cells('H'+str(int(y+7))+':I'+str(int(y+8)))
                    out_wb.active.merge_cells('F'+str(int(y+9))+':G'+str(int(y+10)))
                    out_wb.active.merge_cells('H'+str(int(y+9))+':I'+str(int(y+10)))

                # フォント調整
                out_wb.active.cell(int(y+5), x+1).font = openpyxl.styles.Font(size=22)
                out_wb.active.cell(int(y+5), x+3).font = openpyxl.styles.Font(size=22)
                out_wb.active.cell(int(y+7), x+1).font = openpyxl.styles.Font(size=22)
                out_wb.active.cell(int(y+7), x+3).font = openpyxl.styles.Font(size=22)
                out_wb.active.cell(int(y+9), x+1).font = openpyxl.styles.Font(size=22)
                out_wb.active.cell(int(y+9), x+3).font = openpyxl.styles.Font(size=22)

                out_wb.active.cell(int(y+1), x+1).value = output_list[i][0]
                out_wb.active.cell(int(y+1), x+3).value = output_list[i][1]
                out_wb.active.cell(int(y+2), x+1).value = output_list[i][2]
                out_wb.active.cell(int(y+2), x+2).value = output_list[i][3]
                out_wb.active.cell(int(y+3), x+1).value = output_list[i][4]
                out_wb.active.cell(int(y+4), x+1).value = output_list[i][5]
                out_wb.active.cell(int(y+5), x+1).value = output_list[i][6]
                out_wb.active.cell(int(y+5), x+3).value = output_list[i][7]
                out_wb.active.cell(int(y+7), x+1).value = output_list[i][8]
                out_wb.active.cell(int(y+7), x+3).value = output_list[i][9]
                out_wb.active.cell(int(y+9), x+1).value = output_list[i][10]
                out_wb.active.cell(int(y+9), x+3).value = output_list[i][11]

        out_wb.active.print_area = 'A1:I' + str(int(y+12))

    output_buffer = BytesIO()
    out_wb.save(output_buffer)
    output_buffer.seek(0)
    out_wb.close()

    return output_buffer


def check_excel_make(output_list, excel_name, pg_folder_path):
    """ODチェック表のExcel生成"""
    list_len = len(output_list)

    template_path = os.path.join(pg_folder_path, 'ODチェック表.xlsx')
    with open(template_path, 'rb') as f:
        template_bytes = f.read()

    wb_io = BytesIO(template_bytes)
    out_wb = openpyxl.load_workbook(wb_io)
    out_wb.active.title = excel_name

    if list_len == 0:
        out_wb.active.cell(1, 1).value = "対象データなし"
    else:
        for i in range(list_len):
            out_wb.active.cell(i+5, 1).value = output_list[i][0]  # 月次
            out_wb.active.cell(i+5, 2).value = output_list[i][1]  # 製造番号
            out_wb.active.cell(i+5, 3).value = output_list[i][2]  # 製品名称

        out_wb.active.print_area = 'A1:S' + str(int(4+list_len))

    output_buffer = BytesIO()
    out_wb.save(output_buffer)
    output_buffer.seek(0)
    out_wb.close()

    return output_buffer


def create_error_log(errors_list):
    """エラーログExcel生成"""
    if not errors_list:
        return None

    dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))

    err_wb = openpyxl.Workbook()

    # 列幅設定
    err_wb.active.column_dimensions["A"].width = 5
    err_wb.active.column_dimensions["B"].width = 15
    err_wb.active.column_dimensions["C"].width = 40
    err_wb.active.column_dimensions["D"].width = 100

    # ヘッダー
    err_wb.active["A1"] = "部品数表示超過一覧"
    err_wb.active['A1'].font = openpyxl.styles.Font(size=24)
    err_wb.active["A2"] = "スクリューは7種類、ドライビングバーは4種類がシールへの表示の上限です。"
    err_wb.active["A3"] = "以下のリストにシールに入りきらなかった部品があったものを記載しています。"
    err_wb.active["A4"] = "当リストが表示された場合は、お手数ですが対象のシールエクセルに手作業で追加いただきますようお願いいたします。"

    err_wb.active["A6"] = "No."
    err_wb.active["B6"] = "処理時間"
    err_wb.active["C6"] = "内容"
    err_wb.active["D6"] = "対象の製造番号と超過した部品の一覧"

    # 固定
    err_wb.active.freeze_panes = "A7"

    # 横向き
    err_wb.active.page_setup.orientation = "landscape"

    # エラーデータ追加
    for idx, error in enumerate(errors_list, start=1):
        err_wb.active["A"+str(err_wb.active.max_row + 1)] = idx
        err_wb.active["B"+str(err_wb.active.max_row)] = error['time']
        err_wb.active["C"+str(err_wb.active.max_row)] = error['err1']
        err_wb.active["D"+str(err_wb.active.max_row)] = error['err2']

    output_buffer = BytesIO()
    err_wb.save(output_buffer)
    output_buffer.seek(0)
    err_wb.close()

    return output_buffer


# ===== Streamlitメインアプリケーション =====

st.set_page_config(
    page_title="出荷シール作成システム",
    page_icon="🏷️",
    layout="wide"
)

st.markdown("""
<style>
    .main-header {
        background-color: #f0f2f6;
        padding: 1.5rem;
        border-radius: 0.5rem;
        margin-bottom: 2rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">', unsafe_allow_html=True)
st.title("🏷️ 出荷シール作成システム（完全版）")
st.markdown("</div>", unsafe_allow_html=True)

with st.expander("📖 使用方法"):
    st.markdown("""
    ### 処理手順
    1. **処理対象年月** を入力（YYYYMM形式、例：202501）
    2. **処理開始** ボタンをクリック
    3. 生成された7種類のシールExcelファイルをダウンロード

    ### 生成されるシール（全7種類対応）
    - ✅ ドライビングバー
    - ✅ スクリュー
    - ✅ 電柄Y付きチップ箱
    - ✅ KYK
    - ✅ キャスター
    - ✅ ODチェック表
    - ✅ ガードネットプレート

    ### 入力ファイル自動取得元
    - ネットワークドライブ: `\\\\esrv06\\帳票開発\\{年月}\\`
    - 生産機情報.xls(x)
    - 構成フォルダ内の全ファイル
    """)

# 環境判定
if os.path.exists("/app/network"):
    INPUT_BASE_PATH = "/app/network/帳票開発"
    PG_FOLDER_PATH = "/app/pg"
else:
    INPUT_BASE_PATH = r"\\esrv06\帳票開発"
    PG_FOLDER_PATH = "pg"

col1, col2 = st.columns([2, 8])

with col1:
    nengetsu = st.text_input(
        "処理対象年月",
        placeholder="202501",
        help="YYYYMM形式で入力してください"
    )

with col2:
    st.write("")
    st.write("")
    process_button = st.button("🚀 処理開始", type="primary")

if "processed" not in st.session_state:
    st.session_state.processed = False
    st.session_state.results = {}
    st.session_state.errors = []

if process_button:
    if not nengetsu or len(nengetsu) != 6 or not nengetsu.isdigit():
        st.error("❌ 年月はYYYYMM形式の6桁数字で入力してください")
    else:
        st.session_state.processed = False
        st.session_state.results = {}
        st.session_state.errors = []

        progress_bar = st.progress(0, text="データ読み込み中...")

        try:
            # データ読み込み
            target_path = os.path.join(INPUT_BASE_PATH, nengetsu)

            # 生産機情報ファイル
            seisanki_files = glob(os.path.join(target_path, "生産機情報.xls*"))
            if not seisanki_files:
                st.error(f"❌ 生産機情報ファイルが見つかりません: {target_path}")
                st.stop()

            # 構成ファイル
            kousei_files = glob(os.path.join(target_path, "構成", "*.xls*"))
            if not kousei_files:
                st.error(f"❌ 構成ファイルが見つかりません: {os.path.join(target_path, '構成')}")
                st.stop()

            progress_bar.progress(10, text=f"ファイル読み込み中... ({len(kousei_files)}件)")

            # 生産機情報読み込み
            df1 = pd.read_excel(seisanki_files[0], sheet_name=0, usecols="A:T")
            df1 = df1.iloc[:, [0, 3, 17, 18]]

            # 抽出条件読み込み
            in_wb = openpyxl.load_workbook(os.path.join(PG_FOLDER_PATH, "抽出条件.xlsx"))
            in_sheet = in_wb.active

            # 結果格納用のデータフレーム
            result_df1 = pd.DataFrame()  # ドライビングバー
            result_df2 = pd.DataFrame()  # スクリュー
            result_df3 = pd.DataFrame()  # 電柄Y付きチップ箱
            result_df4 = pd.DataFrame()  # KYK
            result_df5 = pd.DataFrame()  # キャスター
            result_df6 = pd.DataFrame()  # ODチェック表
            result_df7 = pd.DataFrame()  # ガードネットプレート

            # 構成ファイル処理
            total_files = len(kousei_files)
            for file_idx, book1 in enumerate(kousei_files):
                progress_pct = 10 + int((file_idx / total_files) * 40)
                progress_bar.progress(progress_pct, text=f"データ処理中... ({file_idx+1}/{total_files})")

                # ファイル名取得
                fname = os.path.basename(book1)[-16:-8]

                # 構成ファイル読み込み
                df2 = pd.read_excel(book1, sheet_name=0, usecols="A:M")
                df2 = df2.iloc[:, [1, 2, 3, 9, 0]]
                df2.dropna(axis="index", how='all', inplace=True)
                df2.loc[:, "製番"] = fname

                # 結合
                df3 = pd.merge(df1, df2)
                df3.columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

                mask = df3['I'] > df3['I'].shift(1)
                df3.loc[mask, 'J'] = df3['F']
                df3.loc[~mask, 'J'] = 0
                df3['C'] = df3['C'].astype(str)

                tmp_list = ["階層リスト："]
                for idx in range(df3.shape[0]):
                    comp = df3.iloc[idx-1, -2]-df3.iloc[idx, -2]

                    if comp < 0:
                        tmp_list.append(str(df3.iloc[idx-1, -5]) + "＋" + str(df3.iloc[idx-1, -6]))
                    elif comp > 0 and idx != 0:
                        for idx2 in range(int(comp)):
                            del tmp_list[-1]

                    df3.iloc[idx, -1] = str(tmp_list) + "," + str(df3.iloc[idx, -5]) + "＋" + str(df3.iloc[idx, -6])
                    df3.iloc[idx, 2] = "("+str(int(str(df3.iloc[idx, 2])[-3:-1]))+"-"+str(df3.iloc[idx, 2])[-1]+")"

                # ===== 1. ドライビングバー =====
                df4 = df3.copy()
                for row in in_sheet.iter_rows(min_row=6):
                    in_cells = [cell.value for cell in row]
                    if in_cells[0] == "ドライビングバー":
                        if in_cells[1] is not None:
                            df4 = df4[df4['B'].str.contains(in_cells[1], na=False)]
                        if in_cells[2] is not None:
                            df4 = df4[~df4['B'].str.contains(in_cells[2], na=False)]
                        if in_cells[3] is not None:
                            df4 = df4[df4['J'].str.contains(in_cells[3], na=False)]
                        if in_cells[4] is not None:
                            df4 = df4[~df4['J'].str.contains(in_cells[4], na=False)]

                df4 = df4.loc[:, ['A', 'C', 'D', 'B', 'G', 'H']]
                df4 = df4.set_axis(['製造番号', '月次', '国名', '製品名称1', '部品1名称', '部品1数量'], axis=1)
                df4 = df4.groupby(['製造番号', '月次', '国名', '製品名称1', '部品1名称'])['部品1数量'].sum().reset_index()

                df4['部品補足1'] = ""
                df4['部品補足2'] = "Driving Bar Fixed Screws"
                df4['部品2名称'] = ""
                df4['部品2数量'] = ""
                df4['部品3名称'] = ""
                df4['部品3数量'] = ""
                df4['部品4名称'] = ""
                df4['部品4数量'] = ""
                df4 = df4.reindex(['製造番号', '月次', '国名', '製品名称1', '部品補足1', '部品補足2', '部品1名称',
                                  '部品1数量', '部品2名称', '部品2数量', '部品3名称', '部品3数量', '部品4名称', '部品4数量'], axis=1)

                mask = df4['製造番号'] == df4['製造番号'].shift(-1)
                df4.loc[mask, '部品2名称'] = df4['部品1名称'].shift(-1)
                df4.loc[mask, '部品2数量'] = df4['部品1数量'].shift(-1)
                mask = df4['製造番号'] == df4['製造番号'].shift(-2)
                df4.loc[mask, '部品3名称'] = df4['部品1名称'].shift(-2)
                df4.loc[mask, '部品3数量'] = df4['部品1数量'].shift(-2)
                mask = df4['製造番号'] == df4['製造番号'].shift(-3)
                df4.loc[mask, '部品4名称'] = df4['部品1名称'].shift(-3)
                df4.loc[mask, '部品4数量'] = df4['部品1数量'].shift(-3)

                # エラーチェック（5個以上）
                mask = df4['製造番号'] == df4['製造番号'].shift(-4)
                if df4.loc[mask, :].shape[0] > 0:
                    dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
                    st.session_state.errors.append({
                        'time': dt_now.strftime('%H時%M分%S秒'),
                        'err1': "【ドライビングバー】5つ以上の部品が存在します",
                        'err2': f"{df4.iloc[0, 0]}の部品一覧：（詳細省略）"
                    })

                mask = df4['製造番号'] == df4['製造番号'].shift(1)
                df4 = df4.loc[~mask, :]
                result_df1 = pd.concat([result_df1, df4])

                # ===== 2. スクリュー =====
                df5 = df3.copy()
                caster_list = []

                for row in in_sheet.iter_rows(min_row=6):
                    in_cells = [cell.value for cell in row]
                    if in_cells[0] == "スクリュー":
                        if in_cells[1] is not None:
                            df5 = df5[df5['B'].str.contains(in_cells[1], na=False)]
                        if in_cells[2] is not None:
                            df5 = df5[~df5['B'].str.contains(in_cells[2], na=False)]
                        if in_cells[3] is not None:
                            df5 = df5[df5['J'].str.contains(in_cells[3], na=False)]
                        if in_cells[4] is not None:
                            df5 = df5[~df5['J'].str.contains(in_cells[4], na=False)]
                        if in_cells[6] is not None:
                            caster_list.append(in_cells[6])

                df5['部品補足1'] = "なし"
                df5['部品補足2'] = ""
                if len(caster_list) > 0:
                    mask = df5['J'].str.contains(caster_list[0], na=False)
                    df5.loc[mask, '部品補足1'] = 'GUARDNET CASTERS'
                    mask = df5['J'].str.contains(caster_list[0], na=False) & df5['J'].str.contains('(?=.*ASSY. GUARD NET)', na=False)
                    df5.loc[mask, '部品補足2'] = 'Step Fixed Screws'
                if len(caster_list) > 1:
                    mask = df5['J'].str.contains(caster_list[1], na=False)
                    df5.loc[mask, '部品補足1'] = 'STEP CASTERS'
                if len(caster_list) > 2:
                    mask = df5['J'].str.contains(caster_list[2], na=False)
                    df5.loc[mask, '部品補足1'] = 'IN WHEEL MOTOR CASTERS'

                for idx in range(len(caster_list)):
                    if idx > 2:
                        mask = df5['J'].str.contains(caster_list[idx], na=False)
                        i = caster_list[idx].find('(')
                        df5.loc[mask, '部品補足1'] = caster_list[idx][i+1:-2]

                mask = df5['部品補足1'] == 'なし'
                df5 = df5.loc[~mask, :]

                df5 = df5.loc[:, ['A', 'C', 'D', 'B', '部品補足1', '部品補足2', 'G', 'H']]
                df5 = df5.set_axis(['製造番号', '月次', '国名', '製品名称1', '部品補足1', '部品補足2', '部品1名称', '部品1数量'], axis=1)
                df5 = df5.sort_values(['月次', '製造番号', '部品補足1'])
                df5 = df5.groupby(['製造番号', '月次', '国名', '製品名称1', '部品補足1', '部品補足2', '部品1名称'])['部品1数量'].sum().reset_index()

                df5['部品2名称'] = ""
                df5['部品2数量'] = ""
                df5['部品3名称'] = ""
                df5['部品3数量'] = ""
                df5['部品4名称'] = ""
                df5['部品4数量'] = ""
                df5['部品5名称'] = ""
                df5['部品5数量'] = ""
                df5['部品6名称'] = ""
                df5['部品6数量'] = ""
                df5['部品7名称'] = ""
                df5['部品7数量'] = ""
                df5['フラグ'] = 0
                df5 = df5.reindex(['製造番号', '月次', '国名', '製品名称1', '部品補足1', '部品補足2', '部品1名称', '部品1数量', '部品2名称', '部品2数量', '部品3名称', '部品3数量',
                                  '部品4名称', '部品4数量', '部品5名称', '部品5数量', '部品6名称', '部品6数量', '部品7名称', '部品7数量', 'フラグ'], axis=1)

                df5 = df5.sort_values(['月次', '製造番号', '部品補足1'])
                mask = df5['部品補足1'] == df5['部品補足1'].shift(-1)
                df5.loc[mask, '部品2名称'] = df5['部品1名称'].shift(-1)
                df5.loc[mask, '部品2数量'] = df5['部品1数量'].shift(-1)
                mask = df5['部品補足1'] == df5['部品補足1'].shift(-2)
                df5.loc[mask, '部品3名称'] = df5['部品1名称'].shift(-2)
                df5.loc[mask, '部品3数量'] = df5['部品1数量'].shift(-2)
                mask = df5['部品補足1'] == df5['部品補足1'].shift(-3)
                df5.loc[mask, '部品4名称'] = df5['部品1名称'].shift(-3)
                df5.loc[mask, '部品4数量'] = df5['部品1数量'].shift(-3)
                df5.loc[~mask, 'フラグ'] = 1
                mask = df5['部品補足1'] == df5['部品補足1'].shift(-4)
                df5.loc[mask, '部品5名称'] = df5['部品1名称'].shift(-4)
                df5.loc[mask, '部品5数量'] = df5['部品1数量'].shift(-4)
                mask = df5['部品補足1'] == df5['部品補足1'].shift(-5)
                df5.loc[mask, '部品6名称'] = df5['部品1名称'].shift(-5)
                df5.loc[mask, '部品6数量'] = df5['部品1数量'].shift(-5)
                mask = df5['部品補足1'] == df5['部品補足1'].shift(-6)
                df5.loc[mask, '部品7名称'] = df5['部品1名称'].shift(-6)
                df5.loc[mask, '部品7数量'] = df5['部品1数量'].shift(-6)

                # エラーチェック（8個以上）
                mask = df5['部品補足1'] == df5['部品補足1'].shift(-7)
                if df5.loc[mask, :].shape[0] > 0:
                    dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
                    st.session_state.errors.append({
                        'time': dt_now.strftime('%H時%M分%S秒'),
                        'err1': "【スクリュー】8つ以上の部品が存在します",
                        'err2': f"{df5.iloc[0, 0]}の部品一覧：（詳細省略）"
                    })

                mask = df5['部品補足1'] == df5['部品補足1'].shift(1)
                df5 = df5.loc[~mask, :]
                if df5.shape[0] == 1:
                    df5.loc[:, '部品補足2'] = ""
                result_df2 = pd.concat([result_df2, df5])

                # ===== 3. 電柄Y付きチップ箱 =====
                df6 = df3.copy()

                for row in in_sheet.iter_rows(min_row=6):
                    in_cells = [cell.value for cell in row]
                    if in_cells[0] == "電柄Y付チップ箱":
                        if in_cells[1] is not None:
                            df6 = df6[df6['B'].str.contains(in_cells[1], na=False)]
                        if in_cells[2] is not None:
                            df6 = df6[~df6['B'].str.contains(in_cells[2], na=False)]
                        if in_cells[3] is not None:
                            df6 = df6[df6['J'].str.contains(in_cells[3], na=False)]
                        if in_cells[4] is not None:
                            df6 = df6[~df6['J'].str.contains(in_cells[4], na=False)]

                df6 = df6.loc[:, ['A', 'C', 'D', 'B']]
                df6 = df6.set_axis(['製造番号', '月次', '国名', '製品名称1'], axis=1)
                df6 = df6.drop_duplicates()

                df6['部品補足1'] = ""
                df6['部品補足2'] = ""
                df6['部品1名称'] = ""
                df6['部品1数量'] = ""
                df6['部品2名称'] = ""
                df6['部品2数量'] = ""
                df6['部品3名称'] = ""
                df6['部品3数量'] = ""
                df6['部品4名称'] = ""
                df6['部品4数量'] = ""
                df6['部品5名称'] = ""
                df6['部品5数量'] = ""
                df6['部品6名称'] = ""
                df6['部品6数量'] = ""
                df6['部品7名称'] = ""
                df6['部品7数量'] = ""
                df6['フラグ'] = 1
                df6 = df6.reindex(['製造番号', '月次', '国名', '製品名称1', '部品補足1', '部品補足2', '部品1名称', '部品1数量', '部品2名称', '部品2数量', '部品3名称', '部品3数量',
                                  '部品4名称', '部品4数量', '部品5名称', '部品5数量', '部品6名称', '部品6数量', '部品7名称', '部品7数量', 'フラグ'], axis=1)

                # ULM有り
                mask = df6['製品名称1'].str.contains('LEC|SEC', na=False)
                df6.loc[mask, '部品1名称'] = "U.L.M. (USB)"
                df6.loc[mask, '部品1数量'] = 1

                # NEEDLE有り LEC SEC
                mask = df6['製品名称1'].str.contains('LEC|SEC', na=False) & df6['製品名称1'].str.contains('Y', na=False)
                df6.loc[mask, '部品2名称'] = "NEEDLE"
                df6.loc[mask, '部品2数量'] = 1

                # NEEDLE有り LEC SEC 以外
                mask = ~df6['製品名称1'].str.contains('LEC|SEC', na=False) & df6['製品名称1'].str.contains('Y', na=False)
                df6.loc[mask, '部品1名称'] = "NEEDLE"
                df6.loc[mask, '部品1数量'] = 1

                mask = df6['部品1名称'] == ""
                df6 = df6.loc[~mask, :]
                result_df3 = pd.concat([result_df3, df6])

                # ===== 4. KYK =====
                df7 = df3.copy()

                for row in in_sheet.iter_rows(min_row=6):
                    in_cells = [cell.value for cell in row]
                    if in_cells[0] == "KYK":
                        if in_cells[1] is not None:
                            df7 = df7[df7['B'].str.contains(in_cells[1], na=False)]
                        if in_cells[2] is not None:
                            df7 = df7[~df7['B'].str.contains(in_cells[2], na=False)]
                        if in_cells[3] is not None:
                            df7 = df7[df7['J'].str.contains(in_cells[3], na=False)]
                        if in_cells[4] is not None:
                            df7 = df7[~df7['J'].str.contains(in_cells[4], na=False)]

                df7 = df7.loc[:, ['A', 'C', 'D', 'B']]
                df7 = df7.set_axis(['製造番号', '月次', '国名', '製品名称2'], axis=1)
                df7 = df7.drop_duplicates()
                df7 = pd.concat([df7, df7], axis=0)  # 2枚印刷

                df7['分'] = ""
                df7['製品名称1'] = ""
                df7['チェックボックス'] = ""
                df7['空欄1'] = ""
                df7['補足1'] = ""
                df7['空欄2'] = ""
                df7 = df7.reindex(['製造番号', '月次', '分', '製品名称1', 'チェックボックス', '国名', '空欄1', '補足1', '製品名称2', '空欄2'], axis=1)
                result_df4 = pd.concat([result_df4, df7])

                # ===== 5. キャスター =====
                df8 = df3.copy()
                double_list = []

                for row in in_sheet.iter_rows(min_row=6):
                    in_cells = [cell.value for cell in row]
                    if in_cells[0] == "キャスター":
                        if in_cells[1] is not None:
                            df8 = df8[df8['B'].str.contains(in_cells[1], na=False)]
                        if in_cells[2] is not None:
                            df8 = df8[~df8['B'].str.contains(in_cells[2], na=False)]
                        if in_cells[3] is not None:
                            df8 = df8[df8['J'].str.contains(in_cells[3], na=False)]
                        if in_cells[4] is not None:
                            df8 = df8[~df8['J'].str.contains(in_cells[4], na=False)]
                        if in_cells[5] is not None:
                            double_list.append(in_cells[5])

                df8['分'] = ""
                df8['製品名称1'] = ""

                for list_item in double_list:
                    mask = df8['J'].str.contains(list_item, na=False)
                    df8.loc[mask, '分'] = '2／2'
                    df8.loc[mask, '製品名称1'] = 'STEP CASTERS'
                    df8.loc[~mask, '分'] = '1／1'
                    df8.loc[~mask, '製品名称1'] = 'GUARDNET CASTERS'

                df8 = df8.loc[:, ['A', 'C', '分', '製品名称1', 'D', 'B']]
                df8 = df8.set_axis(['製造番号', '月次', '分', '製品名称1', '国名', '製品名称2'], axis=1)
                df8 = df8.drop_duplicates()

                mask = df8['製品名称1'].shift(-1) == 'STEP CASTERS'
                df8.loc[mask, '分'] = '1／2'

                df8['チェックボックス'] = "□"
                df8['空欄1'] = ""
                df8['補足1'] = "SCREW・WASHER"
                df8['空欄2'] = ""
                df8 = df8.reindex(['製造番号', '月次', '分', '製品名称1', 'チェックボックス', '国名', '空欄1', '補足1', '製品名称2', '空欄2'], axis=1)
                result_df5 = pd.concat([result_df5, df8])

                # ===== 6. ODチェック表 =====
                df9 = df3.copy()

                for row in in_sheet.iter_rows(min_row=6):
                    in_cells = [cell.value for cell in row]
                    if in_cells[0] == "ODチェック":
                        if in_cells[1] is not None:
                            df9 = df9[df9['B'].str.contains(in_cells[1], na=False)]
                        if in_cells[2] is not None:
                            df9 = df9[~df9['B'].str.contains(in_cells[2], na=False)]
                        if in_cells[3] is not None:
                            df9 = df9[df9['J'].str.contains(in_cells[3], na=False)]
                        if in_cells[4] is not None:
                            df9 = df9[~df9['J'].str.contains(in_cells[4], na=False)]

                df9 = df9.loc[:, ['C', 'A', 'B']]
                df9 = df9.set_axis(['月次', '製造番号', '製品名称'], axis=1)
                df9 = df9.drop_duplicates()

                for idx in range(df9.shape[0]):
                    df9.iloc[idx, 0] = str(int(str(df9.iloc[idx, 0])[-2:-1]))+"次"
                result_df6 = pd.concat([result_df6, df9])

                # ===== 7. ガードネットプレート =====
                df10 = df3.copy()

                for row in in_sheet.iter_rows(min_row=6):
                    in_cells = [cell.value for cell in row]
                    if in_cells[0] == "ガードネットプレート":
                        if in_cells[1] is not None:
                            df10 = df10[df10['B'].str.contains(in_cells[1], na=False)]
                        if in_cells[2] is not None:
                            df10 = df10[~df10['B'].str.contains(in_cells[2], na=False)]
                        if in_cells[3] is not None:
                            df10 = df10[df10['J'].str.contains(in_cells[3], na=False)]
                        if in_cells[4] is not None:
                            df10 = df10[~df10['J'].str.contains(in_cells[4], na=False)]

                df10['部品補足1'] = "GUARD NET PLATE"
                df10['部品補足2'] = ""

                mask = df10['部品補足1'] == 'なし'
                df10 = df10.loc[~mask, :]

                df10 = df10.loc[:, ['A', 'C', 'D', 'B', '部品補足1', '部品補足2', 'G', 'H']]
                df10 = df10.set_axis(['製造番号', '月次', '国名', '製品名称1', '部品補足1', '部品補足2', '部品1名称', '部品1数量'], axis=1)
                df10 = df10.sort_values(['月次', '製造番号', '部品補足1'])
                df10 = df10.groupby(['製造番号', '月次', '国名', '製品名称1', '部品補足1', '部品補足2', '部品1名称'])['部品1数量'].sum().reset_index()

                df10['部品2名称'] = ""
                df10['部品2数量'] = ""
                df10['部品3名称'] = ""
                df10['部品3数量'] = ""
                df10['部品4名称'] = ""
                df10['部品4数量'] = ""
                df10['部品5名称'] = ""
                df10['部品5数量'] = ""
                df10['部品6名称'] = ""
                df10['部品6数量'] = ""
                df10['部品7名称'] = ""
                df10['部品7数量'] = ""
                df10['フラグ'] = 0
                df10 = df10.reindex(['製造番号', '月次', '国名', '製品名称1', '部品補足1', '部品補足2', '部品1名称', '部品1数量', '部品2名称', '部品2数量', '部品3名称', '部品3数量',
                                    '部品4名称', '部品4数量', '部品5名称', '部品5数量', '部品6名称', '部品6数量', '部品7名称', '部品7数量', 'フラグ'], axis=1)

                df10 = df10.sort_values(['月次', '製造番号', '部品補足1'])
                mask = df10['部品補足1'] == df10['部品補足1'].shift(-1)
                df10.loc[mask, '部品2名称'] = df10['部品1名称'].shift(-1)
                df10.loc[mask, '部品2数量'] = df10['部品1数量'].shift(-1)
                mask = df10['部品補足1'] == df10['部品補足1'].shift(-2)
                df10.loc[mask, '部品3名称'] = df10['部品1名称'].shift(-2)
                df10.loc[mask, '部品3数量'] = df10['部品1数量'].shift(-2)
                mask = df10['部品補足1'] == df10['部品補足1'].shift(-3)
                df10.loc[mask, '部品4名称'] = df10['部品1名称'].shift(-3)
                df10.loc[mask, '部品4数量'] = df10['部品1数量'].shift(-3)
                df10.loc[~mask, 'フラグ'] = 1
                mask = df10['部品補足1'] == df10['部品補足1'].shift(-4)
                df10.loc[mask, '部品5名称'] = df10['部品1名称'].shift(-4)
                df10.loc[mask, '部品5数量'] = df10['部品1数量'].shift(-4)
                mask = df10['部品補足1'] == df10['部品補足1'].shift(-5)
                df10.loc[mask, '部品6名称'] = df10['部品1名称'].shift(-5)
                df10.loc[mask, '部品6数量'] = df10['部品1数量'].shift(-5)
                mask = df10['部品補足1'] == df10['部品補足1'].shift(-6)
                df10.loc[mask, '部品7名称'] = df10['部品1名称'].shift(-6)
                df10.loc[mask, '部品7数量'] = df10['部品1数量'].shift(-6)

                # エラーチェック（8個以上）
                mask = df10['部品補足1'] == df10['部品補足1'].shift(-7)
                if df10.loc[mask, :].shape[0] > 0:
                    dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
                    st.session_state.errors.append({
                        'time': dt_now.strftime('%H時%M分%S秒'),
                        'err1': "【ガードネットプレート】8つ以上の部品が存在します",
                        'err2': f"{df10.iloc[0, 0]}の部品一覧：（詳細省略）"
                    })

                mask = df10['部品補足1'] == df10['部品補足1'].shift(1)
                df10 = df10.loc[~mask, :]
                if df10.shape[0] == 1:
                    df10.loc[:, '部品補足2'] = ""
                result_df7 = pd.concat([result_df7, df10])

            in_wb.close()

            # Excel生成
            progress_bar.progress(60, text="Excel生成中... (1/7)")

            result_df1 = result_df1.sort_values(['月次', '製造番号'])
            result_df2 = result_df2.sort_values(['月次', '製造番号'])
            result_df3 = result_df3.sort_values(['月次', '製造番号'])
            result_df4 = result_df4.sort_values(['月次', '製造番号'])
            result_df5 = result_df5.sort_values(['月次', '製造番号'])
            result_df6 = result_df6.sort_values(['月次', '製造番号'])
            result_df7 = result_df7.sort_values(['月次', '製造番号'])

            dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))

            # 1. ドライビングバー
            progress_bar.progress(65, text="Excel生成中... (1/7) ドライビングバー")
            list_result_df1 = result_df1.to_numpy().tolist()
            buffer1 = print_excel_make1(list_result_df1, "シール_ドライビングバー", PG_FOLDER_PATH)
            st.session_state.results['ドライビングバー'] = {
                'buffer': buffer1,
                'filename': f"シール_ドライビングバー_{dt_now.strftime('%Y%m%d%H%M%S')}.xlsx",
                'count': len(list_result_df1)
            }

            # 2. スクリュー
            progress_bar.progress(70, text="Excel生成中... (2/7) スクリュー")
            list_result_df2 = result_df2.to_numpy().tolist()
            buffer2 = print_excel_make3(list_result_df2, "シール_スクリュー", PG_FOLDER_PATH)
            st.session_state.results['スクリュー'] = {
                'buffer': buffer2,
                'filename': f"シール_スクリュー_{dt_now.strftime('%Y%m%d%H%M%S')}.xlsx",
                'count': len(list_result_df2)
            }

            # 3. 電柄Y付きチップ箱
            progress_bar.progress(75, text="Excel生成中... (3/7) 電柄Y付きチップ箱")
            list_result_df3 = result_df3.to_numpy().tolist()
            buffer3 = print_excel_make3(list_result_df3, "シール_電柄Y付きチップ箱", PG_FOLDER_PATH)
            st.session_state.results['電柄Y付きチップ箱'] = {
                'buffer': buffer3,
                'filename': f"シール_電柄Y付きチップ箱_{dt_now.strftime('%Y%m%d%H%M%S')}.xlsx",
                'count': len(list_result_df3)
            }

            # 4. KYK
            progress_bar.progress(80, text="Excel生成中... (4/7) KYK")
            list_result_df4 = result_df4.to_numpy().tolist()
            buffer4 = print_excel_make2(list_result_df4, "シール_KYK", PG_FOLDER_PATH)
            st.session_state.results['KYK'] = {
                'buffer': buffer4,
                'filename': f"シール_KYK_{dt_now.strftime('%Y%m%d%H%M%S')}.xlsx",
                'count': len(list_result_df4)
            }

            # 5. キャスター
            progress_bar.progress(85, text="Excel生成中... (5/7) キャスター")
            list_result_df5 = result_df5.to_numpy().tolist()
            buffer5 = print_excel_make2(list_result_df5, "シール_キャスター", PG_FOLDER_PATH)
            st.session_state.results['キャスター'] = {
                'buffer': buffer5,
                'filename': f"シール_キャスター_{dt_now.strftime('%Y%m%d%H%M%S')}.xlsx",
                'count': len(list_result_df5)
            }

            # 6. ODチェック表
            progress_bar.progress(90, text="Excel生成中... (6/7) ODチェック表")
            list_result_df6 = result_df6.to_numpy().tolist()
            buffer6 = check_excel_make(list_result_df6, "ODチェック表", PG_FOLDER_PATH)
            st.session_state.results['ODチェック表'] = {
                'buffer': buffer6,
                'filename': f"ODチェック表_{dt_now.strftime('%Y%m%d%H%M%S')}.xlsx",
                'count': len(list_result_df6)
            }

            # 7. ガードネットプレート
            progress_bar.progress(95, text="Excel生成中... (7/7) ガードネットプレート")
            list_result_df7 = result_df7.to_numpy().tolist()
            buffer7 = print_excel_make3(list_result_df7, "シール_ガードネットプレート", PG_FOLDER_PATH)
            st.session_state.results['ガードネットプレート'] = {
                'buffer': buffer7,
                'filename': f"シール_ガードネットプレート_{dt_now.strftime('%Y%m%d%H%M%S')}.xlsx",
                'count': len(list_result_df7)
            }

            progress_bar.progress(100, text="完了!")
            st.session_state.processed = True
            st.success(f"✅ 処理完了！ 全7種類のシールを生成しました（エラー: {len(st.session_state.errors)}件）")

        except Exception as e:
            st.error(f"❌ エラーが発生しました")
            st.exception(e)

if st.session_state.processed and st.session_state.results:
    st.markdown("---")
    st.subheader("📥 ダウンロード")

    cols = st.columns(3)
    for idx, (name, data) in enumerate(st.session_state.results.items()):
        with cols[idx % 3]:
            st.download_button(
                label=f"📄 {name} ({data['count']}件)",
                data=data['buffer'].getvalue(),
                file_name=data['filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    # エラーログがあれば表示
    if st.session_state.errors:
        st.warning(f"⚠️ {len(st.session_state.errors)}件の部品数表示超過があります")
        err_buffer = create_error_log(st.session_state.errors)
        if err_buffer:
            dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
            st.download_button(
                label="📋 部品数表示超過一覧をダウンロード",
                data=err_buffer.getvalue(),
                file_name=f"部品数表示超過一覧_{dt_now.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

st.markdown("---")
st.caption("出荷シール作成システム v1.0 (Streamlit完全版)")
