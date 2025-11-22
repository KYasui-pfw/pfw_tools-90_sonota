################################################################
#　出荷シール作成
#　作成日：20240605 
#　変更：20240611 スクリューの部品数上限を6から7に変更
################################################################

import sys
from glob import glob
import openpyxl
import pandas
import datetime
import shutil

# Pythonのインストールに加えて
# 以下のインストールも行っておく
# pip install openpyxl
# pip install datetime
# pip install pandas
# pip install shutil

################################################################
###  エクセル処理は別関数で定義（※pythonなので上部に記載します) ###
################################################################
###  テンプレート１(ドライビングバー)
def print_excel_make1(output_list,excel_name,pg_folder_path,work_folder_path):

  #引き渡された要素数を取得
  list_len = len(output_list)

  #アウトプットのエクセルの準備
  dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻
  wb = work_folder_path+excel_name+"_"+dt_now.strftime('%Y%m%d%H%M%S') +'.xlsx'
  shutil.copy(pg_folder_path+'テンプレート１.xlsx', wb)
  
  out_wb = openpyxl.load_workbook(wb)
  out_wb.active.title=excel_name
  out_wb.save(wb)

  #処理対象件数0の場合
  if list_len == 0:
    out_wb.active.cell(1,1).value ="対象データなし"

  #x,y座標の定義
  x=0
  y=0
  
  #要素格納ループ
  for i in range(list_len):
    
    #x座標
    x=i%2*5
    #y座標
    y=i/2*9
    if i%2 == 1:
      y=(i-i%2)/2*9

    #格納処理  
    out_wb.active.cell(y+1, x+1).value = output_list[i][0] #製造番号
    out_wb.active.cell(y+1, x+3).value = output_list[i][1] #月次
    out_wb.active.cell(y+2, x+1).value = output_list[i][2] #国名
    out_wb.active.cell(y+2, x+2).value = output_list[i][3] #製品名称
    out_wb.active.cell(y+3, x+1).value = output_list[i][4] #部品補足1
    out_wb.active.cell(y+4, x+1).value = output_list[i][5] #部品補足2
    out_wb.active.cell(y+5, x+1).value = output_list[i][6] #部品1名称
    out_wb.active.cell(y+5, x+3).value = output_list[i][7] #部品1数量
    out_wb.active.cell(y+6, x+1).value = output_list[i][8] #部品2名称
    out_wb.active.cell(y+6, x+3).value = output_list[i][9] #部品2数量
    out_wb.active.cell(y+7, x+1).value = output_list[i][10] #部品3名称
    out_wb.active.cell(y+7, x+3).value = output_list[i][11] #部品3数量
    out_wb.active.cell(y+8, x+1).value = output_list[i][12] #部品4名称
    out_wb.active.cell(y+8, x+3).value = output_list[i][13] #部品4数量

  #印刷範囲を指定
  out_wb.active.print_area = 'A1:I'+str(int(y+8))

  #エクセルを保存
  out_wb.save(wb)
  out_wb.close()

###  テンプレート２(キャスター、KYK)
def print_excel_make2(output_list,excel_name,pg_folder_path,work_folder_path):
  
  #引き渡された要素数を取得
  list_len = len(output_list)

  #アウトプットのエクセルの準備
  dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻
  wb = work_folder_path+excel_name+"_"+dt_now.strftime('%Y%m%d%H%M%S') +'.xlsx'
  shutil.copy(pg_folder_path+'テンプレート２.xlsx', wb)
  
  out_wb = openpyxl.load_workbook(wb)
  out_wb.active.title=excel_name
  out_wb.save(wb)

  #処理対象件数0の場合
  if list_len == 0:
    out_wb.active.cell(1,1).value ="対象データなし"

  #x,y座標の定義
  x=0
  y=0
  
  #要素格納ループ
  for i in range(list_len):
    
    #x座標
    x=i%2*5
    #y座標
    y=i/2*9
    if i%2 == 1:
      y=(i-i%2)/2*9

    #格納処理  
    out_wb.active.cell(y+1, x+1).value = output_list[i][0] #製造番号
    out_wb.active.cell(y+1, x+3).value = output_list[i][1] #月次
    out_wb.active.cell(y+1, x+4).value = output_list[i][2] #分
    out_wb.active.cell(y+2, x+1).value = output_list[i][3] #製品名称1
    out_wb.active.cell(y+2, x+4).value = output_list[i][4] #チェックボックス
    out_wb.active.cell(y+3, x+1).value = output_list[i][5] #国名
    out_wb.active.cell(y+4, x+1).value = output_list[i][6] #空欄1
    out_wb.active.cell(y+3, x+2).value = output_list[i][7] #補足1
    out_wb.active.cell(y+5, x+1).value = output_list[i][8] #製品名称2
    out_wb.active.cell(y+7, x+1).value = output_list[i][9] #空欄2

  #印刷範囲を指定
  out_wb.active.print_area = 'A1:I'+str(int(y+8))

  #エクセルを保存
  out_wb.save(wb)
  out_wb.close()
 
###  テンプレート３(スクリュー、電柄Y付きチップ箱)
def print_excel_make3(output_list,excel_name,pg_folder_path,work_folder_path):
  
  #※特殊処理有り：スクリューはワーク部品数が３以下の時は結合する

  #引き渡された要素数を取得
  list_len = len(output_list)

  #アウトプットのエクセルの準備
  dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻
  wb = work_folder_path+excel_name+"_"+dt_now.strftime('%Y%m%d%H%M%S') +'.xlsx'
  shutil.copy(pg_folder_path+'テンプレート３.xlsx', wb)
  
  out_wb = openpyxl.load_workbook(wb)
  out_wb.active.title=excel_name
  out_wb.save(wb)

  #処理対象件数0の場合
  if list_len == 0:
    out_wb.active.cell(1,1).value ="対象データなし"

  #x,y座標の定義
  x=0
  y=0
  
  #要素格納ループ
  for i in range(list_len):
    
    #x座標
    x=i%2*5
    #y座標
    #y=i/2*11 #DEL_20240605_部品数7対応
    y=i/2*12  #ADD_20240605_部品数7対応
    if i%2 == 1:
      #y=(i-i%2)/2*11  #DEL_20240605_部品数7対応
      y=(i-i%2)/2*12  #ADD_20240605_部品数7対応
      
    #格納処理
    #output_list[i][19]がフラグ。1の時は部品数が3以下のため結合する
    #if output_list[i][19] == 0: # DEL_20240605_部品数7対応
    if output_list[i][20] == 0: # ADD_20240605_部品数7対応
      # #セル結合
      if i%2 == 0:
        out_wb.active.merge_cells('A'+str(int(y+5))+':B'+str(int(y+5)))
        out_wb.active.merge_cells('C'+str(int(y+5))+':D'+str(int(y+5)))
        out_wb.active.merge_cells('A'+str(int(y+6))+':B'+str(int(y+6)))
        out_wb.active.merge_cells('C'+str(int(y+6))+':D'+str(int(y+6)))
        out_wb.active.merge_cells('A'+str(int(y+7))+':B'+str(int(y+7)))
        out_wb.active.merge_cells('C'+str(int(y+7))+':D'+str(int(y+7)))
        out_wb.active.merge_cells('A'+str(int(y+8))+':B'+str(int(y+8)))
        out_wb.active.merge_cells('C'+str(int(y+8))+':D'+str(int(y+8)))
        out_wb.active.merge_cells('A'+str(int(y+9))+':B'+str(int(y+9)))
        out_wb.active.merge_cells('C'+str(int(y+9))+':D'+str(int(y+9)))
        out_wb.active.merge_cells('A'+str(int(y+10))+':B'+str(int(y+10)))
        out_wb.active.merge_cells('C'+str(int(y+10))+':D'+str(int(y+10)))
        out_wb.active.merge_cells('A'+str(int(y+11))+':B'+str(int(y+11))) #ADD_20240605_部品数7対応
        out_wb.active.merge_cells('C'+str(int(y+11))+':D'+str(int(y+11))) #ADD_20240605_部品数7対応
      else:
        out_wb.active.merge_cells('F'+str(int(y+5))+':G'+str(int(y+5)))
        out_wb.active.merge_cells('H'+str(int(y+5))+':I'+str(int(y+5)))
        out_wb.active.merge_cells('F'+str(int(y+6))+':G'+str(int(y+6)))
        out_wb.active.merge_cells('H'+str(int(y+6))+':I'+str(int(y+6)))
        out_wb.active.merge_cells('F'+str(int(y+7))+':G'+str(int(y+7)))
        out_wb.active.merge_cells('H'+str(int(y+7))+':I'+str(int(y+7)))
        out_wb.active.merge_cells('F'+str(int(y+8))+':G'+str(int(y+8)))
        out_wb.active.merge_cells('H'+str(int(y+8))+':I'+str(int(y+8)))
        out_wb.active.merge_cells('F'+str(int(y+9))+':G'+str(int(y+9)))
        out_wb.active.merge_cells('H'+str(int(y+9))+':I'+str(int(y+9)))
        out_wb.active.merge_cells('F'+str(int(y+10))+':G'+str(int(y+10)))
        out_wb.active.merge_cells('H'+str(int(y+10))+':I'+str(int(y+10)))
        out_wb.active.merge_cells('F'+str(int(y+11))+':G'+str(int(y+11))) #ADD_20240605_部品数7対応
        out_wb.active.merge_cells('H'+str(int(y+11))+':I'+str(int(y+11))) #ADD_20240605_部品数7対応
      
      out_wb.active.cell(y+1, x+1).value = output_list[i][0] #製造番号
      out_wb.active.cell(y+1, x+3).value = output_list[i][1] #月次
      out_wb.active.cell(y+2, x+1).value = output_list[i][2] #国名
      out_wb.active.cell(y+2, x+2).value = output_list[i][3] #製品名称
      out_wb.active.cell(y+3, x+1).value = output_list[i][4] #部品補足1
      out_wb.active.cell(y+4, x+1).value = output_list[i][5] #部品補足2
      out_wb.active.cell(y+5, x+1).value = output_list[i][6] #部品1名称
      out_wb.active.cell(y+5, x+3).value = output_list[i][7] #部品1数量
      out_wb.active.cell(y+6, x+1).value = output_list[i][8] #部品2名称
      out_wb.active.cell(y+6, x+3).value = output_list[i][9] #部品2数量
      out_wb.active.cell(y+7, x+1).value = output_list[i][10] #部品3名称
      out_wb.active.cell(y+7, x+3).value = output_list[i][11] #部品3数量
      out_wb.active.cell(y+8, x+1).value = output_list[i][12] #部品4名称
      out_wb.active.cell(y+8, x+3).value = output_list[i][13] #部品4数量
      out_wb.active.cell(y+9, x+1).value = output_list[i][14] #部品5名称
      out_wb.active.cell(y+9, x+3).value = output_list[i][15] #部品5数量
      out_wb.active.cell(y+10, x+1).value = output_list[i][16] #部品6名称
      out_wb.active.cell(y+10, x+3).value = output_list[i][17] #部品6数量
      #out_wb.active.cell(y+11, x+1).value = output_list[i][18] #空欄 DEL_20240605_部品数7対応
      out_wb.active.cell(y+11, x+1).value = output_list[i][18] #部品7名称 ADD_20240605_部品数7対応
      out_wb.active.cell(y+11, x+3).value = output_list[i][19] #部品7数量 ADD_20240605_部品数7対応
      
    #elif output_list[i][19] == 1:  DEL_20240605_部品数7対応
    elif output_list[i][20] == 1: #ADD_20240605_部品数7対応
      #セル結合
      if i%2 == 0:
        out_wb.active.merge_cells('A'+str(int(y+5))+':B'+str(int(y+6)))
        out_wb.active.merge_cells('C'+str(int(y+5))+':D'+str(int(y+6)))
        out_wb.active.merge_cells('A'+str(int(y+7))+':B'+str(int(y+8)))
        out_wb.active.merge_cells('C'+str(int(y+7))+':D'+str(int(y+8)))
        out_wb.active.merge_cells('A'+str(int(y+9))+':B'+str(int(y+10)))
        out_wb.active.merge_cells('C'+str(int(y+9))+':D'+str(int(y+10)))
      else:
        out_wb.active.merge_cells('F'+str(int(y+5))+':G'+str(int(y+6)))
        out_wb.active.merge_cells('H'+str(int(y+5))+':I'+str(int(y+6)))
        out_wb.active.merge_cells('F'+str(int(y+7))+':G'+str(int(y+8)))
        out_wb.active.merge_cells('H'+str(int(y+7))+':I'+str(int(y+8)))
        out_wb.active.merge_cells('F'+str(int(y+9))+':G'+str(int(y+10)))
        out_wb.active.merge_cells('H'+str(int(y+9))+':I'+str(int(y+10)))

      #フォント調整
      out_wb.active.cell(y+5, x+1).font = openpyxl.styles.Font(size = 22)
      out_wb.active.cell(y+5, x+3).font = openpyxl.styles.Font(size = 22)
      out_wb.active.cell(y+7, x+1).font = openpyxl.styles.Font(size = 22)
      out_wb.active.cell(y+7, x+3).font = openpyxl.styles.Font(size = 22)
      out_wb.active.cell(y+9, x+1).font = openpyxl.styles.Font(size = 22)
      out_wb.active.cell(y+9, x+3).font = openpyxl.styles.Font(size = 22)

      #データ格納
      out_wb.active.cell(y+1, x+1).value = output_list[i][0] #製造番号
      out_wb.active.cell(y+1, x+3).value = output_list[i][1] #月次
      out_wb.active.cell(y+2, x+1).value = output_list[i][2] #国名
      out_wb.active.cell(y+2, x+2).value = output_list[i][3] #製品名称
      out_wb.active.cell(y+3, x+1).value = output_list[i][4] #部品補足1
      out_wb.active.cell(y+4, x+1).value = output_list[i][5] #部品補足2
      out_wb.active.cell(y+5, x+1).value = output_list[i][6] #部品1名称
      out_wb.active.cell(y+5, x+3).value = output_list[i][7] #部品1数量
      out_wb.active.cell(y+7, x+1).value = output_list[i][8] #部品2名称
      out_wb.active.cell(y+7, x+3).value = output_list[i][9] #部品2数量
      out_wb.active.cell(y+9, x+1).value = output_list[i][10] #部品3名称
      out_wb.active.cell(y+9, x+3).value = output_list[i][11] #部品3数量
      #out_wb.active.cell(y+11, x+1).value = output_list[i][18] #空欄

    out_wb.save(wb)
  #印刷範囲を指定
  #out_wb.active.print_area = 'A1:I'+str(int(y+11)) #DEL_20240605_部品数7対応
  out_wb.active.print_area = 'A1:I'+str(int(y+12)) #ADD_20240605_部品数7対応

  #エクセルを保存
  out_wb.save(wb)
  out_wb.close()

###  ODチェックリストの出力（簡単なのでそのまま出す）
def check_excel_make(output_list,excel_name,pg_folder_path,work_folder_path):

  #引き渡された要素数を取得
  list_len = len(output_list)

  #アウトプットのエクセルの準備
  dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻
  wb = work_folder_path+excel_name+"_"+dt_now.strftime('%Y%m%d%H%M%S') +'.xlsx'
  shutil.copy(pg_folder_path+'ODチェック表.xlsx', wb)
  
  out_wb = openpyxl.load_workbook(wb)
  out_wb.active.title=excel_name
  out_wb.save(wb)

  #処理対象件数0の場合
  if list_len == 0:
    out_wb.active.cell(1,1).value ="対象データなし"

  #要素格納ループ
  for i in range(list_len):    #格納処理  
    out_wb.active.cell(i+5, 1).value = output_list[i][0] #月次
    out_wb.active.cell(i+5, 2).value = output_list[i][1] #製造番号
    out_wb.active.cell(i+5, 3).value = output_list[i][2] #製品名称

  #印刷範囲を指定
  out_wb.active.print_area = 'A1:S'+str(int(4+list_len))

  #エクセルを保存
  out_wb.save(wb)
  out_wb.close()

###　エラーリストの出力（処理後にPADで開きたいのでエクセルで出力）
def err_excel_make(err1,err2,err_folder_path,time,cnt=[0]):
  dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻
  excel_name = '部品数表示超過一覧'+''"_"+dt_now.strftime('%Y%m%d') +'.xlsx'
  wb = err_folder_path+excel_name
  files = glob(wb)

  #エラーエクセルがある場合は読み取って追記
  if wb in files:
    err_wb = openpyxl.load_workbook(wb)
  else:
    err_wb = openpyxl.Workbook()
    err_wb.save(wb)
    
    #列は固定値を指定する
    err_wb.active.column_dimensions["A"].width = 5
    err_wb.active.column_dimensions["B"].width = 15
    err_wb.active.column_dimensions["C"].width = 40
    err_wb.active.column_dimensions["D"].width = 100

    #シート内に固定文字出力（ヘッダー部）
    err_wb.active["A1"] = "部品数表示超過一覧"
    err_wb.active['A1'].font = openpyxl.styles.Font(size = 24)
    #err_wb.active["A2"] = "スクリューは6種類、ドライビングバーは4種類がシールへの表示の上限です。" DEL_20240605_部品数7対応
    err_wb.active["A2"] = "スクリューは7種類、ドライビングバーは4種類がシールへの表示の上限です。" #ADD_20240605_部品数7対応
    err_wb.active["A3"] = "以下のリストにシールに入りきらなかった部品があったものを記載しています。"
    err_wb.active["A4"] = "当リストが表示された場合は、お手数ですが対象のシールエクセルに手作業で追加いただきますようお願いいたします。"

    err_wb.active["A6"] = "No."
    err_wb.active["B6"] = "処理時間"
    err_wb.active["C6"] = "内容"
    err_wb.active["D6"] = "対象の製造番号と超過した部品の一覧"

    # 固定
    err_wb.active.freeze_panes = "A7"

    #横向き
    err_wb.active.page_setup.orientation="landscape"

  cnt[0] += 1
  err_wb.active["A"+str(err_wb.active.max_row + 1)] = cnt[0]
  err_wb.active["B"+str(err_wb.active.max_row)] = time
  err_wb.active["C"+str(err_wb.active.max_row)] = err1
  err_wb.active["D"+str(err_wb.active.max_row)] = err2

  # 保存
  err_wb.save(wb)
  err_wb.close()

################################################################
###　                     メイン処理　                        ###
################################################################
try:

################################################################
###　               準備＋データ加工処理　                     ###
################################################################

  ## 引数を取得（PAD経由では、引数[0]だと自分自身のpathが取得されてしまう）
  work_dir = sys.argv[1] #workフォルダ
  pg_dir = sys.argv[2] #pgフォルダ
  output_dir = sys.argv[3] #アウトプットフォルダ
  errlog_dir = sys.argv[4] #エラーログフォルダ
  log_level = int(sys.argv[5]) #ログ出力のレベル
  
  # コピーしてきた構成情報エクセルを取得
  books_seisanki = glob(work_dir+r"生産機情報.xls*")
  # コピーしてきた構成情報エクセルを取得
  books_kousei = glob(work_dir+r"構成\*.*xls*")

  # 結果格納用の空のデータフレームを用意する。
  result_df1 = pandas.DataFrame() #ドライビングバー
  result_df2 = pandas.DataFrame() #スクリュー
  result_df3 = pandas.DataFrame() #電柄 Y付きチップ箱
  result_df4 = pandas.DataFrame() #KYK
  result_df5 = pandas.DataFrame() #キャスター
  result_df6 = pandas.DataFrame() #ODチェック表
  all_df = pandas.DataFrame() #記録用の全件df
  
  # 生産機情報のexcel開き、必要な列だけスライス表記で抽出
  df1 = pandas.read_excel(books_seisanki[0],sheet_name=0,usecols="A:T")
  df1 = df1.iloc[:,[0,3,17,18]]

  #Factのエクセルを開き、生産機情報と結合。
  #抽出処理のためにフルパス情報を作成する。
  for book1 in books_kousei: 
    #ファイル名取得
    fname=book1[-16:-8]
    # Factのexcel開く
    df2 = pandas.read_excel(book1,sheet_name=0,usecols="A:M")
    # 必要な列だけスライス表記で抽出
    df2 = df2.iloc[:,[1,2,3,9,0]]
    # 空白行を削除（1レコードに対してすべての列がNanの場合にその行を削除）　不要と思われるが、引っかかってもいやなので記載
    df2.dropna(axis="index",how='all',inplace=True)
    #ファイル名を取得し、df2に追加
    df2.loc[:,"製番"] = fname
    #製番をキーに、両エクセルの内容を結合
    df3 = pandas.merge(df1, df2)
    df3.columns = ['A','B','C','D','E','F','G','H','I']

    mask = df3['I'] > df3['I'].shift(1)
    df3.loc[mask, 'J'] = df3['F']
    df3.loc[~mask, 'J'] = 0 #それ以外は0
    df3['C'] = df3['C'].astype(str) #月次は文字列に型変換しておく

    tmp_list =["階層リスト："]
    for idx in range(df3.shape[0]):
      
      # ひとつ前と比較
      comp = df3.iloc[idx-1,-2]-df3.iloc[idx,-2]

      if comp < 0:
       #tmp_list.append(df3.iloc[idx-1,-5]) #ひとつ前のFを追加
       tmp_list.append(df3.iloc[idx-1,-5]+"＋"+df3.iloc[idx-1,-6]) #ひとつ前のFを追加
      elif comp > 0 and idx != 0:
        for idx2 in range(comp):
          del tmp_list[-1]
      
      df3.iloc[idx,-1] = str(tmp_list)+","+df3.iloc[idx,-5]+"＋"+df3.iloc[idx,-6]

      #月次の修正
      df3.iloc[idx,2] = "("+str(int(str(df3.iloc[idx,2])[-3:-1]))+"-"+str(df3.iloc[idx,2])[-1]+")"

    #記録用の全件df
    #ログを出力するときのみ  
    if log_level == 1:
      all_df = pandas.concat([all_df,df3])
    
################################################################
### ここまででデータの前処理を完了　DF3の情報を用いて、データ抽出 ###
################################################################
#必要とするリスト
#テンプレート１(ドライビングバー)[製造番号,月次,国名,製品名称,部品補足1,部品補足2,部品1名称,部品1数量,部品2名称,部品2数量,部品3名称,部品3数量,部品4名称,部品4数量]
#テンプレート２(キャスター、KYK)[製造番号,月次,分,製品名称1,チェックボックス,国名,空欄1,補足1,製品名称2,空欄3]
#テンプレート３(スクリュー、電柄Y付チップ箱)[製造番号,月次,国名,製品名称1,部品補足1,部品補足2,部品1名称,部品1数量,部品2名称,部品2数量,部品3名称,部品3数量,部品4名称,部品4数量,部品5名称,部品5数量,部品6名称,部品6数量,空欄,フラグ（部品数が3以下で1）] DEL_20240605_部品数7
#テンプレート３(スクリュー、電柄Y付チップ箱)[製造番号,月次,国名,製品名称1,部品補足1,部品補足2,部品1名称,部品1数量,部品2名称,部品2数量,部品3名称,部品3数量,部品4名称,部品4数量,部品5名称,部品5数量,部品6名称,部品6数量,部品7名称,部品7数量,フラグ（部品数が3以下で1）] ADD_20240605_部品数7

    #抽出条件エクセルファイルを開く+アクティブシートの取得
    in_wb = openpyxl.load_workbook(pg_dir+"抽出条件.xlsx")
    in_sheet = in_wb.active

    ##############################
    ###１．ドライビングバー(df4) ###
    ##############################
    
    df4 = df3

    #抽出行を取得（抽出処理は関数化不可 DFが引数化できないため）
    for row in in_sheet.iter_rows(min_row=6):
      in_cells = []

      #セルをlistで取得
      for cell in row:
        in_cells.append(cell.value)

      if in_cells[0] == "ドライビングバー":
        if in_cells[1] is not None:
          df4 = df4[df4['B'].str.contains(in_cells[1],na=False)]
        if in_cells[2] is not None:
          df4 = df4[~df4['B'].str.contains(in_cells[2],na=False)]
        if in_cells[3] is not None:
          df4 = df4[df4['J'].str.contains(in_cells[3],na=False)]
        if in_cells[4] is not None:
          df4 = df4[~df4['J'].str.contains(in_cells[4],na=False)]
    
    #印刷用エクセル関数にデータを渡すための加工処理
    #ラベルの付け替え
    df4 = df4.loc[:,['A','C','D','B','G','H']]
    df4 = df4.set_axis(['製造番号','月次','国名','製品名称1','部品1名称','部品1数量'],axis=1)

    #部品数量を合計表示
    df4 = df4.groupby(['製造番号','月次','国名','製品名称1','部品1名称'])['部品1数量'].sum().reset_index()
    
    #足りない項目を追加、列の入替え
    df4['部品補足1'] = ""
    df4['部品補足2'] = "Driving Bar Fixed Screws"
    df4['部品2名称'] = ""
    df4['部品2数量'] = ""
    df4['部品3名称'] = ""
    df4['部品3数量'] = ""
    df4['部品4名称'] = ""
    df4['部品4数量'] = ""
    df4 = df4.reindex(['製造番号','月次','国名','製品名称1','部品補足1','部品補足2','部品1名称','部品1数量','部品2名称','部品2数量','部品3名称','部品3数量','部品4名称','部品4数量'], axis=1)
    
    #マスク処理でデータフレームを編集
    mask = df4['製造番号'] == df4['製造番号'].shift(-1)
    df4.loc[mask, '部品2名称'] = df4['部品1名称'].shift(-1)
    df4.loc[mask, '部品2数量'] = df4['部品1数量'].shift(-1)
    mask = df4['製造番号'] == df4['製造番号'].shift(-2)
    df4.loc[mask, '部品3名称'] = df4['部品1名称'].shift(-2)
    df4.loc[mask, '部品3数量'] = df4['部品1数量'].shift(-2)
    mask = df4['製造番号'] == df4['製造番号'].shift(-3)
    df4.loc[mask, '部品4名称'] = df4['部品1名称'].shift(-3)
    df4.loc[mask, '部品4数量'] = df4['部品1数量'].shift(-3)
    
    #部品が5列以上存在したとき（5個以上部品が有ったらエラー処理）
    mask = df4['製造番号'] == df4['製造番号'].shift(-4)
    if df4.loc[mask,:].shape[0] > 0:
      df4err1 = "【ドライビングバー】5つ以上の部品が存在します"
      df4err2 = str(df4.iloc[0,0])+'の部品一覧：'
      if log_level == 1:
        for idx in range(df4.loc[mask,:].shape[0]):
          if df4.loc[mask,:].iloc[0,0] == df4.loc[mask,:].iloc[idx,0]:
            df4err2 += '　'+str(idx+4+1)+'：'+str(df4.iloc[idx+4,6])+'×'+str(df4.iloc[idx+4,7])
      else:
        df4err2 += str(df4.loc[mask,'部品1名称'])
      dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻     

      dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻      
      err_excel_make(df4err1,df4err2,errlog_dir,dt_now.strftime('%H時%M分%S秒'))

    mask = df4['製造番号'] == df4['製造番号'].shift(1)
    df4 = df4.loc[~mask,:] 
    
    #処理結果を保存
    result_df1 = pandas.concat([result_df1,df4])

    ##############################
    ### ２．スクリュー(df5)　   ###
    ##############################
    df5 = df3

    #キャスターの種類分け用
    caster_list = []

    #抽出行を取得（抽出処理は関数化不可 DFが引数化できないため）
    for row in in_sheet.iter_rows(min_row=6):
      in_cells = []

      #セルをlistで取得
      for cell in row:
        in_cells.append(cell.value)

      if in_cells[0] == "スクリュー":
        if in_cells[1] is not None:
          df5 = df5[df5['B'].str.contains(in_cells[1],na=False)]
        if in_cells[2] is not None:
          df5 = df5[~df5['B'].str.contains(in_cells[2],na=False)]
        if in_cells[3] is not None:
          df5 = df5[df5['J'].str.contains(in_cells[3],na=False)]
        if in_cells[4] is not None:
          df5 = df5[~df5['J'].str.contains(in_cells[4],na=False)]
        #スクリュー専用の分岐 in_cells[6]　
        if in_cells[6] is not None:
          caster_list.append(in_cells[6])
    
    #印刷用エクセル関数にデータを渡すための加工処理

    #部品補足のセット
    df5['部品補足1'] = "なし"  #一旦格納 
    df5['部品補足2'] = ""
    mask = df5['J'].str.contains(caster_list[0],na=False)
    #df5.loc[mask, '部品補足1'] = 'ガードネットキャスター' DEL_英語にする
    df5.loc[mask, '部品補足1'] = 'GUARDNET CASTERS'
    #df5.loc[mask, '部品補足2'] = 'ステップ取付けネジ'
    mask = df5['J'].str.contains(caster_list[0],na=False)&df5['J'].str.contains('(?=.*ASSY. GUARD NET)',na=False)
    df5.loc[mask, '部品補足2'] = 'Step Fixed Screws'
    mask = df5['J'].str.contains(caster_list[1],na=False)
    #df5.loc[mask, '部品補足1'] = 'ステップキャスター'
    df5.loc[mask, '部品補足1'] = 'STEP CASTERS'
    mask = df5['J'].str.contains(caster_list[2],na=False)
    #df5.loc[mask, '部品補足1'] = 'インホイールモーター'
    df5.loc[mask, '部品補足1'] = 'IN WHEEL MOTOR CASTERS'

    #ADD_20240529 4つ目以降の例外が発生したときようの処理
    for idx in range(len(caster_list)):
      if idx > 2:
        mask = df5['J'].str.contains(caster_list[idx],na=False)
        i = caster_list[idx].find('(')
        df5.loc[mask, '部品補足1'] = caster_list[idx][i+1:-2]

    #部品補足の入っていないデータは省く
    mask = df5['部品補足1']=='なし'
    df5 = df5.loc[~mask,:] 

    #ラベルの付け替え
    df5 = df5.loc[:,['A','C','D','B','部品補足1','部品補足2','G','H']]
    df5 = df5.set_axis(['製造番号','月次','国名','製品名称1','部品補足1','部品補足2','部品1名称','部品1数量'],axis=1)

    df5 = df5.sort_values(['月次','製造番号','部品補足1']) #部品補足1で判定させるため、ソートをかけたうえで、上下比較
    
    #部品数量を合計表示
    df5 = df5.groupby(['製造番号','月次','国名','製品名称1','部品補足1','部品補足2','部品1名称'])['部品1数量'].sum().reset_index() 

    #足りない項目を追加、列の入替え
    df5['部品2名称'] = ""
    df5['部品2数量'] = ""
    df5['部品3名称'] = ""
    df5['部品3数量'] = ""
    df5['部品4名称'] = ""
    df5['部品4数量'] = ""
    df5['部品5名称'] = ""
    df5['部品5数量'] = ""
    df5['部品6名称'] = ""
    df5['部品6数量'] = ""
    #df5['空欄'] = ""  DEL_20240605_部品数7に変更
    df5['部品7名称'] = ""   #ADD_20240605_部品数7に変更
    df5['部品7数量'] = ""   #ADD_20240605_部品数7に変更
    df5['フラグ'] = 0
    #df5 = df5.reindex(['製造番号','月次','国名','製品名称1','部品補足1','部品補足2','部品1名称','部品1数量','部品2名称','部品2数量','部品3名称','部品3数量','部品4名称','部品4数量','部品5名称','部品5数量','部品6名称','部品6数量','空欄','フラグ'], axis=1) #DEL_20240605_部品数7対応
    df5 = df5.reindex(['製造番号','月次','国名','製品名称1','部品補足1','部品補足2','部品1名称','部品1数量','部品2名称','部品2数量','部品3名称','部品3数量','部品4名称','部品4数量','部品5名称','部品5数量','部品6名称','部品6数量','部品7名称','部品7数量','フラグ'], axis=1) #ADD_20240605_部品数7対応

    #マスク処理でデータフレームを編集
    df5 = df5.sort_values(['月次','製造番号','部品補足1']) #部品補足1で判定させるため、ソートをかけたうえで、上下比較
    mask = df5['部品補足1'] == df5['部品補足1'].shift(-1)
    df5.loc[mask, '部品2名称'] = df5['部品1名称'].shift(-1)
    df5.loc[mask, '部品2数量'] = df5['部品1数量'].shift(-1)
    mask = df5['部品補足1'] == df5['部品補足1'].shift(-2)
    df5.loc[mask, '部品3名称'] = df5['部品1名称'].shift(-2)
    df5.loc[mask, '部品3数量'] = df5['部品1数量'].shift(-2)
    mask = df5['部品補足1'] == df5['部品補足1'].shift(-3)
    df5.loc[mask, '部品4名称'] = df5['部品1名称'].shift(-3)
    df5.loc[mask, '部品4数量'] = df5['部品1数量'].shift(-3)
    df5.loc[~mask, 'フラグ'] = 1
    mask = df5['部品補足1'] == df5['部品補足1'].shift(-4)
    df5.loc[mask, '部品5名称'] = df5['部品1名称'].shift(-4)
    df5.loc[mask, '部品5数量'] = df5['部品1数量'].shift(-4)
    mask = df5['部品補足1'] == df5['部品補足1'].shift(-5)
    df5.loc[mask, '部品6名称'] = df5['部品1名称'].shift(-5)
    df5.loc[mask, '部品6数量'] = df5['部品1数量'].shift(-5)
    mask = df5['部品補足1'] == df5['部品補足1'].shift(-6)   #ADD_20240605_部品数7対応
    df5.loc[mask, '部品7名称'] = df5['部品1名称'].shift(-6)   #ADD_20240605_部品数7対応
    df5.loc[mask, '部品7数量'] = df5['部品1数量'].shift(-6)   #ADD_20240605_部品数7対応

    #部品が7列以上存在したとき（7個以上部品が有ったらエラー処理）
    #mask = df5['部品補足1'] == df5['部品補足1'].shift(-6) #DEL_20240605_部品数7対応
    mask = df5['部品補足1'] == df5['部品補足1'].shift(-7) #ADD_20240605_部品数7対応
    if df5.loc[mask,:].shape[0] > 0:
      #df5err1 = "【スクリュー】7つ以上の部品が存在します" #DEL_20240605_部品数7対応
      df5err1 = "【スクリュー】8つ以上の部品が存在します" #ADD_20240605_部品数7対応
      df5err2 = str(df5.iloc[0,0])+'の部品一覧：'
      if log_level == 1:
        for idx in range(df5.loc[mask,:].shape[0]):
          if df5.loc[mask,:].iloc[0,0] == df5.loc[mask,:].iloc[idx,0]:
            #df5err2 += '　'+str(idx+6+1)+'：'+str(df5.iloc[idx+6,6])+'×'+str(df5.iloc[idx+6,7]) #DEL_20240605_部品数7対応
            df5err2 += '　'+str(idx+7+1)+'：'+str(df5.iloc[idx+7,6])+'×'+str(df5.iloc[idx+7,7]) #ADD_20240605_部品数7対応
      else:
        df5err2 += str(df5.loc[mask,'部品1名称'])
      dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻      
      err_excel_make(df5err1,df5err2,errlog_dir,dt_now.strftime('%H時%M分%S秒'))
          
    mask = df5['部品補足1'] == df5['部品補足1'].shift(1) #部品補足1で比較する。製造番号が同じでも、2枚のシールが出ることがある。
    df5 = df5.loc[~mask,:] 
    if df5.shape[0] == 1:
      df5.loc[:, '部品補足2'] = ""
      
    result_df2 = pandas.concat([result_df2,df5])

    ##############################
    # ３．電柄 Y付きチップ箱(df6)  #
    ##############################
    df6 = df3

    #抽出行を取得（抽出処理は関数化不可 DFが引数化できないため）
    for row in in_sheet.iter_rows(min_row=6):
      in_cells = []

      #セルをlistで取得
      for cell in row:
        in_cells.append(cell.value)

      if in_cells[0] == "電柄Y付チップ箱":
        if in_cells[1] is not None:
          df6 = df6[df6['B'].str.contains(in_cells[1],na=False)]
        if in_cells[2] is not None:
          df6 = df6[~df6['B'].str.contains(in_cells[2],na=False)]
        if in_cells[3] is not None:
          df6 = df6[df6['J'].str.contains(in_cells[3],na=False)]
        if in_cells[4] is not None:
          df6 = df6[~df6['J'].str.contains(in_cells[4],na=False)]

    #ラベルの付け替え
    df6 = df6.loc[:,['A','C','D','B']]
    df6 = df6.set_axis(['製造番号','月次','国名','製品名称1'],axis=1)
    #重複を削除
    df6 = df6.drop_duplicates()
    
    #足りない項目を追加
    df6['部品補足1'] =""
    df6['部品補足2'] =""
    df6['部品1名称'] =""
    df6['部品1数量'] =""
    df6['部品2名称'] =""
    df6['部品2数量'] =""
    df6['部品3名称'] =""
    df6['部品3数量'] =""
    df6['部品4名称'] =""
    df6['部品4数量'] =""
    df6['部品5名称'] =""
    df6['部品5数量'] =""
    df6['部品6名称'] =""
    df6['部品6数量'] =""
    df6['部品7名称'] ="" #ADD_20240605_部品数7対応
    df6['部品7数量'] ="" #ADD_20240605_部品数7対応
    #df6['空欄'] ="" #DEL_20240605_部品数7対応
    df6['フラグ'] =1
    #df6 = df6.reindex(['製造番号','月次','国名','製品名称1','部品補足1','部品補足2','部品1名称','部品1数量','部品2名称','部品2数量','部品3名称','部品3数量','部品4名称','部品4数量','部品5名称','部品5数量','部品6名称','部品6数量','空欄','フラグ'], axis=1) #DEL_20240605_部品数7対応
    df6 = df6.reindex(['製造番号','月次','国名','製品名称1','部品補足1','部品補足2','部品1名称','部品1数量','部品2名称','部品2数量','部品3名称','部品3数量','部品4名称','部品4数量','部品5名称','部品5数量','部品6名称','部品6数量','部品7名称','部品7数量','フラグ'], axis=1) #ADD_20240605_部品数7対応

    #個別の条件を設定
    #ULM有り
    mask = df6['製品名称1'].str.contains('LEC|SEC',na=False)
    df6.loc[mask, '部品1名称'] = "U.L.M. (USB)"
    df6.loc[mask, '部品1数量'] = 1
    ### DEL_　ActWireの出力はなし（工具箱にいれるため
    # #Act Wire有り LEC
    # mask = df6['製品名称1'].str.contains('LEC',na=False)
    # df6.loc[mask, '部品2名称'] = "Act Wire"
    # df6.loc[mask, '部品2数量'] = 2
    # mask = df6['製品名称1'].str.contains('LEC.D',na=False)
    # df6.loc[mask, '部品2数量'] = 4
    # #Act Wire有り SEC
    # mask = df6['製品名称1'].str.contains('SEC',na=False)
    # df6.loc[mask, '部品2名称'] = "Act Wire"
    # df6.loc[mask, '部品2数量'] = 2
    # mask = df6['製品名称1'].str.contains('SECP',na=False)
    # df6.loc[mask, '部品2数量'] = 4
    #NEEDLE有り LEC SEC
    mask = df6['製品名称1'].str.contains('LEC|SEC',na=False)&df6['製品名称1'].str.contains('Y',na=False)
    df6.loc[mask, '部品2名称'] = "NEEDLE"
    df6.loc[mask, '部品2数量'] = 1
    #NEEDLE有り LEC SEC 以外
    mask = ~df6['製品名称1'].str.contains('LEC|SEC',na=False)&df6['製品名称1'].str.contains('Y',na=False)
    df6.loc[mask, '部品1名称'] = "NEEDLE"
    df6.loc[mask, '部品1数量'] = 1

    #部品1が無いものを除外
    mask = df6['部品1名称']==""
    df6 = df6.loc[~mask,:] 

    result_df3 = pandas.concat([result_df3,df6])

    ##############################
    ###　　  ４．KYK(df7)　　　 ###
    ##############################    
    df7 = df3

    #抽出行を取得（抽出処理は関数化不可 DFが引数化できないため）
    for row in in_sheet.iter_rows(min_row=6):
      in_cells = []

      #セルをlistで取得
      for cell in row:
        in_cells.append(cell.value)

      if in_cells[0] == "KYK":
        if in_cells[1] is not None:
          df7 = df7[df7['B'].str.contains(in_cells[1],na=False)]
        if in_cells[2] is not None:
          df7 = df7[~df7['B'].str.contains(in_cells[2],na=False)]
        if in_cells[3] is not None:
          df7 = df7[df7['J'].str.contains(in_cells[3],na=False)]
        if in_cells[4] is not None:
          df7 = df7[~df7['J'].str.contains(in_cells[4],na=False)]
    

    #ラベルの付け替え
    df7 = df7.loc[:,['A','C','D','B']]
    df7 = df7.set_axis(['製造番号','月次','国名','製品名称2'],axis=1)
    #重複を削除
    df7 = df7.drop_duplicates()
    #2枚印刷するため、データフレーム自体を追加
    df7 = pandas.concat([df7, df7], axis=0)
    
    #足りない項目を追加、並び替え
    df7['分'] = ""
    df7['製品名称1'] = ""
    df7['チェックボックス'] = ""
    df7['空欄1'] = ""
    df7['補足1'] = ""
    df7['空欄2'] = ""
    df7 = df7.reindex(['製造番号','月次','分','製品名称1','チェックボックス','国名','空欄1','補足1','製品名称2','空欄2'], axis=1)

    #処理結果を保存
    result_df4 = pandas.concat([result_df4,df7])


    ##############################
    ###  ５．キャスター(df8)    ###
    ##############################
    df8 = df3

    #抽出行を取得（抽出処理は関数化不可 DFが引数化できないため）
    #2枚出力チェック用
    double_list = []

    for row in in_sheet.iter_rows(min_row=6):
      in_cells = []

      #セルをlistで取得
      for cell in row:
        in_cells.append(cell.value)

      if in_cells[0] == "キャスター":
        if in_cells[1] is not None:
          df8 = df8[df8['B'].str.contains(in_cells[1],na=False)]
        if in_cells[2] is not None:
          df8 = df8[~df8['B'].str.contains(in_cells[2],na=False)]
        if in_cells[3] is not None:
          df8 = df8[df8['J'].str.contains(in_cells[3],na=False)]
        if in_cells[4] is not None:
          df8 = df8[~df8['J'].str.contains(in_cells[4],na=False)]
        #キャスター専用の分岐 in_cells[5]　を含む場合は2枚印刷する ASSY. STEP
        if in_cells[5] is not None:
          double_list.append(in_cells[5])
    
    #必要項目を先に追加
    df8['分'] = ""
    df8['製品名称1'] = ""
    
    #mask 
    for list in double_list:
      mask = df8['J'].str.contains(list,na=False)
      df8.loc[mask, '分'] = '2／2'
      #df8.loc[mask, '製品名称1'] = 'ステップキャスター' DEL_英語にする
      df8.loc[mask, '製品名称1'] = 'STEP CASTERS'
      df8.loc[~mask, '分'] = '1／1'
      #df8.loc[~mask, '製品名称1'] = 'ガードネットキャスター' DEL_英語にする
      df8.loc[~mask, '製品名称1'] = 'GUARDNET CASTERS'

    #ラベルの付け替え
    df8 = df8.loc[:,['A','C','分','製品名称1','D','B']]
    df8 = df8.set_axis(['製造番号','月次','分','製品名称1','国名','製品名称2'],axis=1)   
    #一旦重複を削除
    df8 = df8.drop_duplicates()

    #mask = df8['製品名称1'].shift(-1) == 'ステップキャスター' DEL_英語にする
    mask = df8['製品名称1'].shift(-1) == 'STEP CASTERS'
    df8.loc[mask, '分'] = '1／2'

    #足りない項目を追加、並び替え
    df8['チェックボックス'] = "□"
    df8['空欄1'] = ""
    df8['補足1'] = "SCREW・WASHER"    
    df8['空欄2'] = ""
    df8 = df8.reindex(['製造番号','月次','分','製品名称1','チェックボックス','国名','空欄1','補足1','製品名称2','空欄2'], axis=1)

    #処理結果を保存
    result_df5 = pandas.concat([result_df5,df8])


    #抽出条件エクセルファイルを閉じる
    in_wb.close()

    ##############################
    ###  ６．ODチェック表(df9)    ###
    ##############################
    df9 = df3

    #抽出行を取得（抽出処理は関数化不可 DFが引数化できないため）
    for row in in_sheet.iter_rows(min_row=6):
      in_cells = []

      #セルをlistで取得
      for cell in row:
        in_cells.append(cell.value)

      if in_cells[0] == "ODチェック":
        if in_cells[1] is not None:
          df9 = df9[df9['B'].str.contains(in_cells[1],na=False)]
        if in_cells[2] is not None:
          df9 = df9[~df9['B'].str.contains(in_cells[2],na=False)]
        if in_cells[3] is not None:
          df9 = df9[df9['J'].str.contains(in_cells[3],na=False)]
        if in_cells[4] is not None:
          df9 = df9[~df9['J'].str.contains(in_cells[4],na=False)]
    
    #ラベルの付け替え
    df9 = df9.loc[:,['C','A','B']]
    df9 = df9.set_axis(['月次','製造番号','製品名称'],axis=1)
    #重複を削除
    df9 = df9.drop_duplicates()
    #月次の再修正
    for idx in range(df9.shape[0]):
      df9.iloc[idx,0] =str(int(str(df9.iloc[idx,0])[-2:-1]))+"次"
    #処理結果を保存
    result_df6 = pandas.concat([result_df6,df9])

    #抽出条件エクセルファイルを閉じる
    in_wb.close()

################################################################
###　　　　　　　　　　　データ抽出結果の保存　　　　　　　　　　　###
################################################################
  #エクセルの保存
  #マージ結果作成用エクセルファイルの新規作成
  #まとめてソート
  result_df1 = result_df1.sort_values(['月次','製造番号'])
  result_df2 = result_df2.sort_values(['月次','製造番号'])
  result_df3 = result_df3.sort_values(['月次','製造番号'])
  result_df4 = result_df4.sort_values(['月次','製造番号'])
  result_df5 = result_df5.sort_values(['月次','製造番号'])
  result_df6 = result_df6.sort_values(['月次','製造番号'])

  # １．ドライビングバー
  #配列化してエクセル関数に渡す
  list_result_df1 = result_df1.to_numpy().tolist()
  print_excel_make1(list_result_df1,"シール_ドライビングバー",pg_dir,output_dir)

  # ２．スクリュー
  #配列化してエクセル関数に渡す
  list_result_df2 = result_df2.to_numpy().tolist()
  print_excel_make3(list_result_df2,"シール_スクリュー",pg_dir,output_dir)

  # ３．電柄Y付きチップ箱
  #配列化してエクセル関数に渡す
  list_result_df3 = result_df3.to_numpy().tolist()
  print_excel_make3(list_result_df3,"シール_電柄Y付きチップ箱",pg_dir,output_dir)
  
  # ４．KYK
  #配列化してエクセル関数に渡す
  list_result_df4 = result_df4.to_numpy().tolist()
  print_excel_make2(list_result_df4,"シール_KYK",pg_dir,output_dir)
  
  # ５．キャスター
  #配列化してエクセル関数に渡す
  list_result_df5 = result_df5.to_numpy().tolist()
  print_excel_make2(list_result_df5,"シール_キャスター",pg_dir,output_dir)  

  # ６．ODチェック表
  #配列化してエクセル関数に渡す
  list_result_df6 = result_df6.to_numpy().tolist()
  check_excel_make(list_result_df6,"ODチェック表",pg_dir,output_dir)

  # 処理結果のバックアップ
  dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻
  
  #ログを出力するときのみ  
  if log_level == 1:
    back_wb = openpyxl.Workbook()
    back_wb_name = dt_now.strftime('%Y%m%d%H%M%S')+"_中間全件データバックアップ"+".xlsx"
    all_df.to_excel(work_dir+"tmp\\"+back_wb_name,sheet_name="中間全件データ",index=False,header=True)
    back_wb.close()

  back_wb = openpyxl.Workbook()
  back_wb_name = dt_now.strftime('%Y%m%d%H%M%S')+"_ドライビングバー処理結果バックアップ"+".xlsx"
  result_df1.to_excel(work_dir+"tmp\\"+back_wb_name,sheet_name="ドライビングバー",index=False,header=True)
  back_wb.close()
  
  back_wb = openpyxl.Workbook()
  back_wb_name = dt_now.strftime('%Y%m%d%H%M%S')+"_スクリュー処理結果バックアップ"+".xlsx"
  result_df2.to_excel(work_dir+"tmp\\"+back_wb_name,sheet_name="スクリュー",index=False,header=True)
  back_wb.close()

  back_wb = openpyxl.Workbook()
  back_wb_name = dt_now.strftime('%Y%m%d%H%M%S')+"_電柄 Y付きチップ箱処理結果バックアップ"+".xlsx"
  result_df3.to_excel(work_dir+"tmp\\"+back_wb_name,sheet_name="電柄 Y付きチップ箱",index=False,header=True)
  back_wb.close()

  back_wb = openpyxl.Workbook()
  back_wb_name = dt_now.strftime('%Y%m%d%H%M%S')+"_KYK処理結果バックアップ"+".xlsx"
  result_df4.to_excel(work_dir+"tmp\\"+back_wb_name,sheet_name="KYK",index=False,header=True)
  back_wb.close()

  back_wb = openpyxl.Workbook()
  back_wb_name = dt_now.strftime('%Y%m%d%H%M%S')+"_キャスター処理結果バックアップ"+".xlsx"
  result_df5.to_excel(work_dir+"tmp\\"+back_wb_name,sheet_name="キャスター",index=False,header=True)
  back_wb.close()

  back_wb = openpyxl.Workbook()
  back_wb_name = dt_now.strftime('%Y%m%d%H%M%S')+"_ODチェック表処理結果バックアップ"+".xlsx"
  result_df6.to_excel(work_dir+"tmp\\"+back_wb_name,sheet_name="ODチェック表",index=False,header=True)
  back_wb.close()

except Exception as e:

  #簡単なエラー処理
  dt_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))) # 日本時刻
  logname = dt_now.strftime('%Y%m%d%H%M%S')+"_errlog.txt"
  with open(errlog_dir+logname, mode='w') as f:
      f.write(str(e))
