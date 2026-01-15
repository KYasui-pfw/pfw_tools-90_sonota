import os
import glob
import win32com.client
from openpyxl import load_workbook

# ===== 設定 =====
b_book_name = "B.xlsm"
sheet_b = "B1"

cell_key_b = "C1"
output_cell_1 = "E1"
output_cell_2 = "E2"

sheet_a = "INPUT"
header_row_a = 8
search_folder = os.path.join(os.path.expanduser("~"), "Desktop")
# =================

# --- 起動中のExcelに接続 ---
excel = win32com.client.GetActiveObject("Excel.Application")

# --- 開いているBブック取得 ---
wb_b = None
for wb in excel.Workbooks:
    if wb.Name == b_book_name:
        wb_b = wb
        break

if wb_b is None:
    raise SystemExit("B.xlsx がExcelで開かれていません")

ws_b = wb_b.Worksheets(sheet_b)

# ① B!C1 の値取得
full_text = ws_b.Range(cell_key_b).Value
if not full_text or len(str(full_text)) < 5:
    raise SystemExit("B!C1 の文字列が不足しています")
#print(f"B!C1 の値: {full_text}")
key_5 = str(full_text)[:5]
#print(f"{key_5}")
#print("検索フォルダ:", search_folder)
# --- Aファイル検索（非起動） ---
pattern = os.path.join(search_folder, f"INPUT*{key_5}*")
matched_files = glob.glob(pattern)

if not matched_files:
    raise SystemExit("条件に一致するExcelファイルAが見つかりません")

file_a = matched_files[0]

# --- Excel A を非起動で開く ---
wb_a = load_workbook(file_a, data_only=True)
ws_a = wb_a[sheet_a]

# ② A!8行目で B!C1 と一致する列を検索
found_col = None
for col in range(1, ws_a.max_column + 1):
    value = ws_a.cell(row=header_row_a, column=col).value
    if value == full_text:
        found_col = col
        break

if not found_col:
    wb_a.close()
    raise SystemExit("Aシートに一致する列がありません")

# ③ 一致列の2行目・3行目を B に書き込み（保存はしない）
ws_b.Range(output_cell_1).Value = ws_a.cell(row=2, column=found_col).value
ws_b.Range(output_cell_2).Value = ws_a.cell(row=3, column=found_col).value

# Aのみクローズ（Bは開いたまま）
wb_a.close()