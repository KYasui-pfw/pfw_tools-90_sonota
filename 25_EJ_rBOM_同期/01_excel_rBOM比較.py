"""
EJ-rBOM同期ツール: Excelからマッピング済みデータを抽出し、D3340と結合

処理内容:
1. \\fsrv24\rBOM\発注情報12月EJとrBOM.xlsx をコピー
2. F列(rBOM発注番号)とG列(発注番号)の両方が入力されている行を抽出
3. D3340テーブルとINNER JOIN (NOTE = G列)
4. 結果をCSV出力
"""

import os
import shutil
import pandas as pd
import requests
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 設定
SOURCE_EXCEL = r"\\fsrv24\rBOM\発注情報12月EJとrBOM.xlsx"
OUTPUT_DIR = Path(__file__).parent / "01_excel_rBOM比較"
API_URL = "http://pfw-api/query"
API_KEY = "oG5^Ls%#20yq"

# 前工程横展開CSVファイル
MAEKOTEI_CSV_FILES = [
    r"\\172.17.107.102\Purchase\EJ前工程\前工程横展開.csv",
    r"\\172.17.107.102\Purchase\EJ前工程\前工程横展開(I).csv",
    r"\\172.17.107.102\Purchase\EJ前工程\前工程横展開(C).csv",
]
# 取得する項目
MAEKOTEI_COLUMNS = ["完成部番", "前工程1", "前工程2", "前工程3", "前工程4", "前工程5",
                    "前工程6", "前工程7", "前工程8", "前工程9", "前工程10"]

# D3340から取得するカラム
D3340_COLUMNS = ["PONO", "LINENO", "SEINO", "HMCD", "HMNM", "DRVDT", "THQTY", "CSBCD", "NOTE"]


def copy_and_merge_maekotei_csv():
    """前工程横展開CSVファイルをコピー・縦結合してCSV出力"""
    print("\n" + "=" * 60)
    print("前工程横展開CSV処理")
    print("=" * 60)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    all_dfs = []

    for src_path in MAEKOTEI_CSV_FILES:
        filename = Path(src_path).name
        dest_path = OUTPUT_DIR / filename

        print(f"\nコピー元: {src_path}")
        print(f"コピー先: {dest_path}")

        try:
            shutil.copy2(src_path, dest_path)
            print(f"コピー完了: {filename}")

            # CSVを読み込み（Shift_JIS想定）
            try:
                df = pd.read_csv(dest_path, encoding='cp932')
            except UnicodeDecodeError:
                df = pd.read_csv(dest_path, encoding='utf-8')

            print(f"読み込み行数: {len(df)}")

            # 指定列のみ抽出（存在する列のみ）
            existing_cols = [col for col in MAEKOTEI_COLUMNS if col in df.columns]
            if existing_cols:
                df_selected = df[existing_cols].copy()
                all_dfs.append(df_selected)
                print(f"抽出列: {existing_cols}")
            else:
                print(f"警告: 指定列が見つかりません。列名: {df.columns.tolist()}")

        except FileNotFoundError:
            print(f"ファイルが見つかりません: {src_path}")
        except Exception as e:
            print(f"エラー: {e}")

    # 縦結合
    if all_dfs:
        df_merged = pd.concat(all_dfs, ignore_index=True)
        print(f"\n縦結合結果: {len(df_merged)}行")

        # CSV出力
        output_path = OUTPUT_DIR / "前工程横展開_結合.csv"
        df_merged.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"CSV出力完了: {output_path}")

        return output_path, df_merged
    else:
        print("結合するデータがありません")
        return None, pd.DataFrame()


def copy_excel_file():
    """Excelファイルをコピー（上書き）、サーバーファイルのタイムスタンプも返す"""
    import os

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    dest_path = OUTPUT_DIR / "発注情報.xlsx"

    print(f"コピー元: {SOURCE_EXCEL}")
    print(f"コピー先: {dest_path}")

    # 1. サーバーファイルのタイムスタンプを記録（書き戻し判定用）
    server_mtime = None
    try:
        server_mtime = os.path.getmtime(SOURCE_EXCEL)
    except (FileNotFoundError, PermissionError, OSError):
        pass  # タイムスタンプ取得失敗は無視

    # 2. サーバーからファイルをコピー（エラーでも続行）
    try:
        shutil.copy2(SOURCE_EXCEL, dest_path)
        print("Excelファイルのコピー完了")
    except PermissionError:
        print(f"[WARN] サーバーファイルが使用中のためコピーできません: {SOURCE_EXCEL}")
        if dest_path.exists():
            print(f"       ローカルファイルを使用: {dest_path}")
        else:
            raise FileNotFoundError(f"ローカルファイルもありません: {dest_path}")

    return dest_path, server_mtime


def copy_to_server(local_path, server_path: str, original_mtime):
    """処理済みExcelをサーバーに書き戻し（タイムスタンプチェック付き）

    Args:
        local_path: ローカルの処理済みファイル
        server_path: サーバーのファイルパス
        original_mtime: コピー前に記録したサーバーファイルのタイムスタンプ（Noneの場合はスキップ）
    """
    import os

    print("\n" + "=" * 60)
    print("サーバーへExcel書き戻し")
    print("=" * 60)

    # タイムスタンプが取得できなかった場合はスキップ
    if original_mtime is None:
        print("[SKIP] サーバーへの書き戻しをスキップ（タイムスタンプ取得不可）")
        return

    try:
        # 現在のサーバーファイルのタイムスタンプを確認
        current_mtime = os.path.getmtime(server_path)

        if current_mtime != original_mtime:
            # タイムスタンプが変更されている = 誰かが更新した
            original_time = datetime.fromtimestamp(original_mtime).strftime('%Y-%m-%d %H:%M:%S')
            current_time = datetime.fromtimestamp(current_mtime).strftime('%Y-%m-%d %H:%M:%S')
            print(f"[SKIP] サーバーファイルが更新されています（上書きしません）")
            print(f"       コピー時: {original_time}")
            print(f"       現在:     {current_time}")
            return

        # 上書きコピー
        shutil.copy2(local_path, server_path)
        print(f"[OK] サーバーに書き戻し完了: {server_path}")

    except PermissionError:
        print(f"[SKIP] ファイルが編集中のため上書きできません: {server_path}")
    except Exception as e:
        print(f"[ERROR] サーバーへの書き戻しに失敗しました: {e}")


def extract_mapped_rows(excel_path):
    """F列(rBOM発注番号)とH列(発注番号)の両方が入力されている行を抽出"""
    print(f"\nExcel読み込み: {excel_path}")

    df = pd.read_excel(excel_path, engine='openpyxl')
    print(f"全行数: {len(df)}")
    print(f"列名: {df.columns.tolist()}")

    # F列(index 5)とH列(index 7)を取得
    col_f = df.columns[5]  # rBOM発注番号
    col_h = df.columns[7]  # 発注番号

    print(f"F列(rBOM発注番号): {col_f}")
    print(f"H列(発注番号): {col_h}")

    # 両方が入力されている行を抽出
    df_filtered = df[df[col_f].notna() & df[col_h].notna()].copy()
    print(f"F列・H列両方入力済み: {len(df_filtered)}行")

    # H列を文字列に変換（JOINのため）
    df_filtered['H列_発注番号'] = df_filtered[col_h].astype(str)

    return df_filtered, col_h


def fetch_d3340_data(order_numbers):
    """D3340テーブルからデータを取得（FastAPI経由）"""
    print(f"\nD3340データ取得中... (対象: {len(order_numbers)}件)")

    headers = {
        "X-API-KEY": API_KEY,
        "accept": "application/json",
        "Content-Type": "application/json"
    }

    # NOTE列がE+数字で始まるデータを全て取得し、Python側でマッチング
    # （NOTE列は「E9057565/植木依頼/...」のような形式のため）
    payload = {
        "table": "D3340",
        "columns": D3340_COLUMNS,
        "where": {
            "NOTE": {"like": "E%"}
        },
        "limit": 10000
    }

    try:
        response = requests.post(API_URL, json=payload, headers=headers, timeout=120)
        response.raise_for_status()

        data = response.json()
        df = pd.DataFrame(data.get("rows", []))
        print(f"D3340から取得（E%）: {len(df)}行")

        if df.empty:
            return df

        # NOTE列から発注番号部分を抽出（スラッシュの前の部分）
        df['発注番号_抽出'] = df['NOTE'].apply(lambda x: x.split('/')[0] if pd.notna(x) and '/' in str(x) else x)

        # 指定された発注番号でフィルタリング
        df_filtered = df[df['発注番号_抽出'].isin(order_numbers)]
        print(f"マッチした行: {len(df_filtered)}行")

        return df_filtered

    except requests.exceptions.RequestException as e:
        print(f"API接続エラー: {e}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"レスポンス: {e.response.text[:500]}")
        print("FastAPIサーバーが起動しているか確認してください")
        return pd.DataFrame()


def join_and_export(df_excel, df_d3340, col_h, suffix=""):
    """ExcelデータとD3340をJOINしてCSV出力"""
    print(f"\nINNER JOIN実行: Excel.{col_h} = D3340.発注番号_抽出 {suffix}")

    if df_d3340.empty:
        print("D3340データが空です")
        df_merged = pd.DataFrame()
    else:
        # JOIN（発注番号_抽出列を使用）
        df_merged = pd.merge(
            df_excel,
            df_d3340,
            left_on='H列_発注番号',
            right_on='発注番号_抽出',
            how='inner'
        )

    print(f"JOIN結果: {len(df_merged)}行")

    # CSV出力（0件でも出力、上書き）
    filename = f"結合結果{suffix}.csv"
    output_path = OUTPUT_DIR / filename

    df_merged.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"CSV出力完了: {output_path}")

    return output_path


def update_excel_with_results(excel_path):
    """CSVの結合結果をExcelに反映し、発注情報_自動追記.xlsxとして保存

    NOTE欄（D3340.NOTE）とH列（EJ発注番号）が突合できた行の処理:
    - F列空欄 or 既存値とPONOが一致 → F列を水色にして値を設定
    - 既存値とPONOが不一致 → F列を赤色にして上書きしない
    """
    print("\n" + "=" * 60)
    print("Excel自動追記処理")
    print("=" * 60)

    # 色定義
    FILL_LIGHT_BLUE = PatternFill(start_color="E0F0FF", end_color="E0F0FF", fill_type="solid")  # 薄い水色
    FILL_RED = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")  # 赤

    # Excelを読み込み（openpyxl）
    wb = load_workbook(excel_path)
    ws = wb.active

    # ヘッダー行を取得してH列（発注番号）のインデックスを特定
    header_row = 1
    col_f_idx = 6  # F列 = rBOM発注番号
    col_g_idx = 7  # G列 = 行番号
    col_h_idx = 8  # H列 = 発注番号

    # H列の値から行番号へのマッピングを作成
    order_to_row = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        h_value = ws.cell(row=row_idx, column=col_h_idx).value
        if h_value is not None:
            order_to_row[str(h_value)] = row_idx

    print(f"Excel内の発注番号（H列）: {len(order_to_row)}件")

    # 両方のCSVを処理
    csv_files = [
        OUTPUT_DIR / "結合結果_FH両方入力済.csv",
        OUTPUT_DIR / "結合結果_H列のみ入力済.csv"
    ]

    updated_count = 0
    mismatch_count = 0
    matched_rows = set()  # 一致済み（水色）の行を追跡

    for csv_path in csv_files:
        if not csv_path.exists():
            print(f"CSVファイルが見つかりません: {csv_path}")
            continue

        print(f"\nCSV読み込み: {csv_path.name}")
        df_csv = pd.read_csv(csv_path, encoding='utf-8-sig')

        if df_csv.empty:
            print("  → 0件（スキップ）")
            continue

        print(f"  → {len(df_csv)}件")

        # 各行を処理
        for _, row in df_csv.iterrows():
            order_no = str(row.get('H列_発注番号', ''))
            pono = row.get('PONO', '')
            lineno = row.get('LINENO', '')

            if order_no not in order_to_row:
                continue

            excel_row = order_to_row[order_no]

            # 既に一致済み（水色）の行はスキップ
            if excel_row in matched_rows:
                continue

            # F列（rBOM発注番号）の処理
            cell_f = ws.cell(row=excel_row, column=col_f_idx)
            existing_f = cell_f.value
            existing_f_str = str(existing_f).strip() if existing_f is not None else ''

            # PONOを0埋め9桁に変換
            if pd.notna(pono) and str(pono) != '':
                pono_str = str(pono).strip()
                # 数値のみ抽出して0埋め
                pono_digits = ''.join(c for c in pono_str if c.isdigit())
                new_f = pono_digits.zfill(9) if pono_digits else ''
            else:
                new_f = ''

            # 既存値が数値のみかチェック
            existing_is_numeric = is_numeric_string(existing_f_str)

            if existing_is_numeric and existing_f_str != new_f:
                # 既存値が数値で、かつ不一致 → 赤色にして上書きしない
                cell_f.fill = FILL_RED
                mismatch_count += 1
            else:
                # それ以外（空欄/テキスト/一致）→ 水色にして値を設定
                cell_f.value = new_f
                cell_f.fill = FILL_LIGHT_BLUE
                updated_count += 1
                matched_rows.add(excel_row)  # 一致済みとしてマーク

                # G列（行番号）の更新 - 数値
                cell_g = ws.cell(row=excel_row, column=col_g_idx)
                if pd.notna(lineno) and str(lineno) != '':
                    try:
                        new_g = int(float(lineno))
                    except (ValueError, TypeError):
                        new_g = lineno
                else:
                    new_g = ''
                cell_g.value = new_g

    # 保存
    output_excel = OUTPUT_DIR / "発注情報_自動追記.xlsx"
    wb.save(output_excel)
    print(f"\n突合成功・一致（水色）: {updated_count}件")
    print(f"突合成功・不一致（赤）: {mismatch_count}件")
    print(f"Excel保存完了: {output_excel}")

    return output_excel


def update_excel_from_rbom(excel_path, maekotei_csv_path=None):
    """発注情報_自動追記.xlsxを読み込み、D3330/D3340からデータを取得してB,C,D,CI〜CMを更新

    Args:
        excel_path: 更新対象のExcelファイルパス
        maekotei_csv_path: 前工程横展開結合CSVファイルパス（A列突合用）
    """
    print("\n" + "=" * 60)
    print("rBOMデータ取得・Excel更新処理")
    print("=" * 60)

    # Excelを読み込み
    wb = load_workbook(excel_path)
    ws = wb.active

    # 列インデックス（1始まり）
    col_a_idx = 1    # A列 = 工程（MK020.KTCD）
    col_b_idx = 2    # B列 = rBOM更新者
    col_c_idx = 3    # C列 = 印刷区分
    col_d_idx = 4    # D列 = 状態（STATUS）
    col_e_idx = 5    # E列 = 納期（DRVDT）
    col_f_idx = 6    # F列 = rBOM発注番号(PONO)
    col_g_idx = 7    # G列 = 行番号(LINENO)
    col_l_idx = 12   # L列 = 品目番号（MK020.NOTEとの突合キー）
    col_cj_idx = 88  # CJ列 = rBOM担当者
    col_ck_idx = 89  # CK列 = rBOM取引先
    col_cl_idx = 90  # CL列 = rBOM品目CD
    col_cm_idx = 91  # CM列 = rBOM発注数
    col_cn_idx = 92  # CN列 = rBOM単価

    # ヘッダー行（1行目）が空欄の場合、項目名を設定
    header_map = {
        col_b_idx: "rBOM更新者",
        col_c_idx: "印刷区分",
        col_d_idx: "状態",
        col_e_idx: "rBOM納期",
        col_cj_idx: "rBOM担当者",
        col_ck_idx: "rBOM取引先",
        col_cl_idx: "rBOM品目CD",
        col_cm_idx: "rBOM発注数",
        col_cn_idx: "rBOM単価",
    }
    for col_idx, header_name in header_map.items():
        cell = ws.cell(row=1, column=col_idx)
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = header_name

    # F列とG列が入力されている行を収集（D3330/D3340更新対象）
    target_rows = []
    # L列が入力されている全行を収集（MK020更新対象）
    all_rows_with_l = []

    for row_idx in range(2, ws.max_row + 1):
        f_value = ws.cell(row=row_idx, column=col_f_idx).value
        g_value = ws.cell(row=row_idx, column=col_g_idx).value
        l_value = ws.cell(row=row_idx, column=col_l_idx).value

        # L列が入力されている行を全て収集（MK020突合用）
        if l_value is not None and str(l_value).strip() != '':
            all_rows_with_l.append({
                'row_idx': row_idx,
                'l_value': str(l_value).strip()
            })

        # F列・G列が入力されている行を収集（D3330/D3340更新用）
        if f_value is not None and g_value is not None:
            target_rows.append({
                'row_idx': row_idx,
                'pono': str(f_value).strip(),
                'lineno': int(g_value) if isinstance(g_value, (int, float)) else g_value,
                'l_value': str(l_value).strip() if l_value is not None else ''
            })

    print(f"対象行数（F列・G列入力済み）: {len(target_rows)}件")
    print(f"対象行数（L列入力済み・MK020突合用）: {len(all_rows_with_l)}件")

    # MK020からデータ取得（L列の値で検索 - 全行対象）
    l_value_list = list(set([r['l_value'] for r in all_rows_with_l]))
    print(f"\nMK020データ取得中... (対象L列ユニーク: {len(l_value_list)}件)")
    df_mk020 = fetch_mk020_by_note(l_value_list)
    print(f"MK020から取得: {len(df_mk020)}行")

    # D3330/D3340はF列・G列入力済み行のみ対象
    df_d3330 = pd.DataFrame()
    df_d3340 = pd.DataFrame()
    if target_rows:
        # ユニークなPONOリストを作成
        pono_list = list(set([r['pono'] for r in target_rows]))
        print(f"\nユニークなPONO: {len(pono_list)}件")

        # D3330からデータ取得（PONOで検索）
        print("D3330データ取得中...")
        df_d3330 = fetch_d3330_by_pono(pono_list)
        print(f"D3330から取得: {len(df_d3330)}行")

        # D3340からデータ取得（PONO+LINENOで検索）
        print("D3340データ取得中...")
        df_d3340 = fetch_d3340_by_pono_lineno(target_rows)
        print(f"D3340から取得: {len(df_d3340)}行")

    # D3330をPONOでインデックス化
    d3330_dict = {}
    if not df_d3330.empty:
        for _, row in df_d3330.iterrows():
            pono = str(row['PONO']).strip()
            d3330_dict[pono] = row

    # D3340をPONO+LINENOでインデックス化
    d3340_dict = {}
    if not df_d3340.empty:
        for _, row in df_d3340.iterrows():
            key = (str(row['PONO']).strip(), int(row['LINENO']))
            d3340_dict[key] = row

    # MK020をNOTEでインデックス化
    mk020_dict = {}
    if not df_mk020.empty:
        for _, row in df_mk020.iterrows():
            note = str(row['NOTE']).strip()
            mk020_dict[note] = row

    # 色定義
    FILL_LIGHT_YELLOW = PatternFill(start_color="FFFFEE", end_color="FFFFEE", fill_type="solid")  # ごく薄い黄色
    FILL_LIGHT_PINK = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")  # 薄いピンク

    # 前工程横展開データの処理（A列突合用辞書を作成）
    maekotei_dict = process_maekotei_for_matching(maekotei_csv_path)

    # A列の更新（L列入力済み全行対象）
    # 処理順序: 前工程突合 → MK020突合（MK020が優先なので後から上書き）
    maekotei_updated_count = 0
    mk020_updated_count = 0

    # 1. 前工程突合（先に処理）
    for row_data in all_rows_with_l:
        row_idx = row_data['row_idx']
        l_value = row_data['l_value']
        maekotei_code = maekotei_dict.get(l_value)
        if maekotei_code is not None:
            cell_a = ws.cell(row=row_idx, column=col_a_idx)
            cell_a.value = maekotei_code
            cell_a.fill = FILL_LIGHT_PINK
            maekotei_updated_count += 1

    print(f"\n前工程突合によるA列更新: {maekotei_updated_count}件")

    # 2. MK020突合（後から処理して上書き - MK020が優先）
    for row_data in all_rows_with_l:
        row_idx = row_data['row_idx']
        l_value = row_data['l_value']
        mk020_row = mk020_dict.get(l_value)
        if mk020_row is not None:
            cell_a = ws.cell(row=row_idx, column=col_a_idx)
            cell_a.value = mk020_row.get('KTCD', '')
            cell_a.fill = FILL_LIGHT_YELLOW
            mk020_updated_count += 1

    print(f"MK020突合によるA列更新: {mk020_updated_count}件")

    # B〜CM列の更新（F列・G列入力済み行対象 - D3330/D3340）
    updated_count = 0
    for target in target_rows:
        row_idx = target['row_idx']
        pono = target['pono']
        lineno = target['lineno']

        # D3330データ取得
        d3330_row = d3330_dict.get(pono)
        # D3340データ取得
        d3340_row = d3340_dict.get((pono, lineno))

        # B列: D3330.UPDTID
        if d3330_row is not None:
            ws.cell(row=row_idx, column=col_b_idx).value = d3330_row.get('UPDTID', '')

        # C列: D3330.PRNKBN（1→未発行、2→発行済、他→空欄）
        if d3330_row is not None:
            prnkbn = d3330_row.get('PRNKBN')
            if prnkbn == 1 or prnkbn == '1':
                ws.cell(row=row_idx, column=col_c_idx).value = "未発行"
            elif prnkbn == 2 or prnkbn == '2':
                ws.cell(row=row_idx, column=col_c_idx).value = "発行済"
            else:
                ws.cell(row=row_idx, column=col_c_idx).value = ""

        # D列: D3340.STATUS（2=承認済、3=一部完納、4=完納、8=強制完納、それ以外=空欄）
        if d3340_row is not None:
            status = d3340_row.get('STATUS')
            status_map = {
                2: "承認済", '2': "承認済",
                3: "一部完納", '3': "一部完納",
                4: "完納", '4': "完納",
                8: "強制完納", '8': "強制完納",
            }
            ws.cell(row=row_idx, column=col_d_idx).value = status_map.get(status, "")

        # E列: D3340.DRVDT
        if d3340_row is not None:
            ws.cell(row=row_idx, column=col_e_idx).value = d3340_row.get('DRVDT', '')

        # CJ列: D3330.TANCD（rBOM担当者）
        if d3330_row is not None:
            ws.cell(row=row_idx, column=col_cj_idx).value = d3330_row.get('TANCD', '')

        # CK列: D3330.SRCD（rBOM取引先）
        if d3330_row is not None:
            ws.cell(row=row_idx, column=col_ck_idx).value = d3330_row.get('SRCD', '')

        # CL列: D3340.HMCD（rBOM品目CD）
        if d3340_row is not None:
            ws.cell(row=row_idx, column=col_cl_idx).value = d3340_row.get('HMCD', '')

        # CM列: D3340.THQTY（rBOM発注数）
        if d3340_row is not None:
            ws.cell(row=row_idx, column=col_cm_idx).value = d3340_row.get('THQTY', '')

        # CN列: D3340.PRICE（rBOM単価）
        if d3340_row is not None:
            ws.cell(row=row_idx, column=col_cn_idx).value = d3340_row.get('PRICE', '')

        updated_count += 1

    # 保存
    wb.save(excel_path)
    print(f"\n更新行数: {updated_count}件")
    print(f"Excel更新完了: {excel_path}")

    return excel_path


def process_maekotei_for_matching(csv_path):
    """前工程横展開CSVを処理し、L列突合用の辞書を作成

    各行の一番左（完成部番）のみを削除し、
    残った項目をキーとして、その項目自身の「最初のハイフンまでの文字列」を値とする辞書を作成
    ただし、ハイフン前が「MA」の場合は上書き対象外（辞書に登録しない）
    例: L1-HUMHJ → L1, LA-1831-406AA30-MC → LA, MA-xxx → (対象外)
    """
    if csv_path is None or not Path(csv_path).exists():
        return {}

    print("\n前工程横展開データの突合準備...")
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    print(f"読み込み行数: {len(df)}")

    # 結果辞書: キー = 項目の値, 値 = その項目の最初のハイフンまでの文字列
    maekotei_dict = {}
    ma_skip_count = 0

    for idx, row in df.iterrows():
        # 非空欄の値を取得（列順序を維持）
        non_empty_values = []
        for col in df.columns:
            val = row[col]
            if pd.notna(val) and str(val).strip() != '':
                non_empty_values.append(str(val).strip())

        # 1つ以下なら対象項目がない（左端削除後に残らない）
        if len(non_empty_values) <= 1:
            continue

        # 左端（完成部番）のみを削除し、残り全てを対象とする
        target_values = non_empty_values[1:]

        # 対象項目をキーとして登録
        # 値はその項目自身の「最初のハイフンまでの文字列」
        for target_val in target_values:
            if target_val not in maekotei_dict:
                # 最初のハイフンまでの文字を抽出
                if '-' in target_val:
                    code_part = target_val.split('-')[0]
                else:
                    code_part = target_val

                # MAの場合は上書き対象外（辞書に登録しない）
                if code_part == 'MA':
                    ma_skip_count += 1
                    continue

                maekotei_dict[target_val] = code_part

    print(f"前工程突合用辞書エントリ数: {len(maekotei_dict)}件（MA除外: {ma_skip_count}件）")
    return maekotei_dict


def fetch_m0820_by_hmcd(hmcd_list):
    """M0820テーブルからHMCD+KNRBUCD=100でデータを取得（バッチ処理）"""
    if not hmcd_list:
        return pd.DataFrame()

    headers = {
        "X-API-KEY": API_KEY,
        "accept": "application/json",
        "Content-Type": "application/json"
    }

    M0820_COLUMNS = ["HMCD", "KNRBUCD", "SRCD"]

    # バッチサイズ（IN句の制限を避けるため500件ずつ）
    BATCH_SIZE = 500
    all_rows = []

    for i in range(0, len(hmcd_list), BATCH_SIZE):
        batch = hmcd_list[i:i + BATCH_SIZE]

        payload = {
            "table": "M0820",
            "columns": M0820_COLUMNS,
            "where": {
                "and": [
                    {"HMCD": {"in": batch}},
                    {"KNRBUCD": {"eq": 100}}
                ]
            },
            "limit": 10000
        }

        try:
            response = requests.post(API_URL, json=payload, headers=headers, timeout=120)
            response.raise_for_status()
            data = response.json()
            all_rows.extend(data.get("rows", []))
        except requests.exceptions.RequestException as e:
            print(f"M0820 API接続エラー (batch {i//BATCH_SIZE + 1}): {e}")
            continue

    return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()


def check_vendor_code_mismatch(excel_path):
    """品目番号をキーにM0820と突合し、取引先コードが一致しない場合は赤背景にする"""
    print("\n" + "=" * 60)
    print("M0820突合・取引先コードチェック")
    print("=" * 60)

    # 色定義
    FILL_RED = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")  # 赤

    # Excelを読み込み
    wb = load_workbook(excel_path)
    ws = wb.active

    # 列インデックス（1始まり）
    col_k_idx = 11   # K列 = 仕入先コード（取引先コード）
    col_l_idx = 12   # L列 = 品目番号

    # L列が入力されている行を収集
    rows_with_l = []
    for row_idx in range(2, ws.max_row + 1):
        l_value = ws.cell(row=row_idx, column=col_l_idx).value
        k_value = ws.cell(row=row_idx, column=col_k_idx).value
        if l_value is not None and str(l_value).strip() != '':
            rows_with_l.append({
                'row_idx': row_idx,
                'l_value': str(l_value).strip(),
                'k_value': str(k_value).strip() if k_value is not None else ''
            })

    print(f"対象行数（L列入力済み）: {len(rows_with_l)}件")

    if not rows_with_l:
        print("チェック対象がありません")
        wb.save(excel_path)
        return excel_path

    # ユニークな品目番号リスト
    hmcd_list = list(set([r['l_value'] for r in rows_with_l]))
    print(f"ユニークな品目番号: {len(hmcd_list)}件")

    # M0820からデータ取得
    print("M0820データ取得中...")
    df_m0820 = fetch_m0820_by_hmcd(hmcd_list)
    print(f"M0820から取得: {len(df_m0820)}行")

    # M0820をHMCDでインデックス化
    m0820_dict = {}
    if not df_m0820.empty:
        for _, row in df_m0820.iterrows():
            hmcd = str(row['HMCD']).strip()
            srcd = str(row['SRCD']).strip() if pd.notna(row['SRCD']) else ''
            m0820_dict[hmcd] = srcd

    # チェックして不一致の場合は赤背景
    mismatch_count = 0
    match_count = 0
    no_m0820_count = 0

    for row_data in rows_with_l:
        row_idx = row_data['row_idx']
        l_value = row_data['l_value']
        k_value = row_data['k_value']

        # M0820から取引先コードを取得
        m0820_srcd = m0820_dict.get(l_value)

        if m0820_srcd is None:
            # M0820にデータがない場合はスキップ
            no_m0820_count += 1
            continue

        # 比較（文字列として比較）
        if k_value != m0820_srcd:
            # 不一致 → K列を赤背景に
            cell_k = ws.cell(row=row_idx, column=col_k_idx)
            cell_k.fill = FILL_RED
            mismatch_count += 1
        else:
            match_count += 1

    # 保存
    wb.save(excel_path)
    print(f"\n一致: {match_count}件")
    print(f"不一致（赤背景）: {mismatch_count}件")
    print(f"M0820データなし: {no_m0820_count}件")
    print(f"Excel保存完了: {excel_path}")

    return excel_path


def fetch_mk020_by_note(note_list):
    """MK020テーブルからNOTEでデータを取得（バッチ処理）"""
    if not note_list:
        return pd.DataFrame()

    headers = {
        "X-API-KEY": API_KEY,
        "accept": "application/json",
        "Content-Type": "application/json"
    }

    MK020_COLUMNS = ["NOTE", "KTCD"]

    # バッチサイズ（IN句の制限を避けるため500件ずつ）
    BATCH_SIZE = 500
    all_rows = []

    for i in range(0, len(note_list), BATCH_SIZE):
        batch = note_list[i:i + BATCH_SIZE]

        payload = {
            "table": "MK020",
            "columns": MK020_COLUMNS,
            "where": {
                "NOTE": {"in": batch}
            },
            "limit": 10000
        }

        try:
            response = requests.post(API_URL, json=payload, headers=headers, timeout=120)
            response.raise_for_status()
            data = response.json()
            all_rows.extend(data.get("rows", []))
        except requests.exceptions.RequestException as e:
            print(f"MK020 API接続エラー (batch {i//BATCH_SIZE + 1}): {e}")
            continue

    return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()


def fetch_d3330_by_pono(pono_list):
    """D3330テーブルからPONOでデータを取得（バッチ処理）"""
    if not pono_list:
        return pd.DataFrame()

    headers = {
        "X-API-KEY": API_KEY,
        "accept": "application/json",
        "Content-Type": "application/json"
    }

    D3330_COLUMNS = ["PONO", "UPDTID", "PRNKBN", "TANCD", "SRCD"]

    # バッチサイズ（IN句の制限を避けるため500件ずつ）
    BATCH_SIZE = 500
    all_rows = []

    for i in range(0, len(pono_list), BATCH_SIZE):
        batch = pono_list[i:i + BATCH_SIZE]

        payload = {
            "table": "D3330",
            "columns": D3330_COLUMNS,
            "where": {
                "PONO": {"in": batch}
            },
            "limit": 10000
        }

        try:
            response = requests.post(API_URL, json=payload, headers=headers, timeout=120)
            response.raise_for_status()
            data = response.json()
            all_rows.extend(data.get("rows", []))
        except requests.exceptions.RequestException as e:
            print(f"D3330 API接続エラー (batch {i//BATCH_SIZE + 1}): {e}")
            continue

    return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()


def fetch_d3340_by_pono_lineno(target_rows):
    """D3340テーブルからPONO+LINENOでデータを取得（バッチ処理）"""
    if not target_rows:
        return pd.DataFrame()

    headers = {
        "X-API-KEY": API_KEY,
        "accept": "application/json",
        "Content-Type": "application/json"
    }

    # ユニークなPONOで取得し、Python側でLINENOフィルタリング
    pono_list = list(set([r['pono'] for r in target_rows]))

    D3340_COLUMNS_DETAIL = ["PONO", "LINENO", "STATUS", "DRVDT", "HMCD", "THQTY", "PRICE"]

    # バッチサイズ（IN句の制限を避けるため500件ずつ）
    BATCH_SIZE = 500
    all_rows = []

    for i in range(0, len(pono_list), BATCH_SIZE):
        batch = pono_list[i:i + BATCH_SIZE]

        payload = {
            "table": "D3340",
            "columns": D3340_COLUMNS_DETAIL,
            "where": {
                "PONO": {"in": batch}
            },
            "limit": 10000
        }

        try:
            response = requests.post(API_URL, json=payload, headers=headers, timeout=120)
            response.raise_for_status()
            data = response.json()
            all_rows.extend(data.get("rows", []))
        except requests.exceptions.RequestException as e:
            print(f"D3340 API接続エラー (batch {i//BATCH_SIZE + 1}): {e}")
            continue

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)

    # PONO+LINENOでフィルタリング
    target_keys = set((r['pono'], r['lineno']) for r in target_rows)
    df['key'] = df.apply(lambda x: (str(x['PONO']), x['LINENO']), axis=1)
    df_filtered = df[df['key'].isin(target_keys)]

    return df_filtered.drop(columns=['key'])


def extract_h_only_rows(excel_path):
    """H列(発注番号)のみが入力されている行を抽出"""
    print(f"\nExcel読み込み（H列のみ）: {excel_path}")

    df = pd.read_excel(excel_path, engine='openpyxl')

    col_h = df.columns[7]  # 発注番号

    # H列が入力されている行を抽出（F列は問わない）
    df_filtered = df[df[col_h].notna()].copy()
    print(f"H列入力済み: {len(df_filtered)}行")

    # H列を文字列に変換（JOINのため）
    df_filtered['H列_発注番号'] = df_filtered[col_h].astype(str)

    return df_filtered, col_h


def is_red_background(cell):
    """セルの背景が赤色かどうかをチェック"""
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = cell.fill.fgColor.rgb
        if len(rgb) >= 6:
            color = rgb[-6:].upper()
            if color == 'FF6666':
                return True
    return False


def is_numeric_string(value):
    """文字列が数字のみで構成されているかチェック（"00000789"等はTrue、"000000011a"等はFalse）"""
    if value is None:
        return False
    s = str(value).strip()
    if s == '':
        return False
    return s.isdigit()


def mark_auto_input_exclusions(excel_path):
    """自動インプット対象外の行を判定し、F列に色付けして理由を記述

    除外条件:
    - A列（工程）が空欄でない → 灰色 + 工程発注
    - L列（品目番号）が空欄 → 灰色 + 品目番号なし
    - K列（仕入先コード）が赤背景 → 灰色 + 取引先不一致(赤)
    - K列がCA/PT → 灰色 + 取引先CA/PT
    - CC列（勘定科目）が12/33でない → 薄いピンク + rBOMで対応する発注の入力をお願いします

    ※F列に数字のみ（PONOなど）が入っている場合はスキップ（上書きしない）
    """
    print("\n" + "=" * 60)
    print("自動インプット除外判定・マーク処理")
    print("=" * 60)

    # 色定義
    FILL_LIGHT_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 薄い灰色
    FILL_LIGHT_PINK = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 薄いピンク

    # Excelを読み込み
    wb = load_workbook(excel_path)
    ws = wb.active

    # 列インデックス（1始まり）
    col_a_idx = 1    # A列 = 工程
    col_f_idx = 6    # F列 = rBOM発注番号（除外理由を追記）
    col_g_idx = 7    # G列 = 行番号（LINENO）
    col_k_idx = 11   # K列 = 仕入先コード（取引先コード）
    col_l_idx = 12   # L列 = 品目番号
    col_bq_idx = 69  # BQ列 = 注文取消伝票発行フラグ
    col_cc_idx = 81  # CC列 = 勘定科目

    # 統計
    stats = {
        'total': 0,
        'excluded_gray': 0,  # 灰色（工程発注/品目番号なし/取引先不一致/CA・PT）
        'excluded_pink': 0,  # 薄いピンク（rBOM対応依頼）
        'target': 0,         # 自動インプット対象
        'skipped': 0,        # F列に数字があるためスキップ
        'g_has_value': 0,    # G列に値があるためスキップ
        'a_not_empty': 0,
        'l_empty': 0,
        'k_red': 0,          # K列赤背景
        'k_ca_pt': 0,
        'bq_canceled': 0,    # BQ列が0以外（発注取消済み）
        'cc_not_valid': 0,   # 勘定科目が12/33以外
        'f_cleared': 0,      # F列クリア件数
    }

    for row_idx in range(2, ws.max_row + 1):
        stats['total'] += 1

        # 各セルの値を取得
        cell_a = ws.cell(row=row_idx, column=col_a_idx)
        cell_f = ws.cell(row=row_idx, column=col_f_idx)
        cell_g = ws.cell(row=row_idx, column=col_g_idx)
        cell_k = ws.cell(row=row_idx, column=col_k_idx)
        cell_l = ws.cell(row=row_idx, column=col_l_idx)
        cell_bq = ws.cell(row=row_idx, column=col_bq_idx)
        cell_cc = ws.cell(row=row_idx, column=col_cc_idx)

        # F列の現在値を取得
        f_value = str(cell_f.value).strip() if cell_f.value is not None else ''

        # BQ列（注文取消伝票発行フラグ）が0以外の場合は発注取消済み
        bq_value = cell_bq.value
        bq_str = str(bq_value).strip() if bq_value is not None else ''
        is_bq_canceled = bq_str != '' and bq_str != '0'

        if is_bq_canceled:
            stats['bq_canceled'] += 1
            # F列に値があれば維持、なければ「EJ発注取消済み」を設定
            if f_value == '':
                cell_f.value = "EJ発注取消済み"
            # 背景を灰色に
            cell_f.fill = FILL_LIGHT_GRAY
            stats['excluded_gray'] += 1
            continue

        # G列（LINENO）に値がある場合はスキップ（rBOMからデータ取得済みの可能性が高い）
        g_value = cell_g.value
        if g_value is not None and str(g_value).strip() != '':
            stats['g_has_value'] += 1
            continue

        # F列が「rBOMで対応する発注の入力をお願いします」の場合は一旦クリア
        if f_value == 'rBOMで対応する発注の入力をお願いします':
            cell_f.value = None
            cell_f.fill = PatternFill()  # 塗りつぶしもクリア
            f_value = ''
            stats['f_cleared'] += 1

        # F列に数字のみ（PONOなど）が入っている場合はスキップ
        if is_numeric_string(f_value):
            stats['skipped'] += 1
            continue

        a_value = cell_a.value
        k_value = str(cell_k.value).strip() if cell_k.value is not None else ''
        l_value = cell_l.value
        cc_value = cell_cc.value

        # 各条件をチェック
        is_a_not_empty = a_value is not None and str(a_value).strip() != ''
        is_l_empty = l_value is None or str(l_value).strip() == ''
        is_k_red = is_red_background(cell_k)
        is_ca_pt = k_value.upper() in ['CA', 'PT']
        cc_str = str(cc_value).strip() if cc_value is not None else ''
        is_cc_not_valid = cc_str not in ['12', '33']  # 12または33以外はNG

        # 統計カウント
        if is_a_not_empty:
            stats['a_not_empty'] += 1
        if is_l_empty:
            stats['l_empty'] += 1
        if is_k_red:
            stats['k_red'] += 1
        if is_ca_pt:
            stats['k_ca_pt'] += 1
        if is_cc_not_valid:
            stats['cc_not_valid'] += 1

        # 除外理由を判定（優先順位順）- 灰色で表示する条件
        gray_reason = None
        if is_a_not_empty:
            gray_reason = "工程発注"
        elif is_l_empty:
            gray_reason = "品目番号なし"
        elif is_k_red:
            gray_reason = "取引先不一致(赤)"
        elif is_ca_pt:
            gray_reason = "取引先CA/PT"

        if gray_reason:
            # 灰色条件に該当
            if gray_reason in ["工程発注", "品目番号なし", "取引先不一致(赤)"]:
                # これらは薄いピンク + 「rBOMで対応する発注の入力をお願いします」
                cell_f.value = "rBOMで対応する発注の入力をお願いします"
                cell_f.fill = FILL_LIGHT_PINK
                stats['excluded_pink'] += 1
            else:
                # 取引先CA/PT → 灰色背景 + 理由を記述
                cell_f.value = gray_reason
                cell_f.fill = FILL_LIGHT_GRAY
                stats['excluded_gray'] += 1
        elif is_cc_not_valid:
            # 勘定科目12/33以外 → 薄いピンク + rBOMで対応する発注の入力をお願いします
            cell_f.value = "rBOMで対応する発注の入力をお願いします"
            cell_f.fill = FILL_LIGHT_PINK
            stats['excluded_pink'] += 1
        else:
            # 全フィルタを通過 = 自動インプット対象（何もしない）
            stats['target'] += 1

    # 保存
    wb.save(excel_path)

    # 統計表示
    print("\n" + "-" * 40)
    print("自動インプット除外判定結果")
    print("-" * 40)
    print(f"全行数: {stats['total']}行")
    print(f"  G列に値あり（スキップ）: {stats['g_has_value']}行")
    print(f"  F列クリア（再判定）: {stats['f_cleared']}行")
    print(f"  F列に数字あり（スキップ）: {stats['skipped']}行")
    print(f"  除外: EJ発注取消済み(BQ列): {stats['bq_canceled']}行")
    print(f"  除外: 工程発注(A列): {stats['a_not_empty']}行")
    print(f"  除外: 品目番号なし(L列): {stats['l_empty']}行")
    print(f"  除外: 取引先不一致/赤(K列): {stats['k_red']}行")
    print(f"  除外: 取引先CA/PT(K列): {stats['k_ca_pt']}行")
    print(f"  除外: 勘定科目12/33以外(CC列): {stats['cc_not_valid']}行")
    print(f"  → 除外（灰色）: {stats['excluded_gray']}行")
    print(f"  → 除外（薄いピンク）: {stats['excluded_pink']}行")
    print(f"  → 自動インプット対象: {stats['target']}行")
    print(f"Excel保存完了: {excel_path}")

    return excel_path


def main():
    """メイン処理"""
    print("=" * 60)
    print("EJ-rBOM同期ツール: Excel-D3340結合")
    print("=" * 60)

    # 0. 前工程横展開CSVをコピー・縦結合
    maekotei_csv_path, _ = copy_and_merge_maekotei_csv()

    # 1. Excelファイルをコピー（タイムスタンプも取得）
    excel_path, server_mtime = copy_excel_file()

    # 2-A. F列・H列両方入力済みの行を抽出
    df_excel_both, col_h = extract_mapped_rows(excel_path)
    order_numbers_both = df_excel_both['H列_発注番号'].unique().tolist() if not df_excel_both.empty else []
    print(f"\n【パターン1】F列・H列両方入力済み: {len(df_excel_both)}行, 発注番号(ユニーク): {len(order_numbers_both)}件")

    # 2-B. H列のみ入力済みの行を抽出
    df_excel_h_only, _ = extract_h_only_rows(excel_path)
    order_numbers_h_only = df_excel_h_only['H列_発注番号'].unique().tolist() if not df_excel_h_only.empty else []
    print(f"【パターン2】H列入力済み: {len(df_excel_h_only)}行, 発注番号(ユニーク): {len(order_numbers_h_only)}件")

    # 3. D3340からデータ取得（全発注番号を対象）
    all_order_numbers = list(set(order_numbers_both + order_numbers_h_only))
    df_d3340 = fetch_d3340_data(all_order_numbers)

    # 4-A. F列・H列両方入力済み → D3340結合
    print("\n" + "-" * 40)
    print("【パターン1】F列・H列両方入力済み × D3340")
    print("-" * 40)
    join_and_export(df_excel_both, df_d3340, col_h, suffix="_FH両方入力済")

    # 4-B. H列のみ入力済み → D3340結合
    print("\n" + "-" * 40)
    print("【パターン2】H列入力済み × D3340")
    print("-" * 40)
    join_and_export(df_excel_h_only, df_d3340, col_h, suffix="_H列のみ入力済")

    # 5. Excel自動追記
    output_excel = update_excel_with_results(excel_path)

    # 6. rBOMデータ取得・Excel更新（前工程CSVパスを渡す）
    update_excel_from_rbom(output_excel, maekotei_csv_path)

    # 6.5. 取引先コード不一致チェック（K列を赤背景にマーク）
    check_vendor_code_mismatch(output_excel)

    # 7. 自動インプット除外判定・マーク処理
    mark_auto_input_exclusions(output_excel)

    # 8. サーバーへ書き戻し（タイムスタンプチェック付き）
    copy_to_server(output_excel, SOURCE_EXCEL, server_mtime)

    print("\n" + "=" * 60)
    print("処理完了")
    print("=" * 60)


if __name__ == "__main__":
    main()
