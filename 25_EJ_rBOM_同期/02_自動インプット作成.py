"""
EJ-rBOM同期ツール: 自動インプット用CSV作成

処理内容:
1. 発注情報_自動追記.xlsx を読み込み
2. フィルタリング:
   - A列（工程）が空欄である行を抽出
   - F列（rBOM発注番号）が空欄である行を抽出
   - L列（品目番号）が空欄である行を省く
   - K列（取引先コード）が赤色背景のデータを省く
   - K列（取引先コード）がCA/PTのデータを省く
   - CC列（勘定科目）が12または33のデータで絞り込む
   - P列（発注伝票発行日）が指定日のデータで絞り込む
3. 抽出データをExcel出力
4. EJ_〇〇〇〇.csv テンプレートにデータ入力して出力
"""

import os
import shutil
import math
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook

# 設定
INPUT_EXCEL = Path(__file__).parent / "01_excel_rBOM比較" / "発注情報_自動追記.xlsx"
TEMPLATE_CSV = Path(__file__).parent / "02_自動インプット用" / "EJ_〇〇〇〇.csv"
OUTPUT_DIR = Path(__file__).parent / "02_自動インプット用"

# ========== フィルタ条件（変更しやすいように分離） ==========
# 対象とする発注伝票発行日（複数指定可）
TARGET_ISSUE_DATES = [
#    "2025/12/16",
#    "2025/12/17",
#    "2025/12/18",
#    "2025/12/19",
#    "2025/12/20",
#    "2025/12/21",
#    "2025/12/22",
#    "2025/12/23",
#    "2025/12/24",
#    "2025/12/25",
#    "2025/12/26",
#    "2025/12/27",
#    "2025/12/28",
#    "2025/12/29",
#    "2025/12/30",
#    "2025/12/31",
    "2026/01/01",
    "2026/01/02",
    "2026/01/03",
    "2026/01/04",
    "2026/01/05",
    "2026/01/06",
    "2026/01/07",
    "2026/01/08",
    "2026/01/09",
    "2026/01/10",
    "2026/01/11",
    "2026/01/12",
    "2026/01/13"
]
# ===========================================================

# 列インデックス（1始まり）
COL_A = 1    # 工程
COL_F = 6    # rBOM発注番号（必須）
COL_H = 8    # 発注番号
COL_I = 9    # 発注担当者
COL_K = 11   # 仕入先コード（取引先コード）
COL_L = 12   # 品目番号
COL_M = 13   # 発注数
COL_N = 14   # 単価
COL_O = 15   # 発注納期
COL_P = 16   # 発注伝票発行日
COL_BQ = 69  # 注文取消伝票発行フラグ
COL_CC = 81  # 勘定科目


def is_red_background(cell):
    """セルの背景が赤色かどうかをチェック（取引先不一致）"""
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = cell.fill.fgColor.rgb
        if len(rgb) >= 6:
            color = rgb[-6:].upper()
            if color == 'FF6666':
                return True
    return False


def normalize_date(value):
    """日付値を YYYY/MM/DD 形式の文字列に正規化"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y/%m/%d')
    # 文字列の場合はそのまま返す（YYYY/MM/DD形式を想定）
    return str(value).split()[0]  # 時刻部分があれば除去


def truncate_decimal(value, decimal_places=2):
    """小数点以下を指定桁数で切り捨て"""
    if value is None:
        return ''
    try:
        num = float(value)
        factor = 10 ** decimal_places
        return math.floor(num * factor) / factor
    except (ValueError, TypeError):
        return value


def to_integer(value):
    """整数に変換（小数点があれば切り捨て）"""
    if value is None:
        return ''
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return str(value).strip()


def extract_filtered_data():
    """Excelからフィルタリングしたデータを抽出"""
    print("=" * 60)
    print("EJ-rBOM同期ツール: 自動インプット用データ抽出")
    print("=" * 60)

    print(f"\n入力ファイル: {INPUT_EXCEL}")
    print(f"対象発注伝票発行日: {', '.join(TARGET_ISSUE_DATES)}")

    if not INPUT_EXCEL.exists():
        print(f"エラー: 入力ファイルが見つかりません: {INPUT_EXCEL}")
        return [], None

    # Excelを読み込み
    wb = load_workbook(INPUT_EXCEL)
    ws = wb.active

    print(f"全行数: {ws.max_row - 1}行（ヘッダー除く）")

    # フィルタリング処理
    filtered_rows = []
    stats = {
        'total': 0,
        'bq_canceled': 0,      # BQ列が0以外（発注取消済み）
        'a_not_empty': 0,
        'f_not_empty': 0,
        'l_empty': 0,
        'k_ca_pt': 0,
        'cc_not_12_33': 0,
        'p_not_target': 0,
        'passed': 0,
        'k_red_exception': 0,  # 取引先不一致（赤）例外出力
    }

    for row_idx in range(2, ws.max_row + 1):
        stats['total'] += 1

        # 各セルの値を取得
        cell_a = ws.cell(row=row_idx, column=COL_A)
        cell_f = ws.cell(row=row_idx, column=COL_F)
        cell_k = ws.cell(row=row_idx, column=COL_K)
        cell_l = ws.cell(row=row_idx, column=COL_L)
        cell_p = ws.cell(row=row_idx, column=COL_P)
        cell_bq = ws.cell(row=row_idx, column=COL_BQ)
        cell_cc = ws.cell(row=row_idx, column=COL_CC)

        a_value = cell_a.value
        f_value = cell_f.value
        k_value = str(cell_k.value).strip() if cell_k.value is not None else ''
        l_value = cell_l.value
        p_value = normalize_date(cell_p.value)
        cc_value = cell_cc.value

        # フィルタ0: BQ列（注文取消伝票発行フラグ）が0以外の場合は除外
        bq_value = cell_bq.value
        bq_str = str(bq_value).strip() if bq_value is not None else ''
        if bq_str != '' and bq_str != '0':
            stats['bq_canceled'] += 1
            continue

        # 例外: K列が赤背景（取引先不一致）の場合は他のフィルタをスキップして出力対象にする
        if is_red_background(cell_k):
            # L列（品目番号）が空欄の場合は除外
            if l_value is None or str(l_value).strip() == '':
                stats['l_empty'] += 1
                continue
            # CC列（勘定科目）が12または33でない場合は除外
            cc_str = str(cc_value).strip() if cc_value is not None else ''
            if cc_str not in ['12', '33']:
                stats['cc_not_12_33'] += 1
                continue
            # P列（発注伝票発行日）が対象日でない場合は除外
            if p_value not in TARGET_ISSUE_DATES:
                stats['p_not_target'] += 1
                continue
            # 例外的に出力対象
            stats['k_red_exception'] += 1
        else:
            # 通常のフィルタ処理

            # フィルタ1: A列（工程）が空欄である行を抽出
            if a_value is not None and str(a_value).strip() != '':
                stats['a_not_empty'] += 1
                continue

            # フィルタ2: F列（rBOM発注番号）が空欄である行を抽出
            if f_value is not None and str(f_value).strip() != '':
                stats['f_not_empty'] += 1
                continue

            # フィルタ3: L列（品目番号）が空欄である行を省く
            if l_value is None or str(l_value).strip() == '':
                stats['l_empty'] += 1
                continue

            # フィルタ4: K列（取引先コード）がCA/PTのデータを省く
            if k_value.upper() in ['CA', 'PT']:
                stats['k_ca_pt'] += 1
                continue

            # フィルタ6: CC列（勘定科目）が12または33のデータで絞り込む
            cc_str = str(cc_value).strip() if cc_value is not None else ''
            if cc_str not in ['12', '33']:
                stats['cc_not_12_33'] += 1
                continue

            # フィルタ7: P列（発注伝票発行日）が対象日のデータで絞り込む
            if p_value not in TARGET_ISSUE_DATES:
                stats['p_not_target'] += 1
                continue

            # 全フィルタを通過
            stats['passed'] += 1

        # 行データを収集
        row_data = {
            'row_idx': row_idx,
            'K_取引先コード': k_value,
            'I_発注担当者': ws.cell(row=row_idx, column=COL_I).value,
            'H_発注番号': ws.cell(row=row_idx, column=COL_H).value,
            'L_品目番号': str(l_value).strip(),
            'M_発注数': ws.cell(row=row_idx, column=COL_M).value,
            'O_発注納期': ws.cell(row=row_idx, column=COL_O).value,
            'N_単価': ws.cell(row=row_idx, column=COL_N).value,
            'CC_勘定科目': cc_value,
            'P_発注伝票発行日': p_value,
        }

        filtered_rows.append(row_data)

    # 統計表示
    print("\n" + "-" * 40)
    print("フィルタリング結果")
    print("-" * 40)
    print(f"全行数: {stats['total']}行")
    print(f"  除外: BQ列（発注取消済み）: {stats['bq_canceled']}行")
    print(f"  除外: A列（工程）が空欄でない: {stats['a_not_empty']}行")
    print(f"  除外: F列（rBOM発注番号）が空欄でない: {stats['f_not_empty']}行")
    print(f"  除外: L列（品目番号）が空欄: {stats['l_empty']}行")
    print(f"  除外: K列がCA/PT: {stats['k_ca_pt']}行")
    print(f"  除外: CC列（勘定科目）が12/33でない: {stats['cc_not_12_33']}行")
    print(f"  除外: P列（発注伝票発行日）が対象外: {stats['p_not_target']}行")
    print(f"  → 抽出結果（通常）: {stats['passed']}行")
    print(f"  → 抽出結果（取引先不一致/赤・例外）: {stats['k_red_exception']}行")
    total_output = stats['passed'] + stats['k_red_exception']
    print(f"  → 抽出結果（合計）: {total_output}行")

    return filtered_rows, wb


def output_filtered_excel(filtered_rows, original_wb):
    """抽出データをExcelとして出力"""
    print("\n" + "-" * 40)
    print("Excel出力")
    print("-" * 40)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 新しいワークブックを作成
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "抽出データ"

    # ヘッダー行
    headers = ['取引先コード', '発注担当者', '発注番号', '品目番号', '発注数', '発注納期', '単価', '勘定科目', '発注伝票発行日']
    for col_idx, header in enumerate(headers, 1):
        ws_out.cell(row=1, column=col_idx, value=header)

    # データ行
    for row_idx, row_data in enumerate(filtered_rows, 2):
        ws_out.cell(row=row_idx, column=1, value=row_data['K_取引先コード'])
        ws_out.cell(row=row_idx, column=2, value=row_data['I_発注担当者'])
        ws_out.cell(row=row_idx, column=3, value=row_data['H_発注番号'])
        ws_out.cell(row=row_idx, column=4, value=row_data['L_品目番号'])
        ws_out.cell(row=row_idx, column=5, value=row_data['M_発注数'])
        ws_out.cell(row=row_idx, column=6, value=row_data['O_発注納期'])
        ws_out.cell(row=row_idx, column=7, value=row_data['N_単価'])
        ws_out.cell(row=row_idx, column=8, value=row_data['CC_勘定科目'])
        ws_out.cell(row=row_idx, column=9, value=row_data['P_発注伝票発行日'])

    # 保存
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = OUTPUT_DIR / f"抽出データ_{timestamp}.xlsx"
    wb_out.save(output_path)
    print(f"Excel出力完了: {output_path}")

    return output_path


def create_csv_output(filtered_rows):
    """CSVテンプレートにデータを入力して出力"""
    print("\n" + "-" * 40)
    print("CSV出力")
    print("-" * 40)

    if not TEMPLATE_CSV.exists():
        print(f"警告: テンプレートCSVが見つかりません: {TEMPLATE_CSV}")
        print("CSVテンプレートを作成してください")
        return None

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # テンプレートを読み込み（ヘッダー2行を取得）
    with open(TEMPLATE_CSV, 'r', encoding='shift_jis') as f:
        lines = f.readlines()

    # ヘッダー行を保持（1-2行目）
    header_lines = lines[:2] if len(lines) >= 2 else lines

    # 出力ファイル名（日時を含む）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f"EJ_{timestamp}.csv"
    output_path = OUTPUT_DIR / output_filename

    # CSV出力
    with open(output_path, 'w', encoding='shift_jis', newline='') as f:
        # ヘッダー行を書き込み
        for line in header_lines:
            f.write(line)

        # データ行を書き込み（3行目から）
        for row_data in filtered_rows:
            # 各フィールドを整形
            shiire_cd = str(row_data['K_取引先コード']).strip() if row_data['K_取引先コード'] else ''
            tanto_cd = str(row_data['I_発注担当者']).strip() if row_data['I_発注担当者'] else ''
            seiban = 'ZAIKOSEIBAN'  # 固定値
            meisai_biko = str(row_data['H_発注番号']).strip() if row_data['H_発注番号'] else ''
            hinmoku_cd = str(row_data['L_品目番号']).strip() if row_data['L_品目番号'] else ''

            # 発注数（整数）
            hacchu_su = to_integer(row_data['M_発注数'])

            # 発注納期（日付形式を YYYY/M/D に）
            nouki = row_data['O_発注納期']
            if isinstance(nouki, datetime):
                nouki_str = nouki.strftime('%Y/%m/%d')
            elif nouki:
                nouki_str = str(nouki).split()[0]  # 時刻部分を除去
            else:
                nouki_str = ''

            # 単価（小数点2桁で切り捨て）
            tanka = row_data['N_単価']
            if tanka is not None:
                tanka_val = truncate_decimal(tanka, 2)
            else:
                tanka_val = ''

            # 原価コード（整数）
            genka_cd = to_integer(row_data['CC_勘定科目'])

            # CSV行を作成（仕入先コード,担当者コード,製番,明細備考,品目コード,発注数,希望納期,単価,原価コード）
            csv_line = f"{shiire_cd},{tanto_cd},{seiban},{meisai_biko},{hinmoku_cd},{hacchu_su},{nouki_str},{tanka_val},{genka_cd}\n"
            f.write(csv_line)

    print(f"CSV出力完了: {output_path}")
    print(f"出力行数: {len(filtered_rows)}行（ヘッダー2行 + データ{len(filtered_rows)}行）")

    return output_path


def main():
    """メイン処理"""
    # データ抽出
    filtered_data, original_wb = extract_filtered_data()

    if not filtered_data:
        print("\n抽出データがありません")
        return

    # Excel出力
    output_filtered_excel(filtered_data, original_wb)

    # CSV出力
    create_csv_output(filtered_data)

    print("\n" + "=" * 60)
    print(f"処理完了: {len(filtered_data)}行")
    print("=" * 60)


if __name__ == "__main__":
    main()
