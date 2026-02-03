# -*- coding: utf-8 -*-
"""
カム課Excelデータ結合
全ファイルから9項目を抽出し、1つのExcelに結合
"""

import os
import re
from pathlib import Path

import xlrd
import openpyxl
from openpyxl import Workbook

TARGET_HEADERS = ["ロットコード", "開始日", "月次", "完成部番", "子部番", "孫部番", "部品名", "必要数", "工程"]


def is_valid_target_sheet(sheet_name: str) -> bool:
    """有効な対象シートかどうかを判定"""
    if not sheet_name.startswith("ロット_ALL"):
        return False
    suffix = sheet_name[len("ロット_ALL"):]
    if suffix.strip() == "":
        return True
    if re.match(r'^\s*\(ALL\)$', suffix):
        return True
    if re.match(r'^\s*\(★★\)?$', suffix):
        return True
    if re.match(r'^\s*\(1\)$', suffix):
        return True
    return False


def get_header_positions(first_row: list) -> dict:
    """ヘッダー位置を取得"""
    positions = {}
    for header in TARGET_HEADERS:
        if header in first_row:
            positions[header] = first_row.index(header)
    return positions


def extract_data_xls(file_path: Path, sheet_name: str) -> list:
    """xlsファイルからデータ抽出"""
    workbook = xlrd.open_workbook(str(file_path))
    sheet = workbook.sheet_by_name(sheet_name)

    if sheet.nrows < 2:
        return []

    # 1行目からヘッダー位置取得
    first_row = [str(cell.value).strip() for cell in sheet.row(0)]
    positions = get_header_positions(first_row)

    if len(positions) != len(TARGET_HEADERS):
        return []

    # 2行目以降のデータ抽出
    data = []
    for row_idx in range(1, sheet.nrows):
        row_data = []
        for header in TARGET_HEADERS:
            col_idx = positions[header]
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value

            # 日付型の処理
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(value, workbook.datemode)
                    value = dt.strftime("%Y/%m/%d")
                except:
                    pass

            row_data.append(value)

        # 空行スキップ（ロットコードが空の場合）
        if row_data[0] is None or str(row_data[0]).strip() == "":
            continue

        data.append(row_data)

    return data


def extract_data_xlsx(file_path: Path, sheet_name: str) -> list:
    """xlsxファイルからデータ抽出"""
    workbook = openpyxl.load_workbook(str(file_path), read_only=True)
    sheet = workbook[sheet_name]

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # 1行目からヘッダー位置取得
    first_row = [str(cell).strip() if cell else "" for cell in rows[0]]
    positions = get_header_positions(first_row)

    if len(positions) != len(TARGET_HEADERS):
        return []

    # 2行目以降のデータ抽出
    data = []
    for row in rows[1:]:
        row_data = []
        for header in TARGET_HEADERS:
            col_idx = positions[header]
            value = row[col_idx] if col_idx < len(row) else None
            row_data.append(value)

        # 空行スキップ
        if row_data[0] is None or str(row_data[0]).strip() == "":
            continue

        data.append(row_data)

    return data


def main():
    base_dir = Path(r"C:\Dev\90_tools\30_工程マスタ作成\02_カム課")
    output_file = base_dir / "全カム課加工検討データ.xlsx"

    # 全Excelファイルを検索
    excel_files = []
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.endswith(('.xls', '.xlsx')) and not f.startswith('~$') and f != "全カム課加工検討データ.xlsx":
                excel_files.append(Path(root) / f)

    excel_files.sort()
    print(f"対象ファイル数: {len(excel_files)}")

    all_data = []
    error_files = []

    for i, file_path in enumerate(excel_files):
        rel_path = file_path.relative_to(base_dir)
        ext = file_path.suffix.lower()

        try:
            # シート名取得
            if ext == '.xls':
                workbook = xlrd.open_workbook(str(file_path))
                sheet_names = workbook.sheet_names()
            else:
                workbook = openpyxl.load_workbook(str(file_path), read_only=True)
                sheet_names = workbook.sheetnames

            # 有効な対象シートを検索
            target_sheet = None
            for s in sheet_names:
                if is_valid_target_sheet(s):
                    target_sheet = s
                    break

            if not target_sheet:
                error_files.append(f"{rel_path} => 対象シートなし")
                continue

            # データ抽出
            if ext == '.xls':
                data = extract_data_xls(file_path, target_sheet)
            else:
                data = extract_data_xlsx(file_path, target_sheet)

            all_data.extend(data)

            if (i + 1) % 20 == 0:
                print(f"処理中: {i + 1}/{len(excel_files)} - 累計行数: {len(all_data)}")

        except Exception as e:
            error_files.append(f"{rel_path} => {e}")

    print(f"処理完了: {len(excel_files)}ファイル - 合計行数: {len(all_data)}")

    # Excel出力
    wb = Workbook()
    ws = wb.active
    ws.title = "全データ"

    # ヘッダー出力
    ws.append(TARGET_HEADERS)

    # データ出力
    for row in all_data:
        ws.append(row)

    wb.save(output_file)
    print(f"出力: {output_file}")
    print(f"合計行数: {len(all_data)} (ヘッダー除く)")

    if error_files:
        print(f"\nエラー ({len(error_files)}件):")
        for e in error_files:
            print(f"  {e}")


if __name__ == "__main__":
    main()
