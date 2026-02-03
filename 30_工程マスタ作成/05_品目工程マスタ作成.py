# -*- coding: utf-8 -*-
"""
05_品目工程マスタ作成.py

04_マスター原本.csvから品目工程マスタ（M0840）データを作成し、
マスター貼り付け用.xlsxに出力する

処理内容:
  1. M_ITEMテーブルチェック（子部番１、子部番２）
     - 子部番１が存在しない → 警告.csv
     - 子部番２が存在しない → 対象外.csv
  2. SRPRICE/SRCDデータ取得（子部番２をキーとして）
     - プライマリ: M_PUCH_UNIT_COST_H + M_PUCH_UNIT_COST
     - フォールバック: T_RLSD_PUCH_ODR
  3. 品目工程マスタデータ作成
     - SEQ=1: 全データ（CAFIN固定）
     - SEQ=2: 工程２が空欄でない場合のみ
  4. Excelファイル書き込み

出力:
  - work/00_警告.csv（子部番１がM_ITEMに存在しない）
  - work/00_対象外.csv（子部番２がM_ITEMに存在しない）
  - マスター貼り付け用.xlsx の [品目工程マスタ] シート（13行目から）
"""

import sys
import csv
import oracledb
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# 入力ファイル
INPUT_CSV = Path(r"C:\Dev\90_tools\30_工程マスタ作成\work\04_マスター原本.csv")

# 出力先
OUTPUT_DIR = Path(r"C:\Dev\90_tools\30_工程マスタ作成\work")
OUTPUT_WARNING_CSV = OUTPUT_DIR / "00_警告.csv"
OUTPUT_EXCLUDED_CSV = OUTPUT_DIR / "00_対象外.csv"
OUTPUT_EXCEL = Path(r"C:\Dev\90_tools\30_工程マスタ作成\10_マスター貼り付け用\マスター貼り付け用.xlsx")

# EJシステム接続情報
EJ_HOST = "172.17.107.102"
EJ_PORT = 1521
EJ_SERVICE = "EXPJ"
EJ_USER = "EXPJ2"
EJ_PASSWORD = "EXPJ2"

# oracledb thick mode初期化フラグ
_thick_mode_initialized = False


def init_oracle_thick_mode():
    """Oracle thick modeを初期化"""
    global _thick_mode_initialized
    if not _thick_mode_initialized:
        try:
            oracledb.init_oracle_client()
            _thick_mode_initialized = True
            print("  oracledb thick mode初期化完了")
        except Exception as e:
            print(f"  thick mode初期化スキップ: {e}")


def get_ej_connection():
    """EJシステムへの接続を取得"""
    init_oracle_thick_mode()
    dsn = oracledb.makedsn(EJ_HOST, EJ_PORT, service_name=EJ_SERVICE)
    return oracledb.connect(user=EJ_USER, password=EJ_PASSWORD, dsn=dsn)


def fetch_m_item_data():
    """
    M_ITEMテーブルからITEM_CD → PUCH_FIXED_LTの辞書を作成

    Returns:
        dict: {ITEM_CD: PUCH_FIXED_LT}
    """
    print("\nM_ITEMテーブル取得中...")
    conn = get_ej_connection()
    try:
        cursor = conn.cursor()
        sql = """
            SELECT ITEM_CD, PUCH_FIXED_LT
            FROM EXPJ2.M_ITEM
            WHERE ITEM_CD IS NOT NULL
        """
        cursor.execute(sql)

        result = {}
        for row in cursor:
            item_cd = row[0]
            puch_fixed_lt = row[1] if row[1] is not None else 0
            result[item_cd] = puch_fixed_lt

        print(f"M_ITEMデータ取得完了: {len(result):,}品目")
        return result

    finally:
        conn.close()


def fetch_srprice_primary_data():
    """
    M_PUCH_UNIT_COST_H + M_PUCH_UNIT_COSTから
    ITEM_CD → {SRCD, SRPRICE}の辞書を作成（プライマリデータソース）

    Returns:
        dict: {ITEM_CD: {'SRCD': VEND_CD, 'SRPRICE': UNIT_COST}}
    """
    print("\n=== プライマリデータソース取得（M_PUCH_UNIT_COST_H + M_PUCH_UNIT_COST）===")
    try:
        conn = get_ej_connection()
        cursor = conn.cursor()

        sql = """
            SELECT h.ITEM_CD, h.VEND_CD, h.PUCH_PRIORITY_REF_NO,
                   c.UNIT_COST, c.EFF_PHASE_IN_DATE, c.PUCH_SIZE
            FROM EXPJ2.M_PUCH_UNIT_COST_H h
            INNER JOIN EXPJ2.M_PUCH_UNIT_COST c
            ON h.ITEM_CD = c.ITEM_CD AND h.VEND_CD = c.VEND_CD
            WHERE h.PUCH_PRIORITY_REF_NO IS NOT NULL
            AND c.EFF_PHASE_IN_DATE < TO_DATE('2025-12-01', 'YYYY-MM-DD')
            ORDER BY h.ITEM_CD, h.PUCH_PRIORITY_REF_NO ASC,
                     c.EFF_PHASE_IN_DATE DESC, c.PUCH_SIZE ASC
        """

        cursor.execute(sql)
        rows = cursor.fetchall()
        conn.close()

        print(f"プライマリデータ取得: {len(rows):,}行")

        # ITEM_CDごとに最初の1行のみ採用（ORDER BY条件により最優先）
        result = {}
        for row in rows:
            item_cd = row[0]
            if item_cd not in result:
                result[item_cd] = {
                    'SRCD': row[1],
                    'SRPRICE': row[3],
                    'EFF_PHASE_IN_DATE': row[4],
                    'PUCH_SIZE': row[5],
                    'PRIORITY': row[2]
                }

        print(f"プライマリ辞書作成: {len(result):,}件")
        return result

    except Exception as e:
        print(f"プライマリデータ取得エラー: {e}")
        return {}


def fetch_srprice_fallback_data():
    """
    T_RLSD_PUCH_ODRから
    ITEM_CD → {SRCD, SRPRICE}の辞書を作成（フォールバックデータソース）

    Returns:
        dict: {ITEM_CD: {'SRCD': VEND_CD, 'SRPRICE': UNIT_COST}}
    """
    print("\n=== フォールバックデータソース取得（T_RLSD_PUCH_ODR）===")
    try:
        conn = get_ej_connection()
        cursor = conn.cursor()

        sql = """
            SELECT ITEM_CD, VEND_CD, UNIT_COST, PUCH_ODR_DLV_DATE
            FROM EXPJ2.T_RLSD_PUCH_ODR
            WHERE ITEM_CD IS NOT NULL AND VEND_CD IS NOT NULL
            ORDER BY ITEM_CD, PUCH_ODR_DLV_DATE DESC
        """

        cursor.execute(sql)
        rows = cursor.fetchall()
        conn.close()

        print(f"フォールバックデータ取得: {len(rows):,}行")

        # ITEM_CDごとに最新日付の1行のみ採用
        result = {}
        for row in rows:
            item_cd = row[0]
            if item_cd not in result:
                result[item_cd] = {
                    'SRCD': row[1],
                    'SRPRICE': row[2],
                    'PUCH_ODR_DLV_DATE': row[3]
                }

        print(f"フォールバック辞書作成: {len(result):,}件")
        return result

    except Exception as e:
        print(f"フォールバックデータ取得エラー: {e}")
        return {}


def check_m_item_existence(rows, m_item_dict):
    """
    M_ITEMテーブルに子部番１、子部番２が存在するかチェック

    Args:
        rows: 入力データ行リスト
        m_item_dict: M_ITEMテーブルの辞書

    Returns:
        tuple: (警告行リスト, 対象外行リスト, 処理対象行リスト)
    """
    print("\nM_ITEMテーブル存在チェック中...")
    warning_rows = []
    excluded_rows = []
    target_rows = []

    for row in rows:
        kobuban1 = row.get('子部番１', '').strip()
        kobuban2 = row.get('子部番２', '').strip()

        has_warning = False
        has_excluded = False

        # 子部番１チェック（空欄でない場合のみ）
        if kobuban1 and kobuban1 not in m_item_dict:
            has_warning = True

        # 子部番２チェック（空欄でない場合のみ）
        if kobuban2 and kobuban2 not in m_item_dict:
            has_excluded = True

        # 分類
        if has_excluded:
            # 対象外（子部番２が存在しない）
            excluded_rows.append(row)
        elif has_warning:
            # 警告のみ（子部番１が存在しない、子部番２はOK）
            warning_rows.append(row)
        else:
            # 処理対象
            target_rows.append(row)

    print(f"  警告のみ: {len(warning_rows):,}行")
    print(f"  対象外含む: {len(excluded_rows):,}行")
    print(f"  処理対象: {len(target_rows):,}行")

    return warning_rows, excluded_rows, target_rows


def create_m0840_data(target_rows, m_item_dict, srprice_primary_dict, srprice_fallback_dict):
    """
    品目工程マスタ（M0840）データを作成

    Args:
        target_rows: 処理対象行リスト
        m_item_dict: M_ITEMテーブルの辞書
        srprice_primary_dict: プライマリSRPRICE辞書
        srprice_fallback_dict: フォールバックSRPRICE辞書

    Returns:
        list: M0840データ行リスト
    """
    print("\n品目工程マスタデータ作成中...")
    output_rows = []

    for row in target_rows:
        kansei_bango = row.get('完成部番', '').strip()
        kotei2 = row.get('工程２', '').strip()
        kobuban2 = row.get('子部番２', '').strip()
        shiire_cd = row.get('仕入先CD', '').strip()

        # ==================== SEQ=1行（必ず作成）====================
        row1 = {
            'HMCD': kansei_bango,
            'SEQ': 1,
            'KTSEQ': 5010,
            'KTCD': 'CAFIN',
            'SRCD': 'CC',
            'SGNCD': '',
            'DDTIME': 0,
            'SGTIME': 0,
            'LDTIME': 2,
            'SRPRICE': 0,
            'CSBCD': '10',
            'SUPCLSCD': '3',
            'SUPCD': 'C',
            'RCVTSTKBN': 2,
            'RCVCHKKBN': 2
        }
        output_rows.append(row1)

        # ==================== SEQ=2行（工程２が空欄でない場合のみ）====================
        if kotei2:
            # 子部番２が空欄の場合
            if not kobuban2:
                ldtime = 0
                srcd = shiire_cd
                srprice = 0
            else:
                # LDTIME取得（子部番２からM_ITEMテーブルを参照）
                ldtime = m_item_dict.get(kobuban2, 0)

                # SRPRICE/SRCD取得（プライマリ → フォールバック）
                # 子部番２をキーとして検索
                srprice_data = srprice_primary_dict.get(kobuban2, {})
                if not srprice_data:
                    srprice_data = srprice_fallback_dict.get(kobuban2, {})

                # VEND_CDを優先使用（一貫性保証）
                # 見つからない場合のみ元の仕入先CDを使用
                srcd = srprice_data.get('SRCD', shiire_cd)
                srprice = srprice_data.get('SRPRICE', 0)

            row2 = {
                'HMCD': kansei_bango,
                'SEQ': 2,
                'KTSEQ': 5020,
                'KTCD': kotei2,
                'SRCD': srcd,
                'SGNCD': '',
                'DDTIME': 0,
                'SGTIME': 0,
                'LDTIME': ldtime,
                'SRPRICE': srprice,
                'CSBCD': '13',
                'SUPCLSCD': '5',
                'SUPCD': 'C',
                'RCVTSTKBN': 2,
                'RCVCHKKBN': 2
            }
            output_rows.append(row2)

    print(f"品目工程マスタデータ作成完了: {len(output_rows):,}行")
    return output_rows


def write_excel_output(output_rows, excel_path):
    """
    マスター貼り付け用.xlsxの[品目工程マスタ]シートに書き込み

    Args:
        output_rows: M0840データ行リスト
        excel_path: Excelファイルパス
    """
    print(f"\nExcelファイル書き込み中: {excel_path}")

    # テンプレートを読み込み
    wb = openpyxl.load_workbook(excel_path)

    # シート名確認
    if '品目工程マスタ' not in wb.sheetnames:
        print(f"エラー: シート '品目工程マスタ' が見つかりません")
        print(f"利用可能なシート: {wb.sheetnames}")
        return False

    ws = wb['品目工程マスタ']

    # 12行目からデータを書き込み
    start_row = 12
    for idx, data_row in enumerate(output_rows):
        row_num = start_row + idx

        # 各列に値を設定
        ws.cell(row_num, 1, data_row['HMCD'])       # A列: HMCD
        ws.cell(row_num, 2, data_row['SEQ'])        # B列: SEQ
        ws.cell(row_num, 3, data_row['KTSEQ'])      # C列: KTSEQ
        ws.cell(row_num, 4, data_row['KTCD'])       # D列: KTCD
        ws.cell(row_num, 5, data_row['SRCD'])       # E列: SRCD
        ws.cell(row_num, 6, data_row['SGNCD'])      # F列: SGNCD
        ws.cell(row_num, 7, data_row['DDTIME'])     # G列: DDTIME
        ws.cell(row_num, 8, data_row['SGTIME'])     # H列: SGTIME
        ws.cell(row_num, 9, data_row['LDTIME'])     # I列: LDTIME
        ws.cell(row_num, 10, data_row['SRPRICE'])   # J列: SRPRICE
        ws.cell(row_num, 11, data_row['CSBCD'])     # K列: CSBCD
        ws.cell(row_num, 12, data_row['SUPCLSCD'])  # L列: SUPCLSCD
        ws.cell(row_num, 13, data_row['SUPCD'])     # M列: SUPCD
        ws.cell(row_num, 14, data_row['RCVTSTKBN']) # N列: RCVTSTKBN
        ws.cell(row_num, 15, data_row['RCVCHKKBN']) # O列: RCVCHKKBN

    # 保存
    wb.save(excel_path)
    print(f"Excelファイル書き込み完了: {len(output_rows):,}行")

    return True


def main():
    print("=" * 60)
    print("05_品目工程マスタ作成")
    print(f"入力: {INPUT_CSV}")
    print(f"出力: {OUTPUT_EXCEL}")
    print("=" * 60)

    # 入力ファイル確認
    if not INPUT_CSV.exists():
        print(f"エラー: 入力ファイルが見つかりません: {INPUT_CSV}")
        return False

    # 出力ディレクトリ確認
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # データ読み込み
    print("\nデータを読み込み中...")
    with open(INPUT_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    print(f"総行数: {len(rows):,}行")

    # M_ITEMテーブル取得
    m_item_dict = fetch_m_item_data()

    # SRPRICE/SRCDデータ取得
    srprice_primary_dict = fetch_srprice_primary_data()
    srprice_fallback_dict = fetch_srprice_fallback_data()

    # M_ITEMテーブル存在チェック
    warning_rows, excluded_rows, target_rows = check_m_item_existence(rows, m_item_dict)

    # 警告.csv出力
    if warning_rows:
        print(f"\n警告ファイル出力中: {OUTPUT_WARNING_CSV}")
        fieldnames = list(rows[0].keys()) if rows else []
        with open(OUTPUT_WARNING_CSV, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(warning_rows)
        print(f"  警告: {len(warning_rows):,}行 → {OUTPUT_WARNING_CSV.name}")

    # 対象外.csv出力
    if excluded_rows:
        print(f"\n対象外ファイル出力中: {OUTPUT_EXCLUDED_CSV}")
        fieldnames = list(rows[0].keys()) if rows else []
        with open(OUTPUT_EXCLUDED_CSV, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(excluded_rows)
        print(f"  対象外: {len(excluded_rows):,}行 → {OUTPUT_EXCLUDED_CSV.name}")

    # 品目工程マスタデータ作成
    m0840_data = create_m0840_data(target_rows, m_item_dict, srprice_primary_dict, srprice_fallback_dict)

    # Excelファイル書き込み
    if not write_excel_output(m0840_data, OUTPUT_EXCEL):
        return False

    # 統計情報
    print("\n処理完了")
    print(f"  警告: {len(warning_rows):,}行")
    print(f"  対象外: {len(excluded_rows):,}行")
    print(f"  処理対象: {len(target_rows):,}行")
    print(f"  M0840出力: {len(m0840_data):,}行")
    print("=" * 60)

    return True


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
