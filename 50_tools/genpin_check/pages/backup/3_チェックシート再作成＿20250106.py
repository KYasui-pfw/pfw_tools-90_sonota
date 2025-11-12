################################
#　新現品票追加作成　　　　　　　 #
################################
#インポート
import streamlit as st
import streamlit.components.v1 as components
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import pandas as pd
import os
import shutil
import pyqrcode
import openpyxl
from PIL import Image
from datetime import datetime, timedelta, timezone
import pythoncom
import random
import sqlite3
import time
from module.conmas_upload import data_upload

try:
  pythoncom.CoInitialize() #サーバーサイドからローカルファイルを動かすときに必要
  
  @st.cache_resource
  def ireporter_data_get(sql):

    # #DB接続定義
    db_url = 'postgresql://postgres:cimtops@ESRV10/irepodb'

    # エンジンを作成
    engine = create_engine(db_url, echo=True)

    # セッションを作成するためのSessionクラスを生成
    Session = sessionmaker(bind=engine)
    session = Session()

    # コネクションを取得
    with engine.connect() as connection:
      # SQLクエリの実行
      df = pd.read_sql(sql, connection)

    # セッションを閉じる
    session.close()   

    return(df)
  
  #sqlite3への接続
  def sqlite_data_get(sql,filepath):
      # #DB接続定義
      db_url = f'sqlite:///{filepath}'

      # エンジンを作成
      engine = create_engine(db_url, echo=True)

      # セッションを作成するためのSessionクラスを生成
      Session = sessionmaker(bind=engine)
      session = Session()

      # コネクションを取得
      with engine.connect() as connection:
        # SQLクエリの実行
        df = pd.read_sql(sql, connection)
  
      # セッションを閉じる
      session.close()   
  
      return(df)


  def irepo_jidoutyouhyou_create(df):

    # 現品票未発行の有無は関係無い

    df = df[(df['ireporter管理flg'] == 1)& (df['QRコードレイアウト区分'] == 0)]
    err_exists = df.empty

    if not err_exists:
      #自動帳票作成用DFの作成
      dt_now = datetime.now(timezone(timedelta(hours=9))) # 日本時刻
      create_day = dt_now.strftime('%Y%m%d')

      df.insert(0, 'defTopId', '') 
      
      #縦型
      mask_df = df['縦横区分'] == 1
      sql = "select def_top_id from view_def_top where def_top_org = 405 and public_status = 2"
      irepo_df = ireporter_data_get(sql)
      df.loc[mask_df, 'defTopId'] = irepo_df['def_top_id'].max()

      #横型
      mask_df = df['縦横区分'] == 0
      sql = "select def_top_id from view_def_top where def_top_org = 406 and public_status = 2"
      irepo_df = ireporter_data_get(sql)
      df.loc[mask_df, 'defTopId'] = irepo_df['def_top_id'].max()

      #必要項目を並べる
      df1 = df.loc[:,['defTopId','部品名','ロット番号','完成部番','月次','組立番号','機種','吋','G','客先名','ﾒｯｷ','工程VERSION']]
      df1.loc[~(df1['ﾒｯｷ'] == '1.0'), 'ﾒｯｷ'] = ''
      df1.loc[df1['ﾒｯｷ'] == '1.0', 'ﾒｯｷ'] = '無電解ニッケルメッキ'
      df1['作成日'] = create_day

      #先頭行に固定値をインサート
      df1.insert(0, 'H', 'R')
      df1.columns = ['H','defTopId','S1C1','S1C2','S1C3','S1C4','S1C5','S1C6','S1C7','S1C8','S1C9','S1C11','S1C644','S1C643']
         
      #CSVに変換
      csv_name = os.path.dirname(os.path.dirname(__file__))+"\\csv\\"+dt_now.strftime('%Y%m%d%H%M%S')+"_自動帳票作成データ"+".csv"
      df1.to_csv(csv_name,index=False,encoding='CP932')
      data_upload(csv_name,df1.shape[0])

  #sqliteから取得
  @st.cache_resource
  def df_set1(lot):

    # #DB接続定義
    dbname = 'genpinhyo.db'
    filepath = os.path.dirname(os.path.dirname(__file__))+f'\\Database\\'+dbname      
    db_url = f'sqlite:///{filepath}'

    # エンジンを作成
    engine = create_engine(db_url, echo=True)

    # セッションを作成するためのSessionクラスを生成
    Session = sessionmaker(bind=engine)
    session = Session()

    # コネクションを取得
    with engine.connect() as connection:
      # SQLクエリの実行
      df = pd.read_sql(f'select * from genpinhyo where ロット番号 = "{lot}"', connection)

    # セッションを閉じる
    session.close()  

    return(df)

  @st.cache_resource
  def df_set2():

    # #DB接続定義
    dbname = 'genpinhyo.db'
    filepath = os.path.dirname(os.path.dirname(__file__))+f'\\Database\\'+dbname      
    db_url = f'sqlite:///{filepath}'

    # エンジンを作成
    engine = create_engine(db_url, echo=True)

    # セッションを作成するためのSessionクラスを生成
    Session = sessionmaker(bind=engine)
    session = Session()

    # コネクションを取得
    with engine.connect() as connection:
      # SQLクエリの実行
      df = pd.read_sql(f'select * from buhin_irepo_mst', connection)

    # セッションを閉じる
    session.close()  

    return(df)

  def main():

    st.set_page_config(
    page_title = '現品票再発行',
      layout="wide" )

    HIDE_ST_STYLE = """
                    <style>
                    div[data-testid="stToolbar"] {
                    visibility: hidden;
                    height: 0%;
                    position: fixed;
                    }
                    div[data-testid="stDecoration"] {
                    visibility: hidden;
                    height: 0%;
                    position: fixed;
                    }
                    #MainMenu {
                    visibility: hidden;
                    height: 0%;
                    }
                    header {
                    visibility: hidden;
                    height: 0%;
                    }
                    footer {
                    visibility: hidden;
                    height: 0%;
                    }
                    .appview-container .main .block-container{
                                padding-top: 1rem;
                                padding-right: 3rem;
                                padding-left: 3rem;
                                padding-bottom: 1rem;
                            }  
                            .reportview-container {
                                padding-top: 0rem;
                                padding-right: 3rem;
                                padding-left: 3rem;
                                padding-bottom: 0rem;
                            }
                            header[data-testid="stHeader"] {
                                z-index: -1;
                            }
                            div[data-testid="stToolbar"] {
                            z-index: 100;
                            }
                            div[data-testid="stDecoration"] {
                            z-index: 100;
                            }
                    .block-container {
                            padding-top: 0rem !important;
                            padding-bottom: 0rem !important;
                            }        
                    </style>
    """
    st.markdown(HIDE_ST_STYLE, unsafe_allow_html=True)
    
    sql = 'SELECT * FROM buhin_irepo_mst'
    filepath = os.path.dirname(os.path.dirname(__file__))+'\\Database\\genpinhyo.db'
    buhin_irepo_df = sqlite_data_get(sql,filepath)
    buhin_irepo_df = buhin_irepo_df.sort_values(['部品名'])
    buhin_list = buhin_irepo_df['部品名'].tolist()
    buhin_list.insert(0, "")

    st.markdown("### チェックシート再作成")
    st.markdown('''この画面は途中で不合格が出た場合等、**同じロット番号で２つ目のチェックシートを発行する場合**に使用します。  
                  現品票が汚れたり破れた等で現品票を再発行したい場合は、通常の「現品票発行」画面より印刷してください。  
                  ※i-Reporterでチェックしていないチェックシートも、通常の「現品票発行」画面より印刷してください''')
    
    # 水平線
    st.divider()       

    bcol1,bcol2,bcol3,bcol4= st.columns([2,1,2,2]) 
    with bcol1:
      lot = st.text_input('ロット番号入力', '')
      df1 = df_set1(lot)
      df2 = df_set2()
      sql = f"select rep_top_id,cluster_1_2_t from view_report_406" #横レイアウト
      df3 = ireporter_data_get(sql)
      sql = f"select rep_top_id,cluster_1_2_t from view_report_405" #縦レイアウト
      df4 = ireporter_data_get(sql)
      df5 = pd.concat([df3,df4], ignore_index=True)
      df = pd.merge(df1, df2, on=['部品名'],how='left')
      df = pd.merge(df, df5, left_on=['ロット番号'],right_on=['cluster_1_2_t'],how='left')
      df = df.rename(columns={'rep_top_id': '帳票発行ID'})
      df = df.drop(columns=['cluster_1_2_t'])
      df['帳票発行ID'] = df["帳票発行ID"].fillna(0) #Noneデータ対
      df['工程VERSION'] = df["工程VERSION"].fillna(0) #Noneデータ対
      df['工程VERSION']  = df['工程VERSION'].astype(float).astype(int)
      max_rep_top_id = df['帳票発行ID'].max()
      df = df[df['帳票発行ID'] == max_rep_top_id]
      df = df.fillna('') #Noneデータ対 
    with bcol2:
      st.write("")
      st.write("")

    if st.button("検索"):
      st.divider() 
      st.write("出力対象データ")
      if not df.empty:
        st.dataframe(df)
      else:
        st.write("出力対象データ無し　※ロット番号が正しくありません")

    # 水平線
    st.divider() 
    
    fcol1,fcol2= st.columns([2,3]) 
    with fcol1:
      if not df.empty:
        if st.button("手入力現品票を印刷"):
          #print(df['ireporter管理flg'].iloc[0])
          with st.spinner('現品票作成処理　実行中'):

            dt = datetime.now(timezone(timedelta(hours=9))) # 日本時刻
            dt_now = dt.strftime('%Y/%m/%d') 

            #自動帳票作成処理
            if df['ireporter管理flg'].iloc[0] == 0: #irepo対象外
              tyohyoID = 0
            else: #対象
              df1 = df.loc[:,['部品名','ロット番号','完成部番','月次','組立番号','機種','吋','G','客先名','ﾒｯｷ','工程VERSION','ireporter管理flg','QRコードレイアウト区分','縦横区分','帳票発行ID']]

              with st.spinner(f'iReporter帳票作成対象のため、20秒ほどお待ちください'):
                irepo_jidoutyouhyou_create(df1)
                st.cache_resource.clear()
                time.sleep(10) #待機しないとDB更新前に帳票を発行してしまう可能性がある
                
                #帳票ID再取得
                sql = f"select rep_top_id,cluster_1_2_t from view_report_406" #横レイアウト
                df3 = ireporter_data_get(sql)
                sql = f"select rep_top_id,cluster_1_2_t from view_report_405" #縦レイアウト
                df4 = ireporter_data_get(sql)
                df5 = pd.concat([df3,df4], ignore_index=True)
                df1 = pd.merge(df1, df5, left_on=['ロット番号'],right_on=['cluster_1_2_t'],how='left')
                tyohyoID = df1['rep_top_id'].max()


            #QRコード作成処理
            if (df['QRコードレイアウト区分'] == 1).any():
              #QRコード作成(旧レイアウトでの作成)
              code = pyqrcode.create(f"{lot},{df['月次'].iloc[0]},{df['組立番号'].iloc[0]},{df['機種'].iloc[0]},{df['吋'].iloc[0]},{df['G'].iloc[0]},{df['完成部番'].iloc[0]},{df['部品名'].iloc[0]},'',{df['客先コード'].iloc[0]},{df['ﾒｯｷ'].iloc[0]}",error='L', version=5, mode='binary')
            else:
              code = pyqrcode.create(f"jp.co.cimtops.ireporter.openreport:repid={tyohyoID}", error='L', version=5, mode='binary')
            code.png(os.path.dirname(__file__)+f"\\work\\qrcode_{lot}_{dt.strftime('%H%M%S')}.png", scale=3)
            #余白を透明化
            img = Image.open(os.path.dirname(__file__)+f"\\work\\qrcode_{lot}_{dt.strftime('%H%M%S')}.png")
            img = img.convert("RGBA")
            datas = img.getdata()
            newData = []
            for item in datas:
                if item[0] == 255 and item[1] == 255 and item[2] == 255:
                    newData.append((255, 255, 255, 0))
                else:
                    newData.append(item)
            img.putdata(newData)
            img.save(os.path.dirname(os.path.dirname(__file__))+f"\\work\\qrcode_{lot}_{dt.strftime('%H%M%S')}.png", "PNG")

            #アウトプットのエクセルの準備
            dt_now = datetime.now(timezone(timedelta(hours=9))) # 日本時刻
            wb = os.path.dirname(__file__)+"\\add_genpinhyo\\再発行用現品票_"+dt_now.strftime('%Y%m%d%H%M%S') +'.xlsx'
            wb_name ="手入力現品票_"+dt_now.strftime('%Y%m%d%H%M%S') +'.xlsx'
            shutil.copy(os.path.dirname(__file__)+'\\template\\add_template.xlsx', wb)

            out_wb = openpyxl.load_workbook(wb)
            out_wb.active.title="手入力現品票"
            out_wb.save(wb)

            chuui1 = "【再発行】"
            if df['ﾒｯｷ'].iloc[0] == "1.0":
              chuui2 = '無電解ニッケルメッキ'
            else:
              chuui2 = ''
            #格納処理 
            out_wb.active.cell(1, 15).value = chuui1 #無電解ニッケルﾒｯｷ
            out_wb.active.cell(2, 10).value = chuui2 #無電解ニッケルﾒｯｷ
            out_wb.active.cell(3, 1).value = dt.strftime('%Y/%m/%d') #発行日
            out_wb.active.cell(4, 21).value = df['必要数'].iloc[0] #個数
            out_wb.active.cell(5, 7).value = df['完成部番'].iloc[0] #完成部番
            out_wb.active.cell(8, 7).value = df['B1図番'].iloc[0] #B1部番
            out_wb.active.cell(9, 7).value = df['B2図番'].iloc[0] #B2部番
            out_wb.active.cell(10, 7).value = df['F1図番'].iloc[0] #F1部番
            out_wb.active.cell(11, 7).value = df['F2図番'].iloc[0] #F2部番
            out_wb.active.cell(12, 7).value = df['L図番'].iloc[0] #L部番
            out_wb.active.cell(13, 7).value = df['R図番'].iloc[0] #R部番
            out_wb.active.cell(14, 7).value = df['C図番'].iloc[0] #C部番
            out_wb.active.cell(1, 30).value = lot #lot番号
            out_wb.active.cell(4, 24).value = df['月次'].iloc[0]  #月次
            out_wb.active.cell(7, 24).value = df['組立番号'].iloc[0] #組立番号
            out_wb.active.cell(10, 24).value = df['機種'].iloc[0] #機種名
            img_to_excel = openpyxl.drawing.image.Image(os.path.dirname(os.path.dirname(__file__))+f"\\work\\qrcode_{lot}_{dt.strftime('%H%M%S')}.png") #QRコード
            out_wb.active.add_image(img_to_excel, f'AI{3}')
            out_wb.active.cell(13, 32).value = df['吋'].iloc[0]  #インチ
            out_wb.active.cell(13, 38).value = df['G'].iloc[0]  #ゲージ
            out_wb.active.cell(15, 6).value = df['工程'].iloc[0]  #工程

            #エクセルを保存
            out_wb.save(wb)
            out_wb.close()

            # ファイルを読み込む
            with open(wb, 'rb') as file:
              filedata = file.read()
            st.markdown('処理終了：処理結果のエクセルを以下からダウンロードしてください')
            st.download_button(
              label='処理結果ダウンロード',
              data=filedata,
              file_name=wb_name,
              mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
              )
            st.cache_resource.clear() 
          
  if __name__ == "__main__":
    main()

except Exception as e:
  #簡単なエラー処理
  print(e)
  dt_now = datetime.now(timezone(timedelta(hours=9))) # 日本時刻
  with open(os.path.dirname(__file__)+"\\"+dt_now.strftime('%Y%m%d%H%M%S')+"_err"+".txt", mode='w') as f:
    f.write(str(e))
