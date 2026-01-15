################################
#　新現品票追加作成　　　　　　　 #
################################
#インポート
import streamlit as st
import streamlit.components.v1 as components
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import pandas as pd
import os
import shutil
import pyqrcode
import openpyxl
from PIL import Image
from datetime import datetime, timedelta, timezone
import pythoncom
import random
import sqlite3
import time
from module.conmas_upload import data_upload

try:
  pythoncom.CoInitialize() #サーバーサイドからローカルファイルを動かすときに必要


  # krdのmachinDBに接続する（SQLite版）
  @st.cache_resource
  def krd_data_get(sql):
    # SQLite接続（KRD MySQL → SQLite同期データベース）
    # \\esrv11\KakouDenpyo\krd_machine.db
    sqlite_db_path = r'\\esrv11\KakouDenpyo\krd_machine.db'

    conn = sqlite3.connect(sqlite_db_path)
    df = pd.read_sql(sql, conn)
    conn.close()

    return df

  # # 【旧版：MySQL接続】コメントアウト（2025-11-22 SQLite移行）
  # @st.cache_resource
  # def krd_data_get(sql):
  #   # #DB接続定義
  #   db_url = 'mysql+pymysql://pfw:mejiriHoo@krd/machin?charset=utf8'
  #
  #   # エンジンを作成
  #   engine = create_engine(db_url, echo=True)
  #
  #   # セッションを作成するためのSessionクラスを生成
  #   Session = sessionmaker(bind=engine)
  #   session = Session()
  #
  #   # コネクションを取得
  #   with engine.connect() as connection:
  #     # SQLクエリの実行
  #     df = pd.read_sql(sql, connection)
  #
  #   # セッションを閉じる
  #   session.close()
  #
  #   return(df)
  
  @st.cache_resource
  def ireporter_data_get(sql):

    # #DB接続定義
    db_url = 'postgresql://postgres:cimtops@ESRV10/irepodb'

    # エンジンを作成
    engine = create_engine(db_url, echo=True)

    # セッションを作成するためのSessionクラスを生成
    Session = sessionmaker(bind=engine)
    session = Session()

    # コネクションを取得
    with engine.connect() as connection:
      # SQLクエリの実行
      df = pd.read_sql(sql, connection)

    # セッションを閉じる
    session.close()   

    return(df)
  
  #sqlite3への接続
  def sqlite_data_get(sql,filepath):
      # #DB接続定義
      db_url = f'sqlite:///{filepath}'

      # エンジンを作成
      engine = create_engine(db_url, echo=True)

      # セッションを作成するためのSessionクラスを生成
      Session = sessionmaker(bind=engine)
      session = Session()

      # コネクションを取得
      with engine.connect() as connection:
        # SQLクエリの実行
        df = pd.read_sql(sql, connection)
  
      # セッションを閉じる
      session.close()   
  
      return(df)

  def irepo_jidoutyouhyou_create(df):

    #現品票未発行の対象があったとき
    df = df[(df['ireporter管理flg'] == 1)& (df['QRコードレイアウト区分'] == 0) & (df['帳票発行ID'] == 0)]
    err_exists = df.empty

    if not err_exists:
      #自動帳票作成用DFの作成
      dt_now = datetime.now(timezone(timedelta(hours=9))) # 日本時刻
      create_day = dt_now.strftime('%Y%m%d')

      df.insert(0, 'defTopId', '') 
      
      #縦型
      mask_df = df['縦横区分'] == 1
      sql = "select def_top_id from view_def_top where def_top_org = 405 and public_status = 2"
      irepo_df = ireporter_data_get(sql)
      df.loc[mask_df, 'defTopId'] = irepo_df['def_top_id'].max()

      #横型
      mask_df = df['縦横区分'] == 0
      sql = "select def_top_id from view_def_top where def_top_org = 406 and public_status = 2"
      irepo_df = ireporter_data_get(sql)
      df.loc[mask_df, 'defTopId'] = irepo_df['def_top_id'].max()

      #必要項目を並べる
      df1 = df.loc[:,['defTopId','部品名','ロット番号','完成部番','月次','組立番号','機種','吋','G','客先名','ﾒｯｷ','工程VERSION']]
      df1.loc[~(df1['ﾒｯｷ'] == 1), 'ﾒｯｷ'] = ''
      df1.loc[df1['ﾒｯｷ'] == 1, 'ﾒｯｷ'] = '無電解ニッケルメッキ'
      df1['作成日'] = create_day

      #先頭行に固定値をインサート
      df1.insert(0, 'H', 'R')
      df1.columns = ['H','defTopId','S1C1','S1C2','S1C3','S1C4','S1C5','S1C6','S1C7','S1C8','S1C9','S1C11','S1C644','S1C643']
      
      #CSVに変換
      csv_name = os.path.dirname(os.path.dirname(__file__))+"\\csv\\"+dt_now.strftime('%Y%m%d%H%M%S')+"_自動帳票作成データ"+".csv"
      df1.to_csv(csv_name,index=False,encoding='CP932')
      data_upload(csv_name,df1.shape[0])

  def main():

    st.set_page_config(
    page_title = '現品票手入力発行',
      layout="wide" )

    HIDE_ST_STYLE = """
                    <style>
                    div[data-testid="stToolbar"] {
                    visibility: hidden;
                    height: 0%;
                    position: fixed;
                    }
                    div[data-testid="stDecoration"] {
                    visibility: hidden;
                    height: 0%;
                    position: fixed;
                    }
                    #MainMenu {
                    visibility: hidden;
                    height: 0%;
                    }
                    header {
                    visibility: hidden;
                    height: 0%;
                    }
                    footer {
                    visibility: hidden;
                    height: 0%;
                    }
                    .appview-container .main .block-container{
                                padding-top: 1rem;
                                padding-right: 3rem;
                                padding-left: 3rem;
                                padding-bottom: 1rem;
                            }  
                            .reportview-container {
                                padding-top: 0rem;
                                padding-right: 3rem;
                                padding-left: 3rem;
                                padding-bottom: 0rem;
                            }
                            header[data-testid="stHeader"] {
                                z-index: -1;
                            }
                            div[data-testid="stToolbar"] {
                            z-index: 100;
                            }
                            div[data-testid="stDecoration"] {
                            z-index: 100;
                            }
                    .block-container {
                            padding-top: 0rem !important;
                            padding-bottom: 0rem !important;
                            }        
                    </style>
    """
    st.markdown(HIDE_ST_STYLE, unsafe_allow_html=True)
    
    sql = 'SELECT * FROM buhin_irepo_mst'
    filepath = os.path.dirname(os.path.dirname(__file__))+'\\Database\\genpinhyo.db'
    buhin_irepo_df = sqlite_data_get(sql,filepath)
    buhin_irepo_df = buhin_irepo_df.sort_values(['部品名'])
    buhin_list = buhin_irepo_df['部品名'].tolist()
    buhin_list.insert(0, "")

    #i-reporterより客先マスター取得
    sql = "SELECT record_key as 客先コード,value as 客先名,field0001 as 国名 FROM view_mst_custom_record WHERE master_key = 'M_CUSTOMER'"
    kyakusaki_df = ireporter_data_get(sql)

    hcol1, hcol2= st.columns([2,5])
    with hcol1:
      st.markdown("### 現品票手入力発行")
    st.markdown('''この画面は生産管理システムで管理していない部番の現品票を発行するためのものです。主にCMRで使用します。  
                年月、CMR番号、部品名称、完成部番を入力のうえ、「手入力現品票を印刷」から印刷してください。  
                部品が「i-Reporter対象」の場合は、i-Reporterのチェックリストも作成されます。''')
    st.divider()
    # # 水平線
    # st.divider()        
    # st.markdown("#### ①入力必須項目（QRコードの発行に必要な項目）")
    # st.markdown("ここに入力された文字がそのままチェックシートの上部（ヘッダー部）に表示されます。  \n \
    #             ※**完成部番**は図番と紐づいているので特に**正確な入力**が必要です。  \n \
    #             ※英数字は**半角**で入力してください。  \n \
    #             ※**客先コード**は画面下部の表を参考にご入力ください。")
    
    l = random.randint(100000,999999)
    lot = "KARI"+str(l)

    # 年月入力欄（年と月であればプルダウンでもよいかもしれない当年前年翌年、１～１２月）
    dt_now = datetime.now(timezone(timedelta(hours=9)))+ timedelta(days=32) # 日本時刻+32日
    dt_nen = int(dt_now.strftime('%Y'))
    dt_tsuki = int(dt_now.strftime('%m'))

    acol1,acol2,acol3,acol4,acol5 = st.columns([1,1,1,1,2])
    with acol1:
      nen = st.selectbox('年',[dt_nen-1,dt_nen,dt_nen+1],index=1)
    with acol2:
      getsu = st.selectbox('月',['01','02','03','04','05','06','07','08','09','10','11','12'],index=dt_tsuki-1)
    with acol3:
      #ji = st.selectbox('次',['0','1','2','3','4','5','6','7','8','9','_'],index=1)
      ji = '_'
    #with acol4:
      #kouteiver = st.number_input('工程version', min_value=1, max_value=99,value=1 ,step=1)
      #kouteiver = 1
    
    # bcol1,bcol2,bcol3,bcol4= st.columns([2,2,1,1]) 
    # with bcol1:
    #   kumi = st.text_input('組立番号（例：K038）', '')
    # with bcol2:
    #   kisyu = st.text_input('機種（例：MAD-6ER11）', '')
    # with bcol3:
    #   inchi = st.text_input('吋', '')
    # with bcol4:
    #   guaji = st.text_input('ｹﾞｰｼﾞ', '')
    kumi = ''
    kisyu = ''
    inchi = ''
    guaji = ''
    a2col1,_= st.columns([1,3]) 
    with a2col1:
      kumi = st.text_input('CMR番号（例：CMR-24003）', '')
      if kumi == '':
        kumi = '_'

    bcol1,bcol2,_= st.columns([1,1,3]) 
    with bcol1:
      buhin = st.selectbox('部品名称',buhin_list)
    with bcol2:
      iflg_df = buhin_irepo_df[buhin_irepo_df['部品名'] == buhin]
      qr_kubun = iflg_df[['QRコードレイアウト区分']]      
      tateyoko_kubun = iflg_df[['縦横区分']]
      iflg_df = iflg_df[['ireporter管理flg']]
      if not iflg_df.empty:
        iflg = iflg_df['ireporter管理flg'].iloc[0]
        if iflg == 0:
          st.write(" ")
          st.write(" ")
          st.markdown("##### i-Reporter対象外")
        else:
          st.write(" ")
          st.write(" ")
          st.markdown("##### i-Reporter対象")     

    iflg = 0
    ccol1,ccol2,ccol3,ccol4,ccol5= st.columns([2.5,2.5,1.3,1.5,1.5]) 
    with ccol1:
      buban = st.text_input('完成部番 (例：565-602BA33)', '')
    with ccol2:
      kouteiver = st.number_input('工程version（不明な場合は 1）', min_value=1, max_value=99,value=1 ,step=1)
    #   buhin = st.selectbox('部品名称',buhin_list)
    # with ccol3:
    #   iflg_df = buhin_irepo_df[buhin_irepo_df['部品名'] == buhin]
    #   qr_kubun = iflg_df[['QRコードレイアウト区分']]      
    #   tateyoko_kubun = iflg_df[['縦横区分']]
    #   iflg_df = iflg_df[['ireporter管理flg']]
    #   if not iflg_df.empty:
    #     iflg = iflg_df['ireporter管理flg'].iloc[0]
    #     if iflg == 0:
    #       st.write(" ")
    #       st.write(" ")
    #       st.markdown("##### i-Reporter対象外")
    #     else:
    #       st.write(" ")
    #       st.write(" ")
    #       st.markdown("##### i-Reporter対象")        
    #with ccol4:
      #mekki = st.number_input('ﾆｯｹﾙﾒｯｷ有は"1"に変更', min_value=0, max_value=1, value=0, step=1)
    mekki = 0


    

    #st.divider()
    #st.markdown("#### ②現品票記載項目（入力必須ではありません）")
    #st.markdown("QRコードの内容には影響はありませんが、現品票に表示される項目です。")
    
    # dcol1,dcol2,dcol3= st.columns([1,4,2]) 
    # with dcol1:
    #   suu = st.number_input('必要数（例：1）', min_value=1, max_value=999, value=1, step=1)
    # with dcol2:
    #   koutei= st.text_input('工程（例：L1-L2-TH1-SL1）', '')
    suu = 1
    #koutei= st.text_input('工程（例：L1-L2-TH1-SL1）', '')
    
    
    b1 = ""
    b2 = ""
    f1 = ""
    f2 = ""
    l1 = ""
    r1 = ""
    c1 = ""
    koutei = ""
    if buban != "":  
      sql = "SELECT FINAL_ITEM_CODE,VERSION,PROCODESTR FROM MSTR_PROCODESTR"
      df_koutei = krd_data_get(sql)
      df_koutei = df_koutei[df_koutei['FINAL_ITEM_CODE']==buban]
      df_koutei = df_koutei[df_koutei['VERSION']==kouteiver]
      df_koutei = df_koutei.fillna('')
      if len(df_koutei) == 1:
        st.write(f"工程情報：{df_koutei['PROCODESTR'].iloc[0]}")
        koutei = df_koutei['PROCODESTR'].iloc[0]
      else:
        st.write("工程情報：無し")
        koutei = ""
      #st.divider()

      #工程図番の取得 ADD_20241115_久原さんより
      sql = "SELECT * FROM DATA_KOUTEIZUKAN"
      df_kouteizu = krd_data_get(sql)
      df_kouteizu = df_kouteizu[df_kouteizu['SETU_F']==buban]
      df_kouteizu = df_kouteizu.fillna('')
      if len(df_kouteizu) == 1:
        st.write("図番情報：")
        b1 = df_kouteizu['B_FIG'].iloc[0]
        b2 = df_kouteizu['B2_FIG'].iloc[0]
        f1 = df_kouteizu['F_FIG'].iloc[0]
        f2 = df_kouteizu['F2_FIG'].iloc[0]
        l1 = df_kouteizu['L_FIG'].iloc[0]
        r1 = df_kouteizu['R_FIG'].iloc[0]
        c1 = df_kouteizu['C_FIG'].iloc[0]
        df_kouteizu2 = df_kouteizu[['SETU_F','F_FIG','F2_FIG','B_FIG','B2_FIG','L_FIG','R_FIG','C_FIG','M_FIG']].rename(columns={'SETU_F':'完成部番','F_FIG': 'F1_図番','F2_FIG': 'F2_図番','B_FIG': 'B1_図番','B2_FIG': 'B2_図番','L_FIG': 'L_図番','R_FIG': 'R_図番','C_FIG': 'C_図番','M_FIG': 'M_図番'})
        st.dataframe(df_kouteizu2,hide_index=True)
      else:
        st.write("図番情報：無し") 
      
    
    #  df_kouteizu2 = pd.merge(merged_df3, df_kouteizu, left_on=['完成部番'],right_on=['SETU_F'],how='left')
    #  df_kouteizu2 = df_kouteizu2[['完成部番','F_FIG','F2_FIG','B_FIG','B2_FIG','L_FIG','R_FIG','C_FIG','M_FIG']].rename(columns={'F_FIG': 'F1_図番','F2_FIG': 'F2_図番','B_FIG': 'B1_図番','B2_FIG': 'B2_図番','L_FIG': 'L_図番','R_FIG': 'R_図番','C_FIG': 'C_図番','M_FIG': 'M_図番'})
    #  df_kouteizu2 = df_kouteizu2.fillna('')
    #  df_kouteizu2 = df_kouteizu2.drop_duplicates() #重複削除
    #  st.dataframe(df_kouteizu2,hide_index=True,use_container_width=True)
   # 図番は差し替え
   # ecol1,ecol2,ecol3,ecol4,ecol5,ecol6,ecol7= st.columns([1,1,1,1,1,1,1]) 
   # with ecol1:
   #   b1 = st.text_input('B1図番', '')    
   # with ecol2:
   #   b2 = st.text_input('B2図番', '')    
   # with ecol3:
   #   f1 = st.text_input('F1図番', '')    
   # with ecol4:
   #   f2 = st.text_input('F2図番', '')    
   # with ecol5:
   #   l1 = st.text_input('L 図番', '')    
   # with ecol6:
   #   r1 = st.text_input('R 図番', '')    
   # with ecol7:
   #   c1 = st.text_input('C 図番', '')
       #ADD_20241115 
        #krdよりプロセスコード取得

    
    c2col1,c2col2,c2col3= st.columns([2,4,1])
    with c2col1:
      kyakusaki = st.text_input('客先コード（特に指定のない場合はJ0003）', value="J0003")
    with c2col2:
      kyakusakimei = ""
      kyakusaki_df2 = kyakusaki_df[kyakusaki_df['客先コード']==kyakusaki]
      if not kyakusaki_df2.empty:
        kyakusakimei = kyakusaki_df2['客先名'].iloc[0]
      if kyakusakimei != "":
        st.write(" ")
        st.write(" ")
        st.markdown(f"##### 客先名： {kyakusakimei}") 

    st.divider()
    
    fcol1,fcol2= st.columns([2,3]) 
    with fcol1:
      if st.button("手入力現品票を印刷"):
        with st.spinner('現品票作成処理　実行中'):
          dt = datetime.now(timezone(timedelta(hours=9))) # 日本時刻
          dt_now = dt.strftime('%Y/%m/%d') 

          #自動帳票作成処理
          if iflg == 0: #irepo対象外
            tyohyoID = 0
          else: #対象

            list1=[[lot,buhin,buban,str(nen)+str(getsu)+str(ji),kumi,kisyu,inchi,guaji,kyakusaki,mekki,kouteiver,iflg,qr_kubun['QRコードレイアウト区分'].iloc[0],tateyoko_kubun['縦横区分'].iloc[0],0]]
            columns1 =['ロット番号','部品名','完成部番','月次','組立番号','機種','吋','G','客先コード','ﾒｯｷ','工程VERSION','ireporter管理flg','QRコードレイアウト区分','縦横区分','帳票発行ID']
            df = pd.DataFrame(data=list1, columns=columns1)

            #客先マスターを結合
            sql = "SELECT record_key as 客先コード,value as 客先名 FROM view_mst_custom_record WHERE master_key = 'M_CUSTOMER'"
            kyakusaki_df = ireporter_data_get(sql)
            df = pd.merge(df, kyakusaki_df, how = 'left')

            with st.spinner(f'iReporter帳票作成対象のため、50秒ほどお待ちください'):
              irepo_jidoutyouhyou_create(df)
              time.sleep(30) #待機しないとDB更新前に帳票を発行してしまう可能性がある

              #帳票ID再取得
              sql = f"select rep_top_id,cluster_1_2_t from view_report_406" #横レイアウト
              df3 = ireporter_data_get(sql)
              sql = f"select rep_top_id,cluster_1_2_t from view_report_405" #縦レイアウト
              df4 = ireporter_data_get(sql)
              df5 = pd.concat([df3,df4], ignore_index=True)
              df = pd.merge(df, df5, left_on=['ロット番号'],right_on=['cluster_1_2_t'],how='left')
              tyohyoID = df['rep_top_id'].iloc[0]

          #QRコード作成処理
          if (qr_kubun['QRコードレイアウト区分'] == 1).any():
            #QRコード作成(旧レイアウトでの作成)
            code = pyqrcode.create(f"{lot},{str(nen)}{str(getsu)}{str(ji)},{kumi},{kisyu},{inchi},{guaji},{buban},{buhin},'',{kyakusaki},{mekki}",error='L', version=5, mode='binary')
          else:
            code = pyqrcode.create(f"jp.co.cimtops.ireporter.openreport:repid={tyohyoID}", error='L', version=5, mode='binary')
          code.png(os.path.dirname(__file__)+f"\\work\\qrcode_{lot}_{dt.strftime('%H%M%S')}.png", scale=3)
          #余白を透明化
          img = Image.open(os.path.dirname(__file__)+f"\\work\\qrcode_{lot}_{dt.strftime('%H%M%S')}.png")
          img = img.convert("RGBA")
          datas = img.getdata()
          newData = []
          for item in datas:
              if item[0] == 255 and item[1] == 255 and item[2] == 255:
                  newData.append((255, 255, 255, 0))
              else:
                  newData.append(item)
          img.putdata(newData)
          img.save(os.path.dirname(__file__)+f"\\work\\qrcode_{lot}_{dt.strftime('%H%M%S')}.png", "PNG")

          #アウトプットのエクセルの準備
          dt_now = datetime.now(timezone(timedelta(hours=9))) # 日本時刻
          wb = os.path.dirname(__file__)+"\\add_genpinhyo\\手入力現品票_"+dt_now.strftime('%Y%m%d%H%M%S') +'.xlsx'
          wb_name ="手入力現品票_"+dt_now.strftime('%Y%m%d%H%M%S') +'.xlsx'
          shutil.copy(os.path.dirname(__file__)+'\\template\\add_template.xlsx', wb)

          out_wb = openpyxl.load_workbook(wb)
          out_wb.active.title="手入力現品票"
          out_wb.save(wb)

          if mekki == 1:
            chuui = '無電解ニッケルメッキ'
          else:
            chuui = ''
          #格納処理  
          out_wb.active.cell(2, 10).value = chuui #無電解ニッケルﾒｯｷ
          out_wb.active.cell(3, 1).value = dt.strftime('%Y/%m/%d') #発行日
          out_wb.active.cell(4, 21).value = suu #個数
          out_wb.active.cell(5, 7).value = buban #完成部番
          out_wb.active.cell(8, 7).value = b1 #B1部番
          out_wb.active.cell(9, 7).value = b2 #B2部番
          out_wb.active.cell(10, 7).value = f1 #F1部番
          out_wb.active.cell(11, 7).value = f2 #F2部番
          out_wb.active.cell(12, 7).value = l1 #L部番
          out_wb.active.cell(13, 7).value = r1 #R部番
          out_wb.active.cell(14, 7).value = c1 #C部番
          out_wb.active.cell(1, 30).value = lot #lot番号
          out_wb.active.cell(4, 24).value = f"{getsu}月{ji}次" #月次
          out_wb.active.cell(7, 24).value = kumi #組立番号
          out_wb.active.cell(10, 24).value = kisyu #機種名
          img_to_excel = openpyxl.drawing.image.Image(os.path.dirname(__file__)+f"\\work\\qrcode_{lot}_{dt.strftime('%H%M%S')}.png") #QRコード
          out_wb.active.add_image(img_to_excel, f'AI{3}')
          out_wb.active.cell(13, 32).value = inchi #インチ
          out_wb.active.cell(13, 38).value = guaji #ゲージ
          out_wb.active.cell(15, 6).value = koutei #工程

          #エクセルを保存
          out_wb.save(wb)
          out_wb.close()

          # ファイルを読み込む
          with open(wb, 'rb') as file:
            filedata = file.read()
          st.markdown('処理終了：処理結果のエクセルを以下からダウンロードしてください')
          st.download_button(
            label='処理結果ダウンロード',
            data=filedata,
            file_name=wb_name,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
          st.cache_resource.clear() 
      # with fcol2:
      #   st.markdown("##### 【参考　客先コード確認用】")

      #   #データフレームを表示（インデックスは非表示）
      #   st.dataframe(kyakusaki_df,hide_index=True)
          
  if __name__ == "__main__":
    main()

except Exception as e:
  #簡単なエラー処理
  print(e)
  dt_now = datetime.now(timezone(timedelta(hours=9))) # 日本時刻
  with open(os.path.dirname(__file__)+"\\"+dt_now.strftime('%Y%m%d%H%M%S')+"_err"+".txt", mode='w') as f:
    f.write(str(e))
