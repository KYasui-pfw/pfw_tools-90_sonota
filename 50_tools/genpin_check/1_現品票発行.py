################################
# 　新現品票作成　　　　　　　      #
#  20241124:削除フラグの考え方追加 #
#  20250122:新規部品名検知のためflgの初期値変更 #
#  20250410:CSV処理の追加
################################
# インポート
import streamlit as st
import streamlit.components.v1 as components
import os
from glob import glob
import time
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import pythoncom
from datetime import datetime, timedelta, timezone
import pandas as pd
import shutil
import pyqrcode
import openpyxl
from PIL import Image
from st_aggrid import AgGrid, JsCode, GridUpdateMode
from st_aggrid.grid_options_builder import GridOptionsBuilder
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import sqlite3
from module.conmas_upload import data_upload
from dateutil.relativedelta import relativedelta


try:
    pythoncom.CoInitialize()  # サーバーサイドからローカルファイルを動かすときに必要

    # ireporterDB
    @st.cache_resource
    def ireporter_data_get(sql):

        # #DB接続定義
        db_url = 'postgresql://postgres:cimtops@ESRV10/irepodb'

        # エンジンを作成
        engine = create_engine(db_url, echo=True)

        # セッションを作成するためのSessionクラスを生成
        Session = sessionmaker(bind=engine)
        session = Session()

        # コネクションを取得
        with engine.connect() as connection:
            # SQLクエリの実行
            df = pd.read_sql(sql, connection)

        # セッションを閉じる
        session.close()

        return (df)

    # krdのmachinDBに接続する（SQLite版）
    def krd_data_get(sql):
        # SQLite接続（KRD MySQL → SQLite同期データベース）
        # \\esrv11\krd_machine\db\krd_machine.db
        sqlite_db_path = r'\\esrv11\krd_machine\db\krd_machine.db'

        conn = sqlite3.connect(sqlite_db_path)
        df = pd.read_sql(sql, conn)
        conn.close()

        return df

    # # 【旧版：MySQL接続】コメントアウト（2025-11-22 SQLite移行）
    # def krd_data_get(sql):
    #     # #DB接続定義
    #     db_url = 'mysql+pymysql://pfw:mejiriHoo@krd/machin?charset=utf8'
    #
    #     # エンジンを作成
    #     engine = create_engine(db_url, echo=True)
    #
    #     # セッションを作成するためのSessionクラスを生成
    #     Session = sessionmaker(bind=engine)
    #     session = Session()
    #
    #     # コネクションを取得
    #     with engine.connect() as connection:
    #         # SQLクエリの実行
    #         df = pd.read_sql(sql, connection)
    #
    #     # セッションを閉じる
    #     session.close()
    #
    #     return (df)

    # sqlite3への接続
    def sqlite_data_get(sql, filepath):
        # #DB接続定義
        db_url = f'sqlite:///{filepath}'

        # エンジンを作成
        engine = create_engine(db_url, echo=True)

        # セッションを作成するためのSessionクラスを生成
        Session = sessionmaker(bind=engine)
        session = Session()

        # コネクションを取得
        with engine.connect() as connection:
            # SQLクエリの実行
            df = pd.read_sql(sql, connection)

        # セッションを閉じる
        session.close()

        return (df)

    # df⇒SQLiteに記録（旧：CSV出力）
    def df_csv_cnv(df, filename):
        # DataFrameをSQLite genpinhyo.db の output_history テーブルに保存
        dt_now = datetime.now(timezone(timedelta(hours=9)))  # 日本時刻
        db_path = os.path.join(os.path.dirname(__file__), 'Database', 'genpinhyo.db')

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # output_history テーブルを作成（存在しない場合）
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS output_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                filename TEXT NOT NULL,
                data TEXT NOT NULL
            )
        ''')

        # DataFrameをJSON形式に変換して保存
        import json
        data_json = df.to_json(orient='records', force_ascii=False)

        cursor.execute('''
            INSERT INTO output_history (created_at, filename, data)
            VALUES (?, ?, ?)
        ''', (dt_now.strftime('%Y-%m-%d %H:%M:%S'), filename, data_json))

        conn.commit()
        conn.close()

        # # 【旧版：CSV出力】コメントアウト（2025-11-22 SQLite移行）
        # csv_name = os.path.dirname(__file__)+"\\"+dt_now.strftime('%Y%m%d%H%M%S')+"_"+filename+".csv"
        # df.to_csv(csv_name, index=False, encoding='CP932')

    def db_update():

        # バックアップ
        dt_now = datetime.now(timezone(timedelta(hours=9)))  # 日本時刻
        if (os.path.isfile(os.path.dirname(__file__)+'\\Database\\genpinhyo.db')):
            shutil.copy(os.path.dirname(__file__)+'\\Database\\genpinhyo.db',
                        os.path.dirname(__file__)+f'\\Database\\backup\\{dt_now.strftime('%Y%m%d')}_genpinhyo.db')

        # 本処理（更新処理）
        dbname = 'genpinhyo.db'
        cdb = os.path.dirname(__file__)+f'\\Database\\'+dbname
        # conn = sqlite3.connect(cdb,isolation_level=None) オートコミットを削除（重かった原因）
        conn = sqlite3.connect(cdb)

        cur = conn.cursor()

        # ■テーブル作成処理
        # EXIT NOTを入れているので、テーブルがあった場合はスキップする
        cur.execute('CREATE TABLE IF NOT EXISTS genpinhyo(\
                  ロット番号 TEXT PRIMARY KEY ,\
                  月次 TEXT,\
                  組立番号 TEXT,\
                  機種 TEXT,\
                  吋 TEXT,\
                  G TEXT,\
                  完成部番 TEXT,\
                  部品名 TEXT,\
                  客先名 TEXT,\
                  国名 TEXT,\
                  B1図番 TEXT,\
                  B2図番 TEXT,\
                  F1図番 TEXT,\
                  F2図番 TEXT,\
                  L図番 TEXT,\
                  R図番 TEXT,\
                  C図番 TEXT,\
                  組立開始日 TEXT,\
                  梱包開始日 TEXT,\
                  指定納期 TEXT,\
                  必要数 TEXT,\
                  SKDK TEXT,\
                  工程 TEXT,\
                  工程VERSION TEXT,\
                  ﾒｯｷ TEXT,\
                  客先コード TEXT\
                  )')
        # ■テーブル作成処理
        # EXIT NOTを入れているので、テーブルがあった場合はスキップする
        cur.execute('CREATE TABLE IF NOT EXISTS buhin_irepo_mst(\
                  部品名 TEXT PRIMARY KEY,\
                  ireporter管理flg INTEGER,\
                  QRコードレイアウト区分 INTEGER,\
                  縦横区分 INTEGER\
                  )')

        # ■テーブル作成処理 ADD_20241125
        # EXIT NOTを入れているので、テーブルがあった場合はスキップする
        cur.execute('CREATE TABLE IF NOT EXISTS delete_mst(\
                  ロット番号 TEXT PRIMARY KEY ,\
                  削除flg INTEGER \
                  )')

        # ▼▼データの取得と結合処理開始
        # EJから出力されたASPKakouDenpyoを取得
        denpyo_df = pd.read_csv(
            r"\\172.17.107.102\PrintOutCsv\4.加工\4-03 ASPKakouDenpyo.csv", encoding='CP932')

        # ADD_20241217_0件dataなら抜ける
        if denpyo_df.empty:
            return

        # i-reporterより客先マスター取得
        sql = "SELECT record_key as 客先コード,value as 客先名m,field0001 as 国名m FROM view_mst_custom_record WHERE master_key = 'M_CUSTOMER'"
        kyakusaki_df = ireporter_data_get(sql)

        # 製番をキーに、両dfの内容を結合
        df1 = pd.merge(denpyo_df, kyakusaki_df, how='left')

        # krdより工程バージョン取得
        sql = "SELECT SLIP_NO as 伝票Ｎｏ,VERSION FROM DATA_ASP2_PUT"
        krd_df1 = krd_data_get(sql)

        # 伝票Noをキーに、両dfの内容を結合
        df2 = pd.merge(df1, krd_df1, how='left')

        # krdよりプロセスコード取得
        sql = "SELECT FINAL_ITEM_CODE as 加工部番,VERSION,PROCODESTR as 工程 FROM MSTR_PROCODESTR"
        krd_df2 = krd_data_get(sql)

        # 加工部番をキーに、両dfの内容を結合
        df3 = pd.merge(df2, krd_df2, how='left')

        # krdより図番取得
        sql = "SELECT SETU_F as 加工部番,\
            B_FIG as B1図番,\
            B2_FIG as B2図番,\
            F_FIG as F1図番,\
            F2_FIG as F2図番,\
            L_FIG as L図番,\
            R_FIG as R図番,\
            C_FIG as C図番\
            FROM DATA_KOUTEIZUKAN"
        krd_df3 = krd_data_get(sql)

        # 加工部番をキーに、両dfの内容を結合
        df4 = pd.merge(df3, krd_df3, how='left')

        # krdより無電解ニッケル電気の対応を取得
        sql = "SELECT FIN_CODE as 加工部番,METAL FROM MSTR_METAL"
        krd_df4 = krd_data_get(sql)

        # 加工部番をキーに、両dfの内容を結合
        merged_df = pd.merge(df4, krd_df4, how='left')

        # ▲▲データの取得と結合処理終了

        # ▼▼dataframeを整える
        mdf = merged_df.filter(items=['伝票Ｎｏ', '生産月次', '組立番号', '機種名', '吋', 'ゲージ', '加工部番', '部品名', '客先名m', '国名m', 'B1図番', 'B2図番',
                               'F1図番', 'F2図番', 'L図番', 'R図番', 'C図番', '組立開始日', '梱包開始日', '指定納期', '必要数', 'SKDK', '工程', 'VERSION', 'METAL', '客先コード'])

        # カラム名付け替え・編集
        mdf = mdf.rename(columns={'伝票Ｎｏ': 'ロット番号'})
        mdf = mdf.rename(columns={'生産月次': '月次'})
        mdf = mdf.rename(columns={'機種名': '機種'})
        mdf = mdf.rename(columns={'ゲージ': 'G'})
        mdf = mdf.rename(columns={'加工部番': '完成部番'})
        mdf = mdf.rename(columns={'客先名m': '客先名'})
        mdf = mdf.rename(columns={'国名m': '国名'})
        mdf = mdf.rename(columns={'VERSION': '工程VERSION'})
        mdf = mdf.rename(columns={'METAL': 'ﾒｯｷ'})
        mdf['吋'] = mdf['吋'].astype('object')
        mdf['G'] = mdf['G'].astype('object')
        mask = (mdf['吋'] != '') & mdf['吋'].apply(
            lambda x: float(x).is_integer() if x != '' else False)
        mdf.loc[mask, '吋'] = mdf.loc[mask, '吋'].apply(
            lambda x: str(int(float(x))))
        mask = (mdf['G'] != '') & mdf['G'].apply(
            lambda x: float(x).is_integer() if x != '' else False)
        mdf.loc[mask, 'G'] = mdf.loc[mask, 'G'].apply(
            lambda x: str(int(float(x))))
        mdf['月次'] = mdf['月次'].astype(str)
        mdf['月次'] = mdf['月次'].apply(
            lambda x: x + '_' if len(x) == 6 else x)  # 6文字は_を追加する
        mdf['完成部番'] = mdf['完成部番'].str.replace('　', ' ')
        mdf['組立番号'] = mdf['組立番号'].str.replace('　', ' ')
        mdf['機種'] = mdf['機種'].str.replace('　', ' ')
        mdf['国名'] = mdf['国名'].str.replace('　', ' ')
        mdf['客先名'] = mdf['客先名'].str.replace('　', ' ')
        mdf = mdf.fillna('')  # Noneデータ対策

        # 部品irepotableの更新用
        buhin_irepo_df = denpyo_df[['部品名']]  # 必要な列だけ抽出
        buhin_irepo_df['i-reporter管理flg'] = 0  # フラグの初期値設定
        # buhin_irepo_df['QRコードレイアウト区分'] = 0  # フラグの初期値設定 DEL_20250122_新規部品名検知のため、flgは2
        # フラグの初期値設定 ADD_20250122_新規部品名検知のため、flgは2
        buhin_irepo_df['QRコードレイアウト区分'] = 2
        buhin_irepo_df['縦横区分'] = 0  # フラグの初期値設定
        buhin_irepo_df = buhin_irepo_df.drop_duplicates()  # 重複削除
        # データベースを更新
        try:    # mdfの全データを一括でリストに変換
            data = [tuple(row.values.tolist())
                    for index, row in mdf.iterrows()]
            # genpinhyoの更新　1000件ずつ分割して処理
            batch_size = 1000
            for i in range(0, len(data), batch_size):
                batch = data[i:i+batch_size]  # 1000件ずつスライス
                cur.executemany('REPLACE INTO genpinhyo(ロット番号,月次,組立番号,機種,吋,G,完成部番,部品名,客先名,国名,B1図番,B2図番,F1図番,F2図番,L図番,\
                          R図番,C図番,組立開始日,梱包開始日,指定納期,必要数,SKDK,工程,工程VERSION,ﾒｯｷ,客先コード)\
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', batch)
            conn.commit()  # 20241212_コミット漏れ追加
            # buhin_irepo_mstの更新
            for index, row in buhin_irepo_df.iterrows():
                buhinmei = row['部品名']
                i_reporter_k_flg = row['i-reporter管理flg']
                qr_kubun = row['QRコードレイアウト区分']
                tateyoko_kubun = row['縦横区分']

                # 部品名が既に存在するか確認し、存在しなければ挿入
                cur.execute('''
            INSERT INTO buhin_irepo_mst (部品名, ireporter管理flg, QRコードレイアウト区分, 縦横区分)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(部品名) DO NOTHING
        ''', (buhinmei, i_reporter_k_flg, qr_kubun, tateyoko_kubun))

            conn.commit()  # 20241212_コミット漏れ追加
            # ADD_20241125_start delete_mst更新を追加
            # DB接続定義
            dbname = 'genpinhyo.db'
            filepath = os.path.dirname(__file__)+f'\\Database\\'+dbname
            db_url = f'sqlite:///{filepath}'
            engine = create_engine(db_url, echo=True)

            # セッションを作成するためのSessionクラスを生成
            Session = sessionmaker(bind=engine)
            session = Session()

            # コネクションを取得
            with engine.connect() as connection:
                # SQLクエリの実行
                del_df = pd.read_sql(f'select * from genpinhyo', connection)
            # セッションを閉じる
            session.close()

            del_df = del_df.loc[:, ['ロット番号']]

            for index, row in del_df.iterrows():

                # 部品名が既に存在するか確認し、存在しなければ挿入
                cur.execute('''
            INSERT INTO delete_mst(ロット番号,削除flg)
            VALUES (?, ?)
            ON CONFLICT(ロット番号) DO NOTHING
        ''', (row['ロット番号'], 0))
            # ADD_20241125_end

            conn.commit()
        except Exception as e:
            # エラーが発生した場合、ロールバックして変更を取り消す
            conn.rollback()
        finally:
            # 最後にカーソルと接続を閉じる
            cur.close()
            conn.close()

        # ADD_20241125_削除flg更新
        # DB接続定義 同じですが、明示します
        dbname = 'genpinhyo.db'
        filepath = os.path.dirname(__file__)+f'\\Database\\'+dbname
        db_url = f'sqlite:///{filepath}'
        engine = create_engine(db_url, echo=True)

        # セッションを作成するためのSessionクラスを生成
        Session = sessionmaker(bind=engine)
        session = Session()

        # コネクションを取得
        with engine.connect() as connection:
            # SQLクエリの実行
            df1 = pd.read_sql(f'select * from genpinhyo', connection)
            df2 = pd.read_sql(f'select * from delete_mst', connection)
        # セッションを閉じる
        session.close()

        delete_df = pd.merge(df1, df2, on=['ロット番号'], how='left')

        # 年月で絞り込み
        dt_now = datetime.now(timezone(timedelta(hours=9)))  # 日本時刻
        next_month = dt_now + relativedelta(months=1)  # 翌月
        next2_month = dt_now + relativedelta(months=2)  # 翌翌月
        next3_month = dt_now + relativedelta(months=3)  # 翌翌翌月
        nengetsu1 = dt_now.strftime('%Y%m')
        nengetsu2 = next_month.strftime('%Y%m')
        nengetsu3 = next2_month.strftime('%Y%m')
        nengetsu4 = next3_month.strftime('%Y%m')
        dfa = delete_df[delete_df['月次'].str.startswith(nengetsu1)]
        dfb = delete_df[delete_df['月次'].str.startswith(nengetsu2)]
        dfc = delete_df[delete_df['月次'].str.startswith(nengetsu3)]
        dfd = delete_df[delete_df['月次'].str.startswith(nengetsu4)]

        delete_df = pd.concat([dfa, dfb, dfc, dfd], ignore_index=True)
        delete_df = delete_df.loc[:, ['ロット番号', '削除flg']]

        # csvを起源としたdf をdelete_df2にセット
        delete_df2 = mdf.loc[:, ['ロット番号']]
        delete_df2['削除flg2'] = 0

        # マージすることにより、削除データをあぶりだす（EJ特有の処理）
        delete_df = pd.merge(delete_df, delete_df2, on=['ロット番号'], how='left')
        delete_df = delete_df.fillna(1)  # Noneデータを1に変更

        # ADD_20241217_DEL対象の確認処理を追加
        chk_df = delete_df[delete_df['削除flg'].astype(
            int) != delete_df['削除flg2'].astype(int)]
        if not chk_df.empty:
            df_csv_cnv(chk_df, "EJ削除flg_on")

        delete_df = delete_df.drop(columns=['削除flg'])
        delete_df = delete_df.rename(columns={'削除flg2': '削除flg'})

        # 本処理（更新処理）
        dbname = 'genpinhyo.db'
        cdb = os.path.dirname(__file__)+f'\\Database\\'+dbname
        conn = sqlite3.connect(cdb)

        cur = conn.cursor()

        try:    # mdfの全データを一括でリストに変換

            data = [tuple(row)
                    for row in delete_df[['削除flg', 'ロット番号']].to_numpy()]

            batch_size = 1000
            for i in range(0, len(data), batch_size):
                # 1000件ずつ分割して処理
                batch = data[i:i+batch_size]  # 1000件ずつスライス
                # 部品名が既に存在するか確認し、存在しなければ挿入
                cur.executemany('''
            UPDATE delete_mst
            SET 削除flg = ?
            WHERE ロット番号 = ?
        ''', batch)

            conn.commit()
        except Exception as e:
            # エラーが発生した場合、ロールバックして変更を取り消す
            conn.rollback()
        finally:
            # 最後にカーソルと接続を閉じる
            cur.close()
            conn.close()

        # ADD_20241126_end

    # sqliteから取得
    @st.cache_resource
    def df_set1(nen, getsu):

        # #DB接続定義
        dbname = 'genpinhyo.db'
        filepath = os.path.dirname(__file__)+f'\\Database\\'+dbname
        db_url = f'sqlite:///{filepath}'

        # エンジンを作成
        engine = create_engine(db_url, echo=True)

        # セッションを作成するためのSessionクラスを生成
        Session = sessionmaker(bind=engine)
        session = Session()

        # コネクションを取得
        with engine.connect() as connection:
            # SQLクエリの実行
            nengetsu = str(nen)+str(getsu)
            df = pd.read_sql(
                f'select * from genpinhyo where 月次 LIKE "{nengetsu}%"', connection)

        # セッションを閉じる
        session.close()

        return (df)

    @st.cache_resource
    def df_set2():

        # #DB接続定義
        dbname = 'genpinhyo.db'
        filepath = os.path.dirname(__file__)+f'\\Database\\'+dbname
        db_url = f'sqlite:///{filepath}'

        # エンジンを作成
        engine = create_engine(db_url, echo=True)

        # セッションを作成するためのSessionクラスを生成
        Session = sessionmaker(bind=engine)
        session = Session()

        # コネクションを取得
        with engine.connect() as connection:
            # SQLクエリの実行
            df = pd.read_sql(f'select * from buhin_irepo_mst', connection)

        # セッションを閉じる
        session.close()

        return (df)

    @st.cache_resource
    def df_set3():

        # #DB接続定義
        dbname = 'genpinhyo.db'
        filepath = os.path.dirname(__file__)+f'\\Database\\'+dbname
        db_url = f'sqlite:///{filepath}'

        # エンジンを作成
        engine = create_engine(db_url, echo=True)

        # セッションを作成するためのSessionクラスを生成
        Session = sessionmaker(bind=engine)
        session = Session()

        # コネクションを取得
        with engine.connect() as connection:
            # SQLクエリの実行
            df = pd.read_sql(f'select * from delete_mst', connection)

        # セッションを閉じる
        session.close()

        return (df)

    def irepo_jidoutyouhyou_create(df):

        # 現品票未発行の対象があったとき　
        df = df[(df['ireporter管理flg'] == 1) & (
            df['QRコードレイアウト区分'] == 0) & (df['帳票発行ID'] == 0)]
        err_exists = df.empty

        if not err_exists:
            # 自動帳票作成用DFの作成
            dt_now = datetime.now(timezone(timedelta(hours=9)))  # 日本時刻
            create_day = dt_now.strftime('%Y%m%d')

            df.insert(0, 'defTopId', '')

            # 縦型
            mask_df = df['縦横区分'] == 1
            sql = "select def_top_id from view_def_top where def_top_org = 405 and public_status = 2"
            irepo_df = ireporter_data_get(sql)
            df.loc[mask_df, 'defTopId'] = irepo_df['def_top_id'].max()

            # 横型
            mask_df = df['縦横区分'] == 0
            sql = "select def_top_id from view_def_top where def_top_org = 406 and public_status = 2"
            irepo_df = ireporter_data_get(sql)
            df.loc[mask_df, 'defTopId'] = irepo_df['def_top_id'].max()

            # 必要項目を並べる
            df1 = df.loc[:, ['defTopId', '部品名', 'ロット番号', '完成部番', '月次',
                             '組立番号', '機種', '吋', 'G', '客先名', 'ﾒｯｷ', '工程VERSION']]
            df1.loc[~(df1['ﾒｯｷ'] == '1.0'), 'ﾒｯｷ'] = ''
            df1.loc[df1['ﾒｯｷ'] == '1.0', 'ﾒｯｷ'] = '無電解ニッケルメッキ'
            df1['作成日'] = create_day

            # 先頭行に固定値をインサート
            df1.insert(0, 'H', 'R')
            df1.columns = ['H', 'defTopId', 'S1C1', 'S1C2', 'S1C3', 'S1C4',
                           'S1C5', 'S1C6', 'S1C7', 'S1C8', 'S1C9', 'S1C11', 'S1C644', 'S1C643']

            # df1 = df1.head(5) ##上から５行（テスト用設定）

            # CSVに変換
            csv_name = os.path.dirname(
                __file__)+"\\csv\\"+dt_now.strftime('%Y%m%d%H%M%S')+"_自動帳票作成データ"+".csv"
            df1.to_csv(csv_name, index=False, encoding='CP932')
            with st.spinner(f'iReporter帳票作成中  \n \
                            処理時間 目安{df1.shape[0]*4+15}秒'):
                data_upload(csv_name, df1.shape[0])

    def qr_create(qr_df):

        # QRコード作成処理
        qr_list = []
        dt = datetime.now(timezone(timedelta(hours=9)))  # 日本時刻
        dt_now = dt.strftime('%Y/%m/%d')

        for index, row in qr_df.iterrows():
            if row['QRコードレイアウト区分'] == 1:
                # QRコード作成(旧レイアウトでの作成)
                code = pyqrcode.create(f"{row['ロット番号']},{row['月次']},{row['組立番号']},{row['機種']},{row['吋']},{row['G']},{
                                       row['完成部番']},{row['部品名']},'',{row['客先コード']},{row['ﾒｯｷ']}", error='L', version=5, mode='binary')
            else:
                code = pyqrcode.create(f"jp.co.cimtops.ireporter.openreport:repid={
                                       str(row['帳票発行ID'])}", error='L', version=5, mode='binary')
            code.png(os.path.dirname(
                __file__)+f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png", scale=3)
            # 余白を透明化
            img = Image.open(os.path.dirname(
                __file__)+f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png")
            img = img.convert("RGBA")
            datas = img.getdata()
            newData = []
            for item in datas:
                if item[0] == 255 and item[1] == 255 and item[2] == 255:
                    newData.append((255, 255, 255, 0))
                else:
                    newData.append(item)
            img.putdata(newData)
            img.save(os.path.dirname(__file__) +
                     f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png", "PNG")
            chuui = ''

            # 特殊条件追加 ADD_20241108 -111 -404 -102 を含む場合で、個数が複数ある場合は個数分出力
            # if ('-111' in row['完成部番'] or '-404' in row['完成部番'] or '-102' in row['完成部番']) and int(float(row['必要数'])) > 1: DEL_20250122_新規部品名検知のため、flgは2
            # ADD_20250122_新規部品名検知のため、flgは2
            if row['QRコードレイアウト区分'] == 2:
                if row['ﾒｯｷ'] != '' and int(float(row['ﾒｯｷ'])) == 1:
                    chuui = '無電解ニッケルメッキ'
                qr_list.append([dt_now, row['必要数'], row['完成部番'], row['B1図番'], row['B2図番'], row['F1図番'], row['F2図番'], row['L図番'], row['R図番'], row['C図番'], row['ロット番号'], row['月次'][4:6] +
                               "月"+row['月次'][6:]+"次", row['組立番号'], row['機種'], os.path.dirname(__file__)+f"\\template\\err.png", row['吋'], row['G'], row['工程'], chuui])

            elif ('-111' in row['完成部番'] or '-404' in row['完成部番'] or '-102' in row['完成部番']) and int(float(row['必要数'])) > 1:
                # 個数分のループ
                cnt = int(float(row['必要数']))
                for i in range(cnt):
                    chuui = "　　　　　　　　　　　　"+str(i+1)+'/'+str(cnt)
                    qr_list.append([dt_now, row['必要数'], row['完成部番'], row['B1図番'], row['B2図番'], row['F1図番'], row['F2図番'], row['L図番'], row['R図番'], row['C図番'], row['ロット番号'], row['月次'][4:6] +
                                   "月"+row['月次'][6:]+"次", row['組立番号'], row['機種'], os.path.dirname(__file__)+f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png", row['吋'], row['G'], row['工程'], chuui])
            else:
                if row['ﾒｯｷ'] != '' and int(float(row['ﾒｯｷ'])) == 1:
                    chuui = '無電解ニッケルメッキ'
                qr_list.append([dt_now, row['必要数'], row['完成部番'], row['B1図番'], row['B2図番'], row['F1図番'], row['F2図番'], row['L図番'], row['R図番'], row['C図番'], row['ロット番号'], row['月次'][4:6] +
                               "月"+row['月次'][6:]+"次", row['組立番号'], row['機種'], os.path.dirname(__file__)+f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png", row['吋'], row['G'], row['工程'], chuui])

        # アウトプットのエクセルの準備
        dt_now = datetime.now(timezone(timedelta(hours=9)))  # 日本時刻
        wb = os.path.dirname(__file__)+"\\work\\現品票_" + \
            dt_now.strftime('%Y%m%d%H%M%S') + '.xlsx'
        wb_name = "現品票_"+dt_now.strftime('%Y%m%d%H%M%S') + '.xlsx'
        if len(qr_list) < 10:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template5.xlsx', wb)
        elif len(qr_list) < 25:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template6.xlsx', wb)
        elif len(qr_list) < 50:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template7.xlsx', wb)
        elif len(qr_list) < 100:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template1.xlsx', wb)
        elif len(qr_list) < 500:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template2.xlsx', wb)
        elif len(qr_list) < 1000:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template3.xlsx', wb)
        elif len(qr_list) < 2000:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template4.xlsx', wb)
        else:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template.xlsx', wb)

        out_wb = openpyxl.load_workbook(wb)
        out_wb.active.title = "現品票"
        out_wb.save(wb)

        # 処理対象件数0の場合
        if len(qr_list) == 0:
            out_wb.active.cell(1, 30).value = "対象データなし"

        # y座標の定義
        y = 0

        # 要素格納ループ
        for i in range(len(qr_list)):
            # y座標
            y = i*18
            # 格納処理
            out_wb.active.cell(y+2, 10).value = qr_list[i][18]  # 無電解ニッケルﾒｯｷ
            out_wb.active.cell(y+3, 1).value = qr_list[i][0]  # 発行日
            out_wb.active.cell(y+4, 21).value = qr_list[i][1]  # 個数
            out_wb.active.cell(y+5, 7).value = qr_list[i][2]  # 完成部番
            out_wb.active.cell(y+8, 7).value = qr_list[i][3]  # B1部番
            out_wb.active.cell(y+9, 7).value = qr_list[i][4]  # B2部番
            out_wb.active.cell(y+10, 7).value = qr_list[i][5]  # F1部番
            out_wb.active.cell(y+11, 7).value = qr_list[i][6]  # F2部番
            out_wb.active.cell(y+12, 7).value = qr_list[i][7]  # L部番
            out_wb.active.cell(y+13, 7).value = qr_list[i][8]  # R部番
            out_wb.active.cell(y+14, 7).value = qr_list[i][9]  # C部番
            out_wb.active.cell(y+1, 30).value = qr_list[i][10]  # lot番号
            out_wb.active.cell(y+4, 24).value = qr_list[i][11]  # 月次
            out_wb.active.cell(y+7, 24).value = qr_list[i][12]  # 組立番号
            out_wb.active.cell(y+10, 24).value = qr_list[i][13]  # 機種名
            img_to_excel = openpyxl.drawing.image.Image(
                qr_list[i][14])  # QRコード
            out_wb.active.add_image(img_to_excel, f'AI{y+3}')
            out_wb.active.cell(y+13, 32).value = qr_list[i][15]  # インチ
            out_wb.active.cell(y+13, 38).value = qr_list[i][16]  # ゲージ
            out_wb.active.cell(y+15, 6).value = qr_list[i][17]  # 工程

        # 印刷範囲を指定
        out_wb.active.print_area = 'A1:AR'+str(int(y+17))

        # エクセルを保存
        out_wb.save(wb)
        out_wb.close()

        # ファイルを読み込む
        with open(wb, 'rb') as file:
            filedata = file.read()

        return (filedata, wb_name)

    # ADD_20250325_ズレる場合の対応追加
    def qr_createb(qr_df):

        # QRコード作成処理
        qr_list = []
        dt = datetime.now(timezone(timedelta(hours=9)))  # 日本時刻
        dt_now = dt.strftime('%Y/%m/%d')

        for index, row in qr_df.iterrows():
            if row['QRコードレイアウト区分'] == 1:
                # QRコード作成(旧レイアウトでの作成)
                code = pyqrcode.create(f"{row['ロット番号']},{row['月次']},{row['組立番号']},{row['機種']},{row['吋']},{row['G']},{
                                       row['完成部番']},{row['部品名']},'',{row['客先コード']},{row['ﾒｯｷ']}", error='L', version=5, mode='binary')
            else:
                code = pyqrcode.create(f"jp.co.cimtops.ireporter.openreport:repid={
                                       str(row['帳票発行ID'])}", error='L', version=5, mode='binary')
            code.png(os.path.dirname(
                __file__)+f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png", scale=3)
            # 余白を透明化
            img = Image.open(os.path.dirname(
                __file__)+f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png")
            img = img.convert("RGBA")
            datas = img.getdata()
            newData = []
            for item in datas:
                if item[0] == 255 and item[1] == 255 and item[2] == 255:
                    newData.append((255, 255, 255, 0))
                else:
                    newData.append(item)
            img.putdata(newData)
            img.save(os.path.dirname(__file__) +
                     f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png", "PNG")
            chuui = ''

            # 特殊条件追加 ADD_20241108 -111 -404 -102 を含む場合で、個数が複数ある場合は個数分出力
            # if ('-111' in row['完成部番'] or '-404' in row['完成部番'] or '-102' in row['完成部番']) and int(float(row['必要数'])) > 1: DEL_20250122_新規部品名検知のため、flgは2
            # ADD_20250122_新規部品名検知のため、flgは2
            if row['QRコードレイアウト区分'] == 2:
                if row['ﾒｯｷ'] != '' and int(float(row['ﾒｯｷ'])) == 1:
                    chuui = '無電解ニッケルメッキ'
                qr_list.append([dt_now, row['必要数'], row['完成部番'], row['B1図番'], row['B2図番'], row['F1図番'], row['F2図番'], row['L図番'], row['R図番'], row['C図番'], row['ロット番号'], row['月次'][4:6] +
                               "月"+row['月次'][6:]+"次", row['組立番号'], row['機種'], os.path.dirname(__file__)+f"\\template\\err.png", row['吋'], row['G'], row['工程'], chuui])

            elif ('-111' in row['完成部番'] or '-404' in row['完成部番'] or '-102' in row['完成部番']) and int(float(row['必要数'])) > 1:
                # 個数分のループ
                cnt = int(float(row['必要数']))
                for i in range(cnt):
                    chuui = "　　　　　　　　　　　　"+str(i+1)+'/'+str(cnt)
                    qr_list.append([dt_now, row['必要数'], row['完成部番'], row['B1図番'], row['B2図番'], row['F1図番'], row['F2図番'], row['L図番'], row['R図番'], row['C図番'], row['ロット番号'], row['月次'][4:6] +
                                   "月"+row['月次'][6:]+"次", row['組立番号'], row['機種'], os.path.dirname(__file__)+f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png", row['吋'], row['G'], row['工程'], chuui])
            else:
                if row['ﾒｯｷ'] != '' and int(float(row['ﾒｯｷ'])) == 1:
                    chuui = '無電解ニッケルメッキ'
                qr_list.append([dt_now, row['必要数'], row['完成部番'], row['B1図番'], row['B2図番'], row['F1図番'], row['F2図番'], row['L図番'], row['R図番'], row['C図番'], row['ロット番号'], row['月次'][4:6] +
                               "月"+row['月次'][6:]+"次", row['組立番号'], row['機種'], os.path.dirname(__file__)+f"\\work\\qrcode_{index}{dt.strftime('%H%M%S')}.png", row['吋'], row['G'], row['工程'], chuui])

        # アウトプットのエクセルの準備
        dt_now = datetime.now(timezone(timedelta(hours=9)))  # 日本時刻
        wb = os.path.dirname(__file__)+"\\work\\現品票_" + \
            dt_now.strftime('%Y%m%d%H%M%S') + '.xlsx'
        wb_name = "現品票B_"+dt_now.strftime('%Y%m%d%H%M%S') + '.xlsx'
        if len(qr_list) < 10:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template5b.xlsx', wb)
        elif len(qr_list) < 25:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template6b.xlsx', wb)
        elif len(qr_list) < 50:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template7b.xlsx', wb)
        elif len(qr_list) < 100:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template1b.xlsx', wb)
        elif len(qr_list) < 500:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template2b.xlsx', wb)
        elif len(qr_list) < 1000:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template3b.xlsx', wb)
        elif len(qr_list) < 2000:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\template4b.xlsx', wb)
        else:
            shutil.copy(os.path.dirname(__file__) +
                        '\\template\\templateb.xlsx', wb)

        out_wb = openpyxl.load_workbook(wb)
        out_wb.active.title = "現品票"
        out_wb.save(wb)

        # 処理対象件数0の場合
        if len(qr_list) == 0:
            out_wb.active.cell(1, 30).value = "対象データなし"

        # y座標の定義
        y = 0

        # 要素格納ループ
        for i in range(len(qr_list)):
            # y座標
            y = i*18
            # 格納処理
            out_wb.active.cell(y+2, 10).value = qr_list[i][18]  # 無電解ニッケルﾒｯｷ
            out_wb.active.cell(y+3, 1).value = qr_list[i][0]  # 発行日
            out_wb.active.cell(y+4, 21).value = qr_list[i][1]  # 個数
            out_wb.active.cell(y+5, 7).value = qr_list[i][2]  # 完成部番
            out_wb.active.cell(y+8, 7).value = qr_list[i][3]  # B1部番
            out_wb.active.cell(y+9, 7).value = qr_list[i][4]  # B2部番
            out_wb.active.cell(y+10, 7).value = qr_list[i][5]  # F1部番
            out_wb.active.cell(y+11, 7).value = qr_list[i][6]  # F2部番
            out_wb.active.cell(y+12, 7).value = qr_list[i][7]  # L部番
            out_wb.active.cell(y+13, 7).value = qr_list[i][8]  # R部番
            out_wb.active.cell(y+14, 7).value = qr_list[i][9]  # C部番
            out_wb.active.cell(y+1, 30).value = qr_list[i][10]  # lot番号
            out_wb.active.cell(y+4, 24).value = qr_list[i][11]  # 月次
            out_wb.active.cell(y+7, 24).value = qr_list[i][12]  # 組立番号
            out_wb.active.cell(y+10, 24).value = qr_list[i][13]  # 機種名
            img_to_excel = openpyxl.drawing.image.Image(
                qr_list[i][14])  # QRコード
            out_wb.active.add_image(img_to_excel, f'AI{y+3}')
            out_wb.active.cell(y+13, 32).value = qr_list[i][15]  # インチ
            out_wb.active.cell(y+13, 38).value = qr_list[i][16]  # ゲージ
            out_wb.active.cell(y+15, 6).value = qr_list[i][17]  # 工程

        # 印刷範囲を指定
        out_wb.active.print_area = 'A1:AR'+str(int(y+17))

        # エクセルを保存
        out_wb.save(wb)
        out_wb.close()

        # ファイルを読み込む
        with open(wb, 'rb') as file:
            filedata = file.read()

        return (filedata, wb_name)

    # ファイル削除処理関数を定義
    def cleanup_old_files():
        # 対象ディレクトリを指定
        directory = os.path.dirname(__file__)+"\\work\\"
        now = datetime.now()  # 現在の時刻を取得
        minutes20 = now - timedelta(minutes=20)  # 20分前

        # ディレクトリ内の全ファイルを確認
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            
            # ファイルであり、拡張子が.pngまたは.xlsxの場合に処理を進める
            if os.path.isfile(file_path) and (filename.lower().endswith('.png') or filename.lower().endswith('.xlsx')):
                # ファイルの最終更新時間を取得
                file_mtime = os.path.getmtime(file_path)
                file_mtime_dt = datetime.fromtimestamp(file_mtime)

                # 20分以上前に更新されたファイルを削除
                if file_mtime_dt < minutes20:
                    try:
                        os.remove(file_path)
                        print(f"古いファイルを削除: {file_path}")
                    except Exception as e:
                        print(f"ファイル削除エラー: {e}")

    def main():

        st.set_page_config(
            page_title='現品票発行',
            layout="wide", page_icon='move.gif'
        )

        HIDE_ST_STYLE = """
                    <style>
                    div[data-testid="stToolbar"] {
                    visibility: hidden;
                    height: 0%;
                    position: fixed;
                    }
                    div[data-testid="stDecoration"] {
                    visibility: hidden;
                    height: 0%;
                    position: fixed;
                    }
                    #MainMenu {
                    visibility: hidden;
                    height: 0%;
                    }
                    header {
                    visibility: hidden;
                    height: 0%;
                    }
                    footer {
                    visibility: hidden;
                    height: 0%;
                    }
                    .appview-container .main .block-container{
                                padding-top: 1rem;
                                padding-right: 3rem;
                                padding-left: 3rem;
                                padding-bottom: 1rem;
                            }  
                            .reportview-container {
                                padding-top: 0rem;
                                padding-right: 3rem;
                                padding-left: 3rem;
                                padding-bottom: 0rem;
                            }
                            header[data-testid="stHeader"] {
                                z-index: -1;
                            }
                            div[data-testid="stToolbar"] {
                            z-index: 100;
                            }
                            div[data-testid="stDecoration"] {
                            z-index: 100;
                            }
                    .block-container {
                            padding-top: 0rem !important;
                            padding-bottom: 0rem !important;
                            }        
                    </style>
    """
        st.markdown(HIDE_ST_STYLE, unsafe_allow_html=True)

        # 年月入力欄（年と月であればプルダウンでもよいかもしれない当年前年翌年、１～１２月）
        dt_now = datetime.now(timezone(timedelta(hours=9))) + \
            timedelta(days=32)  # 日本時刻+32日
        dt_nen = int(dt_now.strftime('%Y'))
        dt_tsuki = int(dt_now.strftime('%m'))

        col1, col2, col3, col4, col5, col6 = st.columns([3, 3, 3, 3, 1, 3])
        with col1:
            st.markdown("### 現品票発行")
        with col2:
            nen = st.selectbox('生産計画年', [dt_nen-1, dt_nen, dt_nen+1], index=1)
        with col3:
            getsu = st.selectbox('生産計画月', [
                                 '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12'], index=dt_tsuki-1)
        with col4:
            ji = st.selectbox(
                '生産計画次', ['', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '_'], index=0)
            # 必要なデータを一通りそろえる　※df_updateに渡すためにここでデータ処理を行う
            # キャッシュリソース化したら再レンダリング問題が解決
            df1 = df_set1(nen, getsu)
            df2 = df_set2()
            sql = f"select rep_top_id,cluster_1_2_t from view_report_406"  # 横レイアウト
            df3 = ireporter_data_get(sql)
            sql = f"select rep_top_id,cluster_1_2_t from view_report_405"  # 縦レイアウト
            df4 = ireporter_data_get(sql)
            df5 = pd.concat([df3, df4], ignore_index=True)
            df = pd.merge(df1, df2, on=['部品名'], how='left')
            df = pd.merge(df, df5, left_on=['ロット番号'], right_on=[
                          'cluster_1_2_t'], how='left')

            df6 = df_set3()  # 削除フラグの追加　ADD_20241124
            # 削除フラグの追加　ADD_20241124
            df = pd.merge(df, df6, on=['ロット番号'], how='left')
            df = df[df['削除flg'] == 0]  # 削除フラグの追加　ADD_20241124

            df = df.rename(columns={'rep_top_id': '帳票発行ID'})
            df = df.drop(columns=['cluster_1_2_t'])
            df['帳票発行ID'] = df["帳票発行ID"].fillna(0)  # Noneデータ対
            df['工程VERSION'] = df["工程VERSION"].replace(
                "", 0).fillna(0)  # ADD_20241217
            # df['工程VERSION'] = df["工程VERSION"].fillna(0) #Noneデータ対
            df['工程VERSION'] = df['工程VERSION'].astype(float).astype(int)
            df = df.fillna('')  # Noneデータ対

            # 再発行している場合は最大のものを表示する
            idx = df.groupby('ロット番号')['帳票発行ID'].idxmax()
            df = df.loc[idx]

        with col6:
            if st.button("データ更新"):
                db_update()
                irepo_jidoutyouhyou_create(df)
                st.cache_resource.clear()

        # 抽出条件の整備
        # 生産開始次
        dcolumns = df.columns
        mdf1 = pd.DataFrame(columns=dcolumns)
        mdf2 = pd.DataFrame(columns=dcolumns)
        mdf3 = pd.DataFrame(columns=dcolumns)
        mdf4 = pd.DataFrame(columns=dcolumns)
        mdf5 = pd.DataFrame(columns=dcolumns)
        mdf6 = pd.DataFrame(columns=dcolumns)
        mdf7 = pd.DataFrame(columns=dcolumns)
        if ji != '':
            mdf1 = df[df['月次'] == str(nen)+getsu+ji]
        else:
            mdf1 = df

        with st.expander('抽出条件'):
            col1, col2, col3, col4, col5, col6 = st.columns(
                [2, 2.5, 3, 1, 1, 2.8])
            with col1:
                # 組立番号
                kumi_list = list(set(df['組立番号'].tolist()))
                kumi_list.sort()
                kumi_list.insert(0, "")
                kumitate = st.selectbox('組立番号', kumi_list, index=0)
                if kumitate != '':
                    mdf2 = df[df['組立番号'] == kumitate]
                else:
                    mdf2 = df

            with col2:
                # 機種名
                kisyu_list = list(set(df['機種'].tolist()))
                kisyu_list.sort()
                kisyu_list.insert(0, "")
                kisyumei = st.selectbox('機種', kisyu_list, index=0)
                if kisyumei != '':
                    mdf3 = df[df['機種'] == kisyumei]
                else:
                    mdf3 = df

            with col3:
                # 部品名
                buhin_list = list(set(df['部品名'].tolist()))
                buhin_list.sort()
                buhin_list.insert(0, "")
                buhinmei = st.selectbox('部品名', buhin_list, index=0)
                if buhinmei != '':
                    mdf4 = df[df['部品名'] == buhinmei]
                else:
                    mdf4 = df

            with col4:
                # 吋
                inchi_list = list(set(df['吋'].tolist()))
                inchi_list.sort()
                inchi_list.insert(0, "")
                inchimei = st.selectbox('吋', inchi_list, index=0)
                if inchimei != '':
                    mdf5 = df[df['吋'] == inchimei]
                else:
                    mdf5 = df

            with col5:
                # ｹﾞｰｼﾞ
                G_list = list(set(df['G'].tolist()))
                G_list.sort()
                G_list.insert(0, "")
                # Gmei = st.selectbox('G', G_list, index=0) DEL_20250106_Gはｹﾞｰｼﾞに名称変更
                Gmei = st.selectbox('ｹﾞｰｼﾞ', G_list, index=0)
                if Gmei != '':
                    mdf6 = df[df['G'] == Gmei]
                else:
                    mdf6 = df

            with col6:
                # 客先
                C_list = list(set(df['客先名'].tolist()))
                C_list.sort()
                C_list.insert(0, "")
                Cmei = st.selectbox('客先名', C_list, index=0)
                if Cmei != '':
                    mdf7 = df[df['客先名'] == Cmei]
                else:
                    mdf7 = df

            cola1, cola2, cola3, cola4, cola5, cola6, cola7 = st.columns(
                [3, 2, 3, 2, 3, 2, 1])

            with cola1:
                # ロット番号
                lotno = st.text_input("ロット番号", '')
            with cola2:
                st.write("")
                syubetsu1 = st.radio(
                    "検索1", ("前方一致", "部分一致"), horizontal=True, label_visibility="collapsed")
                # ロット番号
                if syubetsu1 == "前方一致":
                    mdf8 = df[df['ロット番号'].str.startswith(lotno)]
                elif syubetsu1 == "部分一致":
                    mdf8 = df[df['ロット番号'].str.contains(lotno)]
            with cola3:
                # 完成部番
                buban = st.text_input("完成部番", '')
            with cola4:
                st.write("")
                syubetsu2 = st.radio(
                    "検索2", ("前方一致", "部分一致"), horizontal=True, index=1, label_visibility="collapsed")
                # 完成部番
                if syubetsu2 == "前方一致":
                    mdf9 = df[df['完成部番'].str.startswith(buban)]
                elif syubetsu2 == "部分一致":
                    mdf9 = df[df['完成部番'].str.contains(buban)]
            with cola5:
                # 完成部番
                l_zuban = st.text_input("L図番", '')
            with cola6:
                st.write("")
                syubetsu3 = st.radio(
                    "検索3", ("前方一致", "部分一致"), horizontal=True, label_visibility="collapsed")
                # 完成部番
                if syubetsu3 == "前方一致":
                    mdf10 = df[df['L図番'].str.startswith(l_zuban)]
                elif syubetsu3 == "部分一致":
                    mdf10 = df[df['L図番'].str.contains(l_zuban)]
            with cola7:
                st.write("")
                st.write("")
                st.write("")
                ken = st.empty()

            common_df = mdf1.copy()
            # 2番目から7番目までのデータフレームと共通行を抽出
            # for mdf in [mdf2, mdf3, mdf4, mdf5, mdf6, mdf7]:
            for mdf in [mdf2, mdf3, mdf4, mdf5, mdf6, mdf7, mdf8, mdf9, mdf10]:
                common_df = pd.merge(common_df, mdf, how='inner')

        df = common_df.copy()
        ken.write(str(len(df))+"件")

        # ボタン状態を管理
        if "select_all" not in st.session_state:
            st.session_state["select_all"] = False
        if "deselect_all" not in st.session_state:
            # st.session_state["deselect_all"] = False
            st.session_state["deselect_all"] = True
        if st.session_state["deselect_all"]:
            # ボタンイベントの処理
            if st.button("全選択"):
                st.session_state["select_all"] = True
                st.session_state["deselect_all"] = False
        else:
            if st.button("全解除"):
                st.session_state["select_all"] = False
                st.session_state["deselect_all"] = True

        # if df['帳票発行ID'] != '':
        #     df['帳票発行ID']=df['帳票発行ID'].astype(int)

        # 20250106_Gをｹﾞｰｼﾞに変更対応　表示の変更
        df = df.rename(columns={'G': 'ｹﾞｰｼﾞ'})

        # AgGridの設定
        gd = GridOptionsBuilder.from_dataframe(df)
        gd.configure_selection(selection_mode='multiple', use_checkbox=True)

        # #フィルタリング設定 自由度が低いため削除
        # gd.configure_column('ロット番号',filter='agSetColumnFilter')
        # gd.configure_column('月次',filter='agSetColumnFilter')
        # gd.configure_column('組立番号',filter='agSetColumnFilter')
        # gd.configure_column('機種',filter='agSetColumnFilter')
        # gd.configure_column('吋',filter='agSetColumnFilter')
        # gd.configure_column('G',filter='agSetColumnFilter')
        # gd.configure_column('完成部番',filter='agTextColumnFilter')
        # gd.configure_column('部品名',filter='agSetColumnFilter')
        # gd.configure_column('客先名',filter='agSetColumnFilter')
        # gd.configure_column('国名',filter='agSetColumnFilter')
        # gd.configure_column('B1図番',filter='agTextColumnFilter')
        # gd.configure_column('B2図番',filter='agTextColumnFilter')
        # gd.configure_column('F1図番',filter='agTextColumnFilter')
        # gd.configure_column('F2図番',filter='agTextColumnFilter')
        # gd.configure_column('L図番',filter='agTextColumnFilter')
        # gd.configure_column('R図番',filter='agTextColumnFilter')
        # gd.configure_column('C図番',filter='agTextColumnFilter')
        # gd.configure_column('組立開始日',filter='agTextColumnFilter')
        # gd.configure_column('梱包開始日',filter='agTextColumnFilter')
        # gd.configure_column('指定納期',filter='agTextColumnFilter')
        # gd.configure_column('必要数',filter='agSetColumnFilter')
        # gd.configure_column('SKDK',filter='agSetColumnFilter')
        # gd.configure_column('工程',filter='agSetColumnFilter')
        # gd.configure_column('工程VERSION',filter='agSetColumnFilter')
        # gd.configure_column('ﾒｯｷ',filter='agSetColumnFilter')
        # gd.configure_column('客先コード',filter='agSetColumnFilter')
        # gd.configure_column('ireporter管理flg',filter='agSetColumnFilter')
        # gd.configure_column('QRコードレイアウト区分',filter='agSetColumnFilter')
        # gd.configure_column('縦横区分',filter='agSetColumnFilter')
        # gd.configure_column('帳票発行ID',filter='agSetColumnFilter')

        gd.configure_default_column(filter=False)
        gd.configure_column("ﾒｯｷ", hide=True)
        gridoptions = gd.build()

        # localeTextのカスタマイズ
        gridoptions['localeText'] = {
            'pinColumn': 'ピン止め',
            'noPin': 'ピン止め解除',
            'pinLeft': '左ピン止め',
            'pinRight': '右ピン止め',
            'autosizeThiscolumn': '列幅自動調整',  # ←公式ドキュメントが違う
            'autosizeAllColumns': '全ての列幅自動調整',
            'resetColumns': '設定をリセット',
            "contains": "含む",
            "notContains": "含まない",
            "equals": "等しい",
            "notEqual": "等しくない",
            "startsWith": "で始まる",
            "endsWith": "で終わる",
            "lessThan": "より小さい",
            "lessThanOrEqual": "以下",
            "greaterThan": "より大きい",
            "greaterThanOrEqual": "以上",
            "inRange": "範囲内",
            "blank": "空欄",
            "notBlank": "空欄でない",
            "before": "以前",
            "after": "以後",
            'value': '値',
            'noRowsToShow': '表示する行がありません',
            'page': 'ページ',
            'of': 'の',
            'cut': '切り取り',
            'copy': 'コピー',
            'copyWithHeaders': 'ヘッダー付コピー',
            'copyWithGroupHeaders': 'グループヘッダー付コピー',
            'paste': '貼り付け',
            'export': 'ファイル出力',
            'csvExport': 'csvファイル出力',
            'excelExport': 'Excelファイル出力',
            'selectAll': '(全て選択)',
            'blanks': '(空欄)'
        }

        # ADD_20241119_全選択、全解除の追加
        # # カスタム JavaScript コード
        # select_all_code = JsCode("""
        # function selectAllRows(e) {
        #     e.api.selectAll();  // すべての行を選択
        # }
        # """)

        # deselect_all_code = JsCode("""
        # function deselectAllRows(e) {
        #     e.api.deselectAll();  // すべての行の選択を解除
        # }
        # """)
        # カスタム JavaScript コード

        select_all_code = JsCode("""
        function selectAllRows(api) {
            api.api.selectAll();  // すべての行を選択
        }
        """)

        deselect_all_code = JsCode("""
        function deselectAllRows(api) {
            api.api.deselectAll();  // すべての行の選択を解除
        }
        """)

        # JavaScriptの適用
        if st.session_state["select_all"]:
            gridoptions["onFirstDataRendered"] = select_all_code
        elif st.session_state["deselect_all"]:
            gridoptions["onFirstDataRendered"] = deselect_all_code

        # AgGridをStreamlitに表示
        grid_table = AgGrid(df, height=470, gridOptions=gridoptions, update_mode=GridUpdateMode.SELECTION_CHANGED |
                            GridUpdateMode.MODEL_CHANGED, enable_enterprise_modules=False, allow_unsafe_jscode=True, reload_data=True)
        # enable_enterprise_modules=Falseを入れないと、デフォルトで試用

        #pcol1, pcol2, pcol3 = st.columns([1, 1, 2])
        pcol1, pcol2, pcol3, pcol4 = st.columns([10, 12, 10,8])
        with pcol1:
            # 初回実行時のキャッシュクリア処理を追加（タイミングによって、正しく帳票IDの更新情報が表示されないケースがあるため）ADD_20241127
            if "initialized" not in st.session_state:
                st.session_state["initialized"] = True
                st.cache_resource.clear()
                db_update()
                st.write("キャッシュクリア中……")
                # st.experimental_rerun()
                st.rerun()
                st.cache_resource.clear()
            else:
                if st.button("チェック済の印刷"):
                    # grid_table['selected_rows'] で選択された行を取得する
                    selected_rows = grid_table['selected_rows']

                    with st.spinner('現品票作成処理　実行中'):
                        if selected_rows is not None and len(selected_rows) > 0:

                            # 20250106_Gをｹﾞｰｼﾞに変更対応　表示を元に戻す
                            selected_rows = selected_rows.rename(columns={'ｹﾞｰｼﾞ': 'G'})

                            df_csv_cnv(selected_rows, "履歴")

                            check_df = selected_rows[(selected_rows['ireporter管理flg'] == 1) & (
                                selected_rows['QRコードレイアウト区分'] == 0) & (selected_rows['帳票発行ID'] == 0)]
                            err_exists = check_df.empty
                            if err_exists:
                                selected_rows = selected_rows.sort_values(
                                    ['月次', '組立番号', 'ロット番号'])
                                filedata, wb_name = qr_create(selected_rows)

                                st.markdown('処理終了：処理結果のエクセルを以下からダウンロードしてください')
                                st.download_button(
                                    label='処理結果ダウンロード',
                                    data=filedata,
                                    file_name=wb_name,
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                )
                            else:
                                st.markdown(
                                    'ireporter管理flgが1かつ、帳票発行IDが未発行のデータがあります（カムリング、ダイヤルキャップは除く）。右上の「データ更新」を行ってください。')
                                st.dataframe(check_df, hide_index=True)
                            st.cache_resource.clear()
                        else:
                            st.markdown('データ無し')
                            st.cache_resource.clear()
                st.write("　")
                st.write("　")
                st.write("　")

        with pcol2:
            # ADD_20250325_start ズレる場合の処理追加
            if st.button("チェック済の印刷（ずれる場合）"):
                # grid_table['selected_rows'] で選択された行を取得する
                selected_rows = grid_table['selected_rows']

                with st.spinner('現品票作成処理　実行中'):
                    if selected_rows is not None and len(selected_rows) > 0:

                        # 20250106_Gをｹﾞｰｼﾞに変更対応　表示を元に戻す
                        selected_rows = selected_rows.rename(columns={'ｹﾞｰｼﾞ': 'G'})
                        
                        df_csv_cnv(selected_rows, "履歴")

                        check_df = selected_rows[(selected_rows['ireporter管理flg'] == 1) & (
                            selected_rows['QRコードレイアウト区分'] == 0) & (selected_rows['帳票発行ID'] == 0)]
                        err_exists = check_df.empty
                        if err_exists:
                            selected_rows = selected_rows.sort_values(
                                ['月次', '組立番号', 'ロット番号'])
                            filedatab, wb_nameb = qr_createb(selected_rows)

                            st.markdown('処理終了：処理結果のエクセルを以下からダウンロードしてください')
                            st.download_button(
                                label='処理結果ダウンロード',
                                data=filedatab,
                                file_name=wb_nameb,
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
                        else:
                            st.markdown(
                                'ireporter管理flgが1かつ、帳票発行IDが未発行のデータがあります（カムリング、ダイヤルキャップは除く）。右上の「データ更新」を行ってください。')
                            st.dataframe(check_df, hide_index=True)
                        st.cache_resource.clear()
                    else:
                        st.markdown('データ無し')
                        st.cache_resource.clear()
                # ADD_20250325_end
        with pcol3:
            st.write("　　")
        with pcol4:
            #ADD_Start_20250410
            selected_rows = grid_table['selected_rows']
            
            # CSV出力ボタン - データが選択されているかに関わらず表示
            if not selected_rows is None:

                selected_rows = selected_rows.rename(columns={'ｹﾞｰｼﾞ': 'G'})
                #df_csv_cnv(selected_rows, "csv履歴")
                    
                csv = selected_rows.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    label="CSVダウンロード",
                    data=csv,
                    file_name=f'現品票リスト_{datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d%H%M%S")}.csv',
                    mime='text/csv',
                )
            #ADD_End_20250410

        # ファイル削除処理を実行
        cleanup_old_files()
        # time.sleep(300)

    if __name__ == "__main__":
        main()

except Exception as e:
    # エラーをSQLite genpinhyo.db の error_log テーブルに記録
    print(e)
    dt_now = datetime.now(timezone(timedelta(hours=9)))  # 日本時刻
    db_path = os.path.join(os.path.dirname(__file__), 'Database', 'genpinhyo.db')

    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # error_log テーブルを作成（存在しない場合）
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS error_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                error_message TEXT NOT NULL
            )
        ''')

        # エラーメッセージを保存
        cursor.execute('''
            INSERT INTO error_log (created_at, error_message)
            VALUES (?, ?)
        ''', (dt_now.strftime('%Y-%m-%d %H:%M:%S'), str(e)))

        conn.commit()
        conn.close()
    except Exception as db_error:
        # SQLite保存に失敗した場合は従来のファイル出力にフォールバック
        print(f"DB保存失敗: {db_error}")
        with open(os.path.dirname(__file__)+"\\"+dt_now.strftime('%Y%m%d%H%M%S')+"_err"+".txt", mode='w') as f:
            f.write(str(e))

    # # 【旧版：TXT出力】コメントアウト（2025-11-22 SQLite移行）
    # with open(os.path.dirname(__file__)+"\\"+dt_now.strftime('%Y%m%d%H%M%S')+"_err"+".txt", mode='w') as f:
    #     f.write(str(e))
